import glob
import io
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
# FILE PATHS
# ══════════════════════════════════════════════════════════════════════════════
OUTPUT_FILE  = '/mnt/user-data/outputs/AVS_MidWeek_Labor_Report.xlsx'
REPORT_DATES = input("Enter the report date range (e.g. 3.6.26 - 3.9.26): ").strip()

# ── Auto-detect ADP file ──────────────────────────────────────────────────────
def find_file(folder, pattern, label):
    matches = glob.glob(os.path.join(folder, pattern))
    if not matches:
        raise FileNotFoundError(
            f"No {label} file found in {folder} matching '{pattern}'. "
            "Please upload the file and try again."
        )
    if len(matches) > 1:
        matches = sorted(matches, key=os.path.getmtime, reverse=True)
        print(f"Warning: Multiple {label} files found — using most recent: {os.path.basename(matches[0])}")
    else:
        print(f"Found {label}: {os.path.basename(matches[0])}")
    return matches[0]

UPLOADS  = '/mnt/user-data/uploads'
ADP_FILE = find_file(UPLOADS, 'adp_payroll*.csv', 'ADP Payroll CSV')

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Clean ADP export (fix OH store delineation)
# ══════════════════════════════════════════════════════════════════════════════
# WHY THIS IS NEEDED (do not remove):
#   The ADP CSV export is a third-party file we cannot control. Some Ohio store
#   names are exported with an unquoted comma — e.g. "112-0019 Streetsboro, OH"
#   — which causes pandas to read 10 fields instead of the expected 9, throwing
#   a ParserError. This pre-processing step fixes those rows in memory before
#   pandas ever sees the file. No data is lost or altered; the store name is
#   simply re-quoted so the CSV is valid. This will be required on every run
#   until ADP changes their export format.
with open(ADP_FILE, 'r', encoding='utf-8-sig') as f:
    lines = f.readlines()
fixed_lines = []
for line in lines:
    fields = line.rstrip('\r\n').split(',')
    if len(fields) == 10:
        # Columns 3 & 4 are the two halves of a store name split on its comma.
        # Re-join them into one quoted field so the row is 9 fields again.
        merged = '"' + fields[3] + ',' + fields[4] + '"'
        fields = fields[:3] + [merged] + fields[5:]
    fixed_lines.append(','.join(fields) + '\n')

raw_payroll = pd.read_csv(io.StringIO(''.join(fixed_lines)), dtype=str)
raw_payroll['Batch Description'] = raw_payroll['Batch Description'].str.replace(
    r'(112-\d{4}\s+\w[\w\s]*?)\s+OH\b', lambda m: m.group(1) + ' (OH)', regex=True
)
for col in ['Temp Rate', 'Reg Hours', 'O/T Hours']:
    raw_payroll[col] = pd.to_numeric(raw_payroll[col], errors='coerce').fillna(0.0)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Reference tables (edit directly when DMs, bands, or goals change)
# ══════════════════════════════════════════════════════════════════════════════

# ── DM Mapping (update when DM assignments change) ───────────────────────────
dm = pd.DataFrame([
    {'Store #': '112-0019', 'Store Name': 'Streetsboro (OH)',              'DM': 'Alex'},
    {'Store #': '112-0020', 'Store Name': 'Strongsville (OH)',             'DM': 'Alex'},
    {'Store #': '112-0022', 'Store Name': 'Avon (OH)',                     'DM': 'Alex'},
    {'Store #': '112-0023', 'Store Name': 'Lorain (OH)',                   'DM': 'Alex'},
    {'Store #': '112-0026', 'Store Name': 'Fairfield (OH)',                'DM': 'Alex'},
    {'Store #': '112-0027', 'Store Name': 'Florence (KY)',                 'DM': 'Alex'},
    {'Store #': '112-0015', 'Store Name': 'Centerville (OH)',              'DM': 'Cody'},
    {'Store #': '112-0017', 'Store Name': 'West Chester (OH)',             'DM': 'Cody'},
    {'Store #': '112-0024', 'Store Name': 'Beavercreek (OH)',              'DM': 'Cody'},
    {'Store #': '112-0028', 'Store Name': 'Englewood - S Main (OH)',       'DM': 'Cody'},
    {'Store #': '112-0029', 'Store Name': 'Grove City (OH)',               'DM': 'Cody'},
    {'Store #': '112-0030', 'Store Name': 'Hilliard (OH)',                 'DM': 'Cody'},
    {'Store #': '112-0002', 'Store Name': 'Muskogee (OK)',                 'DM': 'Dottie'},
    {'Store #': '112-0031', 'Store Name': 'McAlester (OK)',                'DM': 'Dottie'},
    {'Store #': '112-0035', 'Store Name': 'Tulsa - 51st (OK)',             'DM': 'Dottie'},
    {'Store #': '112-0038', 'Store Name': 'Ada (OK)',                      'DM': 'Dottie'},
    {'Store #': '112-0001', 'Store Name': 'Stillwater (OK)',               'DM': 'Manny'},
    {'Store #': '112-0003', 'Store Name': 'Lawton (OK)',                   'DM': 'Manny'},
    {'Store #': '112-0004', 'Store Name': 'Ardmore (OK)',                  'DM': 'Manny'},
    {'Store #': '112-0005', 'Store Name': 'Durant (OK)',                   'DM': 'Manny'},
    {'Store #': '112-0008', 'Store Name': 'Lufkin (TX)',                   'DM': 'Miriam'},
    {'Store #': '112-0009', 'Store Name': 'Nacogdoches (TX)',              'DM': 'Miriam'},
    {'Store #': '112-0011', 'Store Name': 'Princeton (TX)',                'DM': 'Miriam'},
    {'Store #': '112-0032', 'Store Name': 'Plano (TX)',                    'DM': 'Miriam'},
    {'Store #': '112-0034', 'Store Name': 'Marshall (TX)',                 'DM': 'Miriam'},
    {'Store #': '112-0037', 'Store Name': 'McKinney - Hardin Blvd (TX)',   'DM': 'Miriam'},
    {'Store #': '112-0012', 'Store Name': 'Amelia (OH)',                   'DM': 'Renee'},
    {'Store #': '112-0013', 'Store Name': 'Milford (OH)',                  'DM': 'Renee'},
    {'Store #': '112-0014', 'Store Name': 'Cincinnati - Winton Rd (OH)',   'DM': 'Renee'},
    {'Store #': '112-0016', 'Store Name': 'Cincinnati - Red Bank Rd (OH)', 'DM': 'Renee'},
    {'Store #': '112-0018', 'Store Name': 'Harrison (OH)',                 'DM': 'Renee'},
    {'Store #': '112-0021', 'Store Name': 'Loveland (OH)',                 'DM': 'Renee'},
    {'Store #': '112-0006', 'Store Name': 'Wichita Falls (TX)',            'DM': 'Steve'},
    {'Store #': '112-0007', 'Store Name': 'Gainesville (TX)',              'DM': 'Steve'},
    {'Store #': '112-0010', 'Store Name': 'Little Elm (TX)',               'DM': 'Steve'},
    {'Store #': '112-0033', 'Store Name': 'Brownwood (TX)',                'DM': 'Steve'},
    {'Store #': '112-0036', 'Store Name': 'Denton (TX)',                   'DM': 'Steve'},
])

# ── Revenue Band Hourly Goals (update when goals change) ─────────────────────
# NOTE: These are FULL weekly goals. % Used is calculated as Thu-Sun hours / full goal.
rev_bands = {
    '<25k':         422,
    '25k-30k':      420,
    '30k-35k':      504,
    '35k-40k':      588,
    '40k-45k':      605,
    '45k-50k':      849,
    '50k+':         756,
    'NRO Seasoned': 895,
    'NRO':          979,
}

# ── Store Revenue Band Assignments (update weekly as needed) ─────────────────
store_rev_bands = {
    '112-0001': '35k-40k', '112-0002': '35k-40k', '112-0003': '<25k',
    '112-0004': '<25k',    '112-0005': '<25k',    '112-0006': '30k-35k',
    '112-0007': '<25k',    '112-0008': '30k-35k', '112-0009': '25k-30k',
    '112-0010': '<25k',    '112-0011': '<25k',    '112-0012': '<25k',
    '112-0013': '<25k',    '112-0014': '<25k',    '112-0015': '45k-50k',
    '112-0016': '<25k',    '112-0017': '30k-35k', '112-0018': '30k-35k',
    '112-0019': '<25k',    '112-0020': '<25k',    '112-0021': '30k-35k',
    '112-0022': '<25k',    '112-0023': '<25k',    '112-0024': '25k-30k',
    '112-0026': '<25k',    '112-0027': '<25k',    '112-0028': '30k-35k',
    '112-0029': '<25k',    '112-0030': '<25k',    '112-0031': '<25k',
    '112-0032': '<25k',    '112-0033': '<25k',    '112-0034': '<25k',
    '112-0035': 'NRO Seasoned', '112-0036': '35k-40k', '112-0037': '35k-40k',
    '112-0038': 'NRO',
}

# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Aggregate Thu-Sun payroll hours (exclude 112-9001)
# ══════════════════════════════════════════════════════════════════════════════
pf = raw_payroll[~raw_payroll['Batch Description'].str.contains('112-9001', na=False)].copy()
pf['store_num']   = pf['Batch Description'].str.extract(r'(112-\d{4})')
pf['total_hours'] = pf['Reg Hours'] + pf['O/T Hours']
pagg = pf.groupby('store_num').agg(
    actual_hours=('total_hours', 'sum')
).reset_index()
pagg['actual_hours'] = pagg['actual_hours'].round(2)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — Merge and calculate
# ══════════════════════════════════════════════════════════════════════════════
df = dm.merge(pagg[['store_num', 'actual_hours']], left_on='Store #', right_on='store_num', how='left')
df['actual_hours'] = df['actual_hours'].fillna(0.0)
df['Rev Band']     = df['Store #'].map(store_rev_bands)
df['Weekly Goal']  = df['Rev Band'].map(rev_bands)
df['Variance']     = (df['actual_hours'] - df['Weekly Goal']).round(2)
df['Pct Used']     = (df['actual_hours'] / df['Weekly Goal']).where(df['Weekly Goal'] > 0, 0.0)
df = df.sort_values(['DM', 'Store #']).reset_index(drop=True)

# ══════════════════════════════════════════════════════════════════════════════
# STYLES
# ══════════════════════════════════════════════════════════════════════════════
DARK_NAVY = '1F3864'
MID_BLUE  = '2E5FA3'
OVER_RED  = 'FFDADA'
UNDER_GRN = 'D6F0DA'
NEUTRAL   = 'F2F2F2'
SUBTOTAL  = 'E2E8F0'
WHITE     = 'FFFFFF'

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def subtotal_border():
    thick = Side(style='medium', color='2E5FA3')
    thin  = Side(style='thin',   color='CCCCCC')
    return Border(left=thin, right=thin, top=thick, bottom=thin)

center = Alignment(horizontal='center', vertical='center')
left   = Alignment(horizontal='left',   vertical='center', indent=1)

FMT_HOURS    = '#,##0.00'
FMT_VARIANCE = '+#,##0.00;[Red]-#,##0.00;"-"'
FMT_PCT      = '0.0%'

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — Mid-Week Labor Pulse
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = 'Mid-Week Labor Pulse'
ws.sheet_view.showGridLines = False

headers    = ['Store #', 'Store Name', 'DM', 'Weekly Hourly Goal',
              'Thu–Sun Actual Hours', 'Variance (Hrs)', '% of Weekly Goal Used']
col_widths = [12, 34, 10, 20, 22, 18, 24]

# Title row
ws.row_dimensions[1].height = 36
ws.merge_cells('A1:G1')
c = ws['A1']
c.value     = f'AvS Mid-Week Labor Pulse — Thu–Sun Hours — {REPORT_DATES}'
c.font      = Font(name='Arial', size=16, bold=True, color=WHITE)
c.fill      = fill(DARK_NAVY)
c.alignment = Alignment(horizontal='center', vertical='center')

# Header row
ws.row_dimensions[2].height = 24
for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
    c = ws.cell(row=2, column=ci, value=hdr)
    c.font      = Font(name='Arial', size=11, bold=True, color=WHITE)
    c.fill      = fill(MID_BLUE)
    c.alignment = center
    c.border    = thin_border()
    ws.column_dimensions[get_column_letter(ci)].width = w

row_num        = 3
grand_goal     = 0.0
grand_hours    = 0.0
grand_variance = 0.0

for dm_name, group in df.groupby('DM', sort=True):
    dm_goal     = 0.0
    dm_hours    = 0.0
    dm_variance = 0.0

    for _, row in group.iterrows():
        variance  = float(row['Variance'])    if pd.notna(row['Variance'])    else 0.0
        hours     = float(row['actual_hours']) if pd.notna(row['actual_hours']) else 0.0
        goal_v    = float(row['Weekly Goal'])  if pd.notna(row['Weekly Goal'])  else 0.0
        pct_used  = float(row['Pct Used'])     if pd.notna(row['Pct Used'])     else 0.0

        dm_goal     += goal_v
        dm_hours    += hours
        dm_variance += variance

        # Color logic: red = >90%, grey = 87%–89.999%, green = <85%
        if pct_used > 0.90:
            row_fill = fill(OVER_RED)
        elif 0.87 <= pct_used <= 0.89999:
            row_fill = fill(NEUTRAL)
        elif pct_used < 0.85:
            row_fill = fill(UNDER_GRN)
        else:
            row_fill = fill(WHITE)
        ws.row_dimensions[row_num].height = 18

        vals   = [row['Store #'], row['Store Name'], row['DM'],
                  goal_v, hours, variance, pct_used]
        fmts   = [None, None, None, FMT_HOURS, FMT_HOURS, FMT_VARIANCE, FMT_PCT]
        aligns = [center, left, center, center, center, center, center]

        for ci, (val, fmt, aln) in enumerate(zip(vals, fmts, aligns), 1):
            c = ws.cell(row=row_num, column=ci, value=val)
            c.font      = Font(name='Arial', size=10)
            c.fill      = row_fill
            c.alignment = aln
            c.border    = thin_border()
            if fmt:
                c.number_format = fmt
        row_num += 1

    # DM subtotal
    dm_pct = dm_hours / dm_goal if dm_goal else 0.0
    ws.row_dimensions[row_num].height = 20
    sub_vals   = ['', f'{dm_name} — Subtotal', '', dm_goal, dm_hours, round(dm_variance, 2), dm_pct]
    sub_fmts   = [None, None, None, FMT_HOURS, FMT_HOURS, FMT_VARIANCE, FMT_PCT]
    sub_aligns = [center, left, center, center, center, center, center]

    for ci, (val, fmt, aln) in enumerate(zip(sub_vals, sub_fmts, sub_aligns), 1):
        c = ws.cell(row=row_num, column=ci, value=val)
        c.font      = Font(name='Arial', size=10, bold=True, color='1F3864')
        c.fill      = fill(SUBTOTAL)
        c.alignment = aln
        c.border    = subtotal_border()
        if fmt:
            c.number_format = fmt
    row_num += 1

    grand_goal     += dm_goal
    grand_hours    += dm_hours
    grand_variance += dm_variance

# Grand total row
grand_pct = grand_hours / grand_goal if grand_goal else 0.0
ws.row_dimensions[row_num].height = 24
gt_vals  = ['', 'GRAND TOTAL', '', grand_goal, grand_hours, round(grand_variance, 2), grand_pct]
gt_fmts  = [None, None, None, FMT_HOURS, FMT_HOURS, FMT_VARIANCE, FMT_PCT]
gt_aligns = [center, left, center, center, center, center, center]

for ci, (val, fmt, aln) in enumerate(zip(gt_vals, gt_fmts, gt_aligns), 1):
    c = ws.cell(row=row_num, column=ci, value=val)
    c.font      = Font(name='Arial', size=11, bold=True, color=WHITE)
    c.fill      = fill(DARK_NAVY)
    c.alignment = aln
    c.border    = thin_border()
    if fmt:
        c.number_format = fmt

# Legend
row_num += 2
ws.merge_cells(f'A{row_num}:C{row_num}')
ws[f'A{row_num}'].value = 'Legend'
ws[f'A{row_num}'].font  = Font(name='Arial', size=10, bold=True)
row_num += 1
for hex_c, label in [(OVER_RED, '> 90% of Weekly Goal Used'), (NEUTRAL, '87% – 89.999% of Weekly Goal Used'), (UNDER_GRN, '< 85% of Weekly Goal Used')]:
    ws.cell(row=row_num, column=1).fill   = fill(hex_c)
    ws.cell(row=row_num, column=1).border = thin_border()
    c = ws.cell(row=row_num, column=2, value=label)
    c.font      = Font(name='Arial', size=10)
    c.alignment = left
    row_num += 1

# Footnote
row_num += 1
ws.merge_cells(f'A{row_num}:G{row_num}')
note = ws[f'A{row_num}']
note.value     = '* Variance and % Used are measured against the full weekly hourly goal. This report covers Thu–Sun only.'
note.font      = Font(name='Arial', size=9, italic=True, color='666666')
note.alignment = left

ws.freeze_panes = 'A3'

wb.save(OUTPUT_FILE)
print(f'Done — Report saved to {OUTPUT_FILE}')
print(f'Total Thu–Sun Hours: {grand_hours:,.2f}  |  Weekly Goal: {grand_goal:,.2f}  |  % Used: {grand_pct:.1%}')
