"""
AVS Labor Report Engine
-----------------------
Shared processing functions for AVS report types:
  - Weekly Full Report (AvS Summary + Store Rankings + DM Rankings)
  - Mid-Week Pulse (day-specific pacing thresholds via DAY_THRESHOLDS)

All functions accept uploaded file objects and reference data DataFrames,
returning (BytesIO, DataFrame) tuples.
"""

import io
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Excel style constants (shared across all reports)
# ---------------------------------------------------------------------------
DARK_NAVY = "1F3864"
MID_BLUE = "2E5FA3"
OVER_RED = "FFDADA"
UNDER_GRN = "D6F0DA"
NEUTRAL = "F2F2F2"
SUBTOTAL_CLR = "E2E8F0"
WHITE = "FFFFFF"

FMT_CURRENCY = '$#,##0.00'
FMT_HOURS = '#,##0.00'
FMT_VARIANCE = '+#,##0.00;[Red]-#,##0.00;"-"'
FMT_PCT = '0.0%'

_center = Alignment(horizontal="center", vertical="center")
_left = Alignment(horizontal="left", vertical="center", indent=1)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _subtotal_border():
    thick = Side(style="medium", color="2E5FA3")
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thick, bottom=thin)


# ---------------------------------------------------------------------------
# ADP CSV pre-processor
# ---------------------------------------------------------------------------
def preprocess_adp_csv(uploaded_file) -> pd.DataFrame:
    """
    Read an ADP payroll CSV, fixing Ohio store names that have unquoted commas
    (e.g. '112-0019 Streetsboro, OH' → '112-0019 Streetsboro (OH)').
    Returns a cleaned DataFrame.
    """
    if hasattr(uploaded_file, "read"):
        raw_bytes = uploaded_file.read()
        if hasattr(uploaded_file, "seek"):
            uploaded_file.seek(0)
        text = raw_bytes.decode("utf-8-sig")
    else:
        with open(uploaded_file, "r", encoding="utf-8-sig") as f:
            text = f.read()

    lines = text.splitlines(keepends=True)
    fixed_lines = []
    for line in lines:
        fields = line.rstrip("\r\n").split(",")
        if len(fields) == 10:
            merged = '"' + fields[3] + "," + fields[4] + '"'
            fields = fields[:3] + [merged] + fields[5:]
        fixed_lines.append(",".join(fields) + "\n")

    raw_payroll = pd.read_csv(io.StringIO("".join(fixed_lines)), dtype=str)
    raw_payroll["Batch Description"] = raw_payroll["Batch Description"].str.replace(
        r"(112-\d{4}\s+\w[\w\s]*?)\s+OH\b",
        lambda m: m.group(1) + " (OH)",
        regex=True,
    )
    for col in ["Temp Rate", "Reg Hours", "O/T Hours"]:
        raw_payroll[col] = pd.to_numeric(raw_payroll[col], errors="coerce").fillna(0.0)
    return raw_payroll


# ---------------------------------------------------------------------------
# Payroll aggregation
# ---------------------------------------------------------------------------
PAYROLL_TAX_RATE = 0.0765
WORKERS_COMP_RATE = 0.0100
LOAD_FACTOR = 1 + PAYROLL_TAX_RATE + WORKERS_COMP_RATE  # 1.0865


def _aggregate_payroll(raw_payroll, include_wages=False):
    """Aggregate hours (and optionally wages) by store, excluding 112-9001."""
    pf = raw_payroll[
        ~raw_payroll["Batch Description"].str.contains("112-9001", na=False)
    ].copy()
    pf["store_num"] = pf["Batch Description"].str.extract(r"(112-\d{4})")
    pf["total_hours"] = pf["Reg Hours"] + pf["O/T Hours"]

    agg_dict = {"total_hours": ("total_hours", "sum")}
    if include_wages:
        pf["gross_wages"] = pf["Temp Rate"] * pf["total_hours"]
        agg_dict["reg_hours"] = ("Reg Hours", "sum")
        agg_dict["ot_hours"] = ("O/T Hours", "sum")
        agg_dict["gross_wages"] = ("gross_wages", "sum")

    pagg = pf.groupby("store_num").agg(**agg_dict).reset_index()
    pagg["total_hours"] = pagg["total_hours"].round(2)
    if include_wages:
        pagg["gross_wages"] = pagg["gross_wages"].round(2)
        pagg["loaded_payroll"] = (pagg["gross_wages"] * LOAD_FACTOR).round(2)
    return pagg


# ---------------------------------------------------------------------------
# Sales loader & normalizer (for weekly full report)
# ---------------------------------------------------------------------------
def _normalize(name):
    return (
        str(name)
        .lower()
        .replace(",", "")
        .replace("(", "")
        .replace(")", "")
        .replace("-", " ")
        .strip()
    )


_NAME_OVERRIDES = {"denton   teasley tx": "denton tx"}


def load_net_sales(uploaded_file) -> pd.DataFrame:
    """Read End_Of_Week_Net_Sales Excel, return normalized DataFrame."""
    sales_raw = pd.read_excel(uploaded_file)
    sales = sales_raw[["Restaurant", "Last Week Net Sales"]].copy()
    sales = sales[
        ~sales["Restaurant"].str.contains(
            "Net Sale Averages|Stillwater- OSU", na=False
        )
    ]
    sales["norm"] = sales["Restaurant"].apply(_normalize).replace(_NAME_OVERRIDES)
    return sales


# ---------------------------------------------------------------------------
# Merge helper
# ---------------------------------------------------------------------------
def _build_merged_df(ref_data, band_goals, pagg, sales=None):
    """
    Merge reference data + payroll aggregates + (optionally) sales.
    Returns a DataFrame ready for report generation.
    """
    dm = ref_data.rename(
        columns={"location_id": "Store #", "store_name": "Store Name", "dm": "DM", "revenue_band": "Rev Band"}
    ).copy()
    # Drop locked-config-only columns that would otherwise create duplicates
    # (e.g. 'hourly_goal' from the DB conflicts with the computed 'Hourly Goal' below)
    for _col in ["hourly_goal", "week_start", "source", "status"]:
        if _col in dm.columns:
            dm = dm.drop(columns=[_col])
    dm["Hourly Goal"] = dm["Rev Band"].map(band_goals)

    if sales is not None:
        dm["norm"] = dm["Store Name"].apply(_normalize)
        dm = dm.merge(sales[["norm", "Last Week Net Sales"]], on="norm", how="left")

    dm = dm.merge(
        pagg.rename(columns={"total_hours": "actual_hours"})[
            ["store_num", "actual_hours"]
            + (["loaded_payroll"] if "loaded_payroll" in pagg.columns else [])
        ],
        left_on="Store #",
        right_on="store_num",
        how="left",
    )
    dm["actual_hours"] = dm["actual_hours"].fillna(0.0)
    dm["Variance"] = (dm["actual_hours"] - dm["Hourly Goal"]).round(2)

    if "loaded_payroll" in dm.columns:
        dm["loaded_payroll"] = dm["loaded_payroll"].fillna(0.0)

    dm = dm.sort_values(["DM", "Store #"]).reset_index(drop=True)
    return dm


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT GENERATORS
# ═══════════════════════════════════════════════════════════════════════════════

def generate_weekly_report(adp_file, sales_file, ref_data, band_goals, report_dates):
    """
    Generate the full weekly AVS Labor Report (3 tabs).
    Returns BytesIO with the Excel workbook.
    """
    raw_payroll = preprocess_adp_csv(adp_file)
    pagg = _aggregate_payroll(raw_payroll, include_wages=True)
    sales = load_net_sales(sales_file)
    df = _build_merged_df(ref_data, band_goals, pagg, sales=sales)

    wb = Workbook()

    # === TAB 1: AvS Summary ===
    ws = wb.active
    ws.title = "AvS Summary"
    ws.sheet_view.showGridLines = False

    headers = [
        "Store #", "Store Name", "DM", "Rev Band", "Hourly Goal",
        "Actual Weekly Sales", "Actual Labor Hours", "Variance (Hrs)",
        "Est. Payroll Expense", "Est. Labor %",
    ]
    col_widths = [12, 34, 10, 12, 13, 22, 20, 18, 22, 14]

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"AvS Weekly Labor Report - {report_dates}"
    c.font = Font(name="Arial", size=16, bold=True, color=WHITE)
    c.fill = _fill(DARK_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 24
    for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = _fill(MID_BLUE)
        c.alignment = _center
        c.border = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    row_num = 3
    grand_sales = grand_hours = grand_variance = grand_goal = grand_payroll = 0.0

    for dm_name, group in df.groupby("DM", sort=True):
        dm_sales = dm_hours = dm_variance = dm_goal = dm_payroll = 0.0
        dm_labor_num = dm_labor_den = 0.0

        for _, row in group.iterrows():
            variance = float(row["Variance"]) if pd.notna(row["Variance"]) else 0.0
            hours = float(row["actual_hours"]) if pd.notna(row["actual_hours"]) else 0.0
            sales_v = float(row["Last Week Net Sales"]) if pd.notna(row.get("Last Week Net Sales")) else 0.0
            goal_v = float(row["Hourly Goal"]) if pd.notna(row["Hourly Goal"]) else 0.0
            payroll_v = float(row["loaded_payroll"]) if pd.notna(row.get("loaded_payroll")) else 0.0
            labor_pct = payroll_v / sales_v if sales_v else 0.0

            dm_sales += sales_v
            dm_hours += hours
            dm_variance += variance
            dm_goal += goal_v
            dm_payroll += payroll_v
            dm_labor_num += payroll_v
            dm_labor_den += sales_v

            row_fill = _fill(OVER_RED) if variance > 0 else _fill(UNDER_GRN) if variance < 0 else _fill(NEUTRAL)
            ws.row_dimensions[row_num].height = 18

            vals = [row["Store #"], row["Store Name"], row["DM"], row["Rev Band"],
                    goal_v, sales_v, hours, variance, payroll_v, labor_pct]
            fmts = [None, None, None, None, FMT_HOURS, FMT_CURRENCY, FMT_HOURS, FMT_VARIANCE, FMT_CURRENCY, "0.0%"]
            aligns = [_center, _left, _center, _center, _center, _center, _center, _center, _center, _center]

            for ci, (val, fmt, aln) in enumerate(zip(vals, fmts, aligns), 1):
                c = ws.cell(row=row_num, column=ci, value=val)
                c.font = Font(name="Arial", size=10)
                c.fill = row_fill
                c.alignment = aln
                c.border = _thin_border()
                if fmt:
                    c.number_format = fmt
            row_num += 1

        dm_labor_pct = dm_labor_num / dm_labor_den if dm_labor_den else 0.0
        ws.row_dimensions[row_num].height = 20
        sub_vals = ["", f"{dm_name} — Subtotal", "", "", dm_goal, dm_sales, dm_hours, dm_variance, dm_payroll, dm_labor_pct]
        sub_fmts = [None, None, None, None, FMT_HOURS, FMT_CURRENCY, FMT_HOURS, FMT_VARIANCE, FMT_CURRENCY, "0.0%"]
        for ci, (val, fmt) in enumerate(zip(sub_vals, sub_fmts), 1):
            c = ws.cell(row=row_num, column=ci, value=val)
            c.font = Font(name="Arial", size=10, bold=True, color="1F3864")
            c.fill = _fill(SUBTOTAL_CLR)
            c.alignment = _left if ci == 2 else _center
            c.border = _subtotal_border()
            if fmt:
                c.number_format = fmt
        row_num += 1

        grand_sales += dm_sales
        grand_hours += dm_hours
        grand_variance += dm_variance
        grand_goal += dm_goal
        grand_payroll += dm_payroll

    # Grand total
    grand_labor_pct = grand_payroll / grand_sales if grand_sales else 0.0
    ws.row_dimensions[row_num].height = 24
    gt_vals = ["", "GRAND TOTAL", "", "", grand_goal, grand_sales, grand_hours, grand_variance, grand_payroll, grand_labor_pct]
    gt_fmts = [None, None, None, None, FMT_HOURS, FMT_CURRENCY, FMT_HOURS, FMT_VARIANCE, FMT_CURRENCY, "0.0%"]
    for ci, (val, fmt) in enumerate(zip(gt_vals, gt_fmts), 1):
        c = ws.cell(row=row_num, column=ci, value=val)
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = _fill(DARK_NAVY)
        c.alignment = _center if ci != 2 else _left
        c.border = _thin_border()
        if fmt:
            c.number_format = fmt

    row_num += 2
    ws.merge_cells(f"A{row_num}:C{row_num}")
    ws[f"A{row_num}"].value = "Legend"
    ws[f"A{row_num}"].font = Font(name="Arial", size=10, bold=True)
    row_num += 1
    for hex_c, label in [(OVER_RED, "Over Goal"), (UNDER_GRN, "Under Goal"), (NEUTRAL, "At Goal")]:
        ws.cell(row=row_num, column=1).fill = _fill(hex_c)
        ws.cell(row=row_num, column=1).border = _thin_border()
        c = ws.cell(row=row_num, column=2, value=label)
        c.font = Font(name="Arial", size=10)
        c.alignment = _left
        row_num += 1

    row_num += 1
    ws.merge_cells(f"A{row_num}:J{row_num}")
    note = ws[f"A{row_num}"]
    note.value = f"* Est. Payroll Expense = Gross Wages (Temp Rate x Hours) x {LOAD_FACTOR:.4f} load factor (incl. {PAYROLL_TAX_RATE*100:.2f}% payroll tax + {WORKERS_COMP_RATE*100:.1f}% workers comp)"
    note.font = Font(name="Arial", size=9, italic=True, color="666666")
    note.alignment = _left
    ws.freeze_panes = "A3"

    # === TAB 2: Store Rankings ===
    wr = wb.create_sheet("Store Rankings")
    wr.sheet_view.showGridLines = False
    rank_headers = ["Rank", "Store #", "Store Name", "DM", "Hourly Goal", "Actual Labor Hours", "Variance (Hrs)", "Est. Labor %"]
    rank_widths = [8, 12, 34, 12, 13, 20, 18, 14]

    wr.row_dimensions[1].height = 36
    wr.merge_cells("A1:H1")
    c = wr["A1"]
    c.value = f"AvS Weekly Labor Report - Store Rankings - {report_dates}"
    c.font = Font(name="Arial", size=16, bold=True, color=WHITE)
    c.fill = _fill(DARK_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    wr.row_dimensions[2].height = 24
    for ci, (hdr, w) in enumerate(zip(rank_headers, rank_widths), 1):
        c = wr.cell(row=2, column=ci, value=hdr)
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = _fill(MID_BLUE)
        c.alignment = _center
        c.border = _thin_border()
        wr.column_dimensions[get_column_letter(ci)].width = w

    ranked = df.copy()
    ranked["labor_pct"] = ranked["loaded_payroll"] / ranked["Last Week Net Sales"].replace(0, float("nan"))
    ranked["labor_pct"] = ranked["labor_pct"].fillna(0)
    ranked = ranked.sort_values("Variance", ascending=True).reset_index(drop=True)

    for rank_idx, row in ranked.iterrows():
        rn = rank_idx + 3
        variance = float(row["Variance"]) if pd.notna(row["Variance"]) else 0.0
        hours = float(row["actual_hours"]) if pd.notna(row["actual_hours"]) else 0.0
        goal_v = float(row["Hourly Goal"]) if pd.notna(row["Hourly Goal"]) else 0.0
        lp = float(row["labor_pct"]) if pd.notna(row["labor_pct"]) else 0.0

        row_fill = _fill(OVER_RED) if variance > 0 else _fill(UNDER_GRN) if variance < 0 else _fill(NEUTRAL)
        wr.row_dimensions[rn].height = 18
        vals = [rank_idx + 1, row["Store #"], row["Store Name"], row["DM"], goal_v, hours, variance, lp]
        fmts = [None, None, None, None, FMT_HOURS, FMT_HOURS, FMT_VARIANCE, "0.0%"]
        aligns = [_center, _center, _left, _center, _center, _center, _center, _center]
        for ci, (val, fmt, aln) in enumerate(zip(vals, fmts, aligns), 1):
            c = wr.cell(row=rn, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = row_fill
            c.alignment = aln
            c.border = _thin_border()
            if fmt:
                c.number_format = fmt

    leg_row = len(ranked) + 5
    wr.merge_cells(f"A{leg_row}:C{leg_row}")
    wr[f"A{leg_row}"].value = "Legend"
    wr[f"A{leg_row}"].font = Font(name="Arial", size=10, bold=True)
    leg_row += 1
    for hex_c, label in [(UNDER_GRN, "Under Goal — Best"), (OVER_RED, "Over Goal — Worst"), (NEUTRAL, "At Goal")]:
        wr.cell(row=leg_row, column=1).fill = _fill(hex_c)
        wr.cell(row=leg_row, column=1).border = _thin_border()
        c = wr.cell(row=leg_row, column=2, value=label)
        c.font = Font(name="Arial", size=10)
        c.alignment = _left
        leg_row += 1
    wr.freeze_panes = "A3"

    # === TAB 3: DM Rankings ===
    wdm = wb.create_sheet("DM Rankings")
    wdm.sheet_view.showGridLines = False
    dm_headers = ["Rank", "DM", "Hourly Goal", "Actual Labor Hours", "Variance (Hrs)", "Est. Payroll Expense", "Est. Labor %"]
    dm_widths = [8, 14, 13, 20, 18, 22, 14]

    wdm.row_dimensions[1].height = 36
    wdm.merge_cells("A1:G1")
    c = wdm["A1"]
    c.value = f"AvS Weekly Labor Report - DM Rankings - {report_dates}"
    c.font = Font(name="Arial", size=16, bold=True, color=WHITE)
    c.fill = _fill(DARK_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    wdm.row_dimensions[2].height = 24
    for ci, (hdr, w) in enumerate(zip(dm_headers, dm_widths), 1):
        c = wdm.cell(row=2, column=ci, value=hdr)
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = _fill(MID_BLUE)
        c.alignment = _center
        c.border = _thin_border()
        wdm.column_dimensions[get_column_letter(ci)].width = w

    dm_summary = []
    for dm_name, group in df.groupby("DM", sort=True):
        dm_sales_t = group["Last Week Net Sales"].fillna(0).sum()
        dm_hours_t = group["actual_hours"].fillna(0).sum()
        dm_goal_t = group["Hourly Goal"].fillna(0).sum()
        dm_payroll_t = group["loaded_payroll"].fillna(0).sum()
        dm_var_t = round(dm_hours_t - dm_goal_t, 2)
        dm_lp = dm_payroll_t / dm_sales_t if dm_sales_t else 0.0
        dm_summary.append({
            "DM": dm_name, "Hourly Goal": round(dm_goal_t, 2),
            "Actual Labor Hours": round(dm_hours_t, 2), "Variance": dm_var_t,
            "Est. Payroll Expense": round(dm_payroll_t, 2), "Est. Labor %": dm_lp,
        })

    dm_ranked = sorted(dm_summary, key=lambda x: x["Variance"])
    for rank_idx, dm_row in enumerate(dm_ranked):
        rn = rank_idx + 3
        variance = dm_row["Variance"]
        row_fill = _fill(OVER_RED) if variance > 0 else _fill(UNDER_GRN) if variance < 0 else _fill(NEUTRAL)
        wdm.row_dimensions[rn].height = 20
        vals = [rank_idx + 1, dm_row["DM"], dm_row["Hourly Goal"], dm_row["Actual Labor Hours"],
                variance, dm_row["Est. Payroll Expense"], dm_row["Est. Labor %"]]
        fmts = [None, None, FMT_HOURS, FMT_HOURS, FMT_VARIANCE, FMT_CURRENCY, "0.0%"]
        for ci, (val, fmt) in enumerate(zip(vals, fmts), 1):
            c = wdm.cell(row=rn, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = row_fill
            c.alignment = _center
            c.border = _thin_border()
            if fmt:
                c.number_format = fmt

    leg_row = len(dm_ranked) + 5
    wdm.merge_cells(f"A{leg_row}:C{leg_row}")
    wdm[f"A{leg_row}"].value = "Legend"
    wdm[f"A{leg_row}"].font = Font(name="Arial", size=10, bold=True)
    leg_row += 1
    for hex_c, label in [(UNDER_GRN, "Under Goal — Best"), (OVER_RED, "Over Goal — Worst"), (NEUTRAL, "At Goal")]:
        wdm.cell(row=leg_row, column=1).fill = _fill(hex_c)
        wdm.cell(row=leg_row, column=1).border = _thin_border()
        c = wdm.cell(row=leg_row, column=2, value=label)
        c.font = Font(name="Arial", size=10)
        c.alignment = _left
        leg_row += 1
    wdm.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, df



# ---------------------------------------------------------------------------
# Day-specific color thresholds for Mid-Week Pulse
# Orange = below green_min (under-pacing)
# Green  = green_min to green_max (on pace)
# Red    = above green_max (over-pacing)
# Note: red_above is set equal to green_max so there is no gap
# ---------------------------------------------------------------------------
DAY_THRESHOLDS = {
    "Friday":    {"green_min": 0.1179, "green_max": 0.1679, "red_above": 0.1679},
    "Saturday":  {"green_min": 0.2607, "green_max": 0.3107, "red_above": 0.3107},
    "Sunday":    {"green_min": 0.54,   "green_max": 0.599,  "red_above": 0.599},
    "Monday":    {"green_min": 0.5464, "green_max": 0.5964, "red_above": 0.5964},
    "Tuesday":   {"green_min": 0.6893, "green_max": 0.7393, "red_above": 0.7393},
    "Wednesday": {"green_min": 0.8321, "green_max": 0.8821, "red_above": 0.8821},
}

# Color constants for the new 3-color scheme
LIGHT_ORANGE = "FFCC80"  # Under-pacing (behind schedule)
LIGHT_GREEN = "A5D6A7"   # On track
LIGHT_RED = "EF9A9A"     # Over-pacing (ahead of schedule / too many hours)


def generate_midweek_report(adp_file, ref_data, band_goals, report_dates, through_day="Friday"):
    """Mid-Week Pulse with day-specific color thresholds."""
    thresholds = DAY_THRESHOLDS.get(through_day, DAY_THRESHOLDS["Friday"])
    green_min = thresholds["green_min"]
    green_max = thresholds["green_max"]
    red_above = thresholds["red_above"]

    raw_payroll = preprocess_adp_csv(adp_file)
    pagg = _aggregate_payroll(raw_payroll, include_wages=False)
    df = _build_merged_df(ref_data, band_goals, pagg)
    df["Pct Used"] = (df["actual_hours"] / df["Hourly Goal"]).where(df["Hourly Goal"] > 0, 0.0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mid-Week Labor Pulse"
    ws.sheet_view.showGridLines = False

    headers = ["Store #", "Store Name", "DM", "Weekly Hourly Goal",
               "Actual Hours", "Variance (Hrs)", "% of Weekly Goal Used", "Status"]
    col_widths = [12, 34, 10, 20, 22, 18, 24, 18]

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"AvS Mid-Week Labor Pulse — Through {through_day} — {report_dates}"
    c.font = Font(name="Arial", size=16, bold=True, color=WHITE)
    c.fill = _fill(DARK_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 24
    for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = _fill(MID_BLUE)
        c.alignment = _center
        c.border = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    row_num = 3
    grand_goal = grand_hours = grand_variance = 0.0

    for dm_name, group in df.groupby("DM", sort=True):
        dm_goal = dm_hours = dm_variance = 0.0

        for _, row in group.iterrows():
            variance = float(row["Variance"]) if pd.notna(row["Variance"]) else 0.0
            hours = float(row["actual_hours"]) if pd.notna(row["actual_hours"]) else 0.0
            goal_v = float(row["Hourly Goal"]) if pd.notna(row["Hourly Goal"]) else 0.0
            pct_used = float(row["Pct Used"]) if pd.notna(row["Pct Used"]) else 0.0

            dm_goal += goal_v
            dm_hours += hours
            dm_variance += variance

            # Day-specific color coding and status
            if pct_used > red_above:
                row_fill = _fill(LIGHT_RED)
                status = "Over Pacing"
            elif green_min <= pct_used <= green_max:
                row_fill = _fill(LIGHT_GREEN)
                status = "On Pace"
            elif pct_used < green_min:
                row_fill = _fill(LIGHT_ORANGE)
                status = "Under Pacing"
            else:
                row_fill = _fill(WHITE)
                status = ""
            ws.row_dimensions[row_num].height = 18

            vals = [row["Store #"], row["Store Name"], row["DM"], goal_v, hours, variance, pct_used, status]
            fmts = [None, None, None, FMT_HOURS, FMT_HOURS, FMT_VARIANCE, FMT_PCT, None]
            aligns = [_center, _left, _center, _center, _center, _center, _center, _center]
            for ci, (val, fmt, aln) in enumerate(zip(vals, fmts, aligns), 1):
                c = ws.cell(row=row_num, column=ci, value=val)
                c.font = Font(name="Arial", size=10)
                c.fill = row_fill
                c.alignment = aln
                c.border = _thin_border()
                if fmt:
                    c.number_format = fmt
            row_num += 1

        dm_pct = dm_hours / dm_goal if dm_goal else 0.0
        dm_status = "Over Pacing" if dm_pct > red_above else ("On Pace" if green_min <= dm_pct <= green_max else ("Under Pacing" if dm_pct < green_min else ""))
        ws.row_dimensions[row_num].height = 20
        sub_vals = ["", f"{dm_name} — Subtotal", "", dm_goal, dm_hours, round(dm_variance, 2), dm_pct, dm_status]
        sub_fmts = [None, None, None, FMT_HOURS, FMT_HOURS, FMT_VARIANCE, FMT_PCT, None]
        for ci, (val, fmt) in enumerate(zip(sub_vals, sub_fmts), 1):
            c = ws.cell(row=row_num, column=ci, value=val)
            c.font = Font(name="Arial", size=10, bold=True, color="1F3864")
            c.fill = _fill(SUBTOTAL_CLR)
            c.alignment = _left if ci == 2 else _center
            c.border = _subtotal_border()
            if fmt:
                c.number_format = fmt
        row_num += 1
        grand_goal += dm_goal
        grand_hours += dm_hours
        grand_variance += dm_variance

    # Grand total
    grand_pct = grand_hours / grand_goal if grand_goal else 0.0
    grand_status = "Over Pacing" if grand_pct > red_above else ("On Pace" if green_min <= grand_pct <= green_max else ("Under Pacing" if grand_pct < green_min else ""))
    ws.row_dimensions[row_num].height = 24
    gt_vals = ["", "GRAND TOTAL", "", grand_goal, grand_hours, round(grand_variance, 2), grand_pct, grand_status]
    gt_fmts = [None, None, None, FMT_HOURS, FMT_HOURS, FMT_VARIANCE, FMT_PCT, None]
    for ci, (val, fmt) in enumerate(zip(gt_vals, gt_fmts), 1):
        c = ws.cell(row=row_num, column=ci, value=val)
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = _fill(DARK_NAVY)
        c.alignment = _center if ci != 2 else _left
        c.border = _thin_border()
        if fmt:
            c.number_format = fmt

    # Legend
    row_num += 2
    ws.merge_cells(f"A{row_num}:C{row_num}")
    ws[f"A{row_num}"].value = "Legend"
    ws[f"A{row_num}"].font = Font(name="Arial", size=10, bold=True)
    row_num += 1
    legend_items = [
        (LIGHT_RED, f"Above {red_above:.0%} of Weekly Goal — Over Pacing"),
        (LIGHT_GREEN, f"{green_min:.0%} – {green_max:.1%} of Weekly Goal — On Pace"),
        (LIGHT_ORANGE, f"Below {green_min:.0%} of Weekly Goal — Under Pacing"),
    ]
    for hex_c, label in legend_items:
        ws.cell(row=row_num, column=1).fill = _fill(hex_c)
        ws.cell(row=row_num, column=1).border = _thin_border()
        c = ws.cell(row=row_num, column=2, value=label)
        c.font = Font(name="Arial", size=10)
        c.alignment = _left
        row_num += 1

    row_num += 1
    ws.merge_cells(f"A{row_num}:H{row_num}")
    note = ws[f"A{row_num}"]
    note.value = f"* Thresholds based on expected cumulative usage through {through_day}."
    note.font = Font(name="Arial", size=9, italic=True, color="666666")
    note.alignment = _left
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, df
