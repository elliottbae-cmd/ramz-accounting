import glob
import os
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
# FILE PATHS — reference files (update only when tables change)
# ══════════════════════════════════════════════════════════════════════════════
OUTPUT_FILE  = '/mnt/user-data/outputs/AVS_Labor_Report.xlsx'
REPORT_DATES = input("Enter the report date range (e.g. 3.5.26 - 3.11.26): ").strip()

# ── Auto-detect uploaded files by pattern ────────────────────────────────────
def find_file(folder, pattern, label):
    matches = glob.glob(os.path.join(folder, pattern))
    if not matches:
        raise FileNotFoundError(f"No {label} file found in {folder} matching '{pattern}'. Please upload the file and try again.")
    if len(matches) > 1:
        # Pick the most recently modified if multiple matches
        matches = sorted(matches, key=os.path.getmtime, reverse=True)
        print(f"Warning: Multiple {label} files found — using most recent: {os.path.basename(matches[0])}")
    else:
        print(f"Found {label}: {os.path.basename(matches[0])}")
    return matches[0]

UPLOADS = '/mnt/user-data/uploads'
ADP_FILE   = find_file(UPLOADS, 'adp_payroll*.csv',              'ADP Payroll CSV')
SALES_FILE = find_file(UPLOADS, 'End_Of_Week_Net_Sales*.xlsx',   'End of Week Sales XLSX')

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Clean ADP export (fix OH store delineation)
# ══════════════════════════════════════════════════════════════════════════════
# WHY THIS IS NEEDED (do not remove):
#   The ADP CSV export is a third-party file we cannot control. Some Ohio store
#   names are exported with an unquoted comma — e.g. "112-0019 Streetsboro, OH"
#   — which causes pandas to read 10 fields instead of the expected 9, throwing
#   a ParserError. This pre-processing step fixes those rows in memory before
#   pandas ever sees the file. No data is lost or altered; the store name is
#   simply re-quoted so the CSV is valid. This will be required on every run
#   until ADP changes their export format.
import io
with open(ADP_FILE, 'r', encoding='utf-8-sig') as f:
    lines = f.readlines()
fixed_lines = []
for line in lines:
    fields = line.rstrip('\r\n').split(',')
    if len(fields) == 10:
        # Columns 3 & 4 are the two halves of a store name split on its comma.
        # Re-join them into one quoted field so the row is 9 fields again.
        merged = '"' + fields[3] + ',' + fields[4] + '"'
        fields = fields[:3] + [merged] + fields[5:]
    fixed_lines.append(','.join(fields) + '\n')
raw_payroll = pd.read_csv(io.StringIO(''.join(fixed_lines)), dtype=str)
raw_payroll['Batch Description'] = raw_payroll['Batch Description'].str.replace(
    r'(112-\d{4}\s+\w[\w\s]*?)\s+OH\b', lambda m: m.group(1) + ' (OH)', regex=True
)
# Convert numeric columns back
for col in ['Temp Rate', 'Reg Hours', 'O/T Hours']:
    raw_payroll[col] = pd.to_numeric(raw_payroll[col], errors='coerce').fillna(0.0)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Reference tables (edit directly when DMs, bands, or goals change)
# ══════════════════════════════════════════════════════════════════════════════

# ── DM Mapping (update when DM assignments change) ───────────────────────────
dm = pd.DataFrame([
    {'Store #': '112-0019', 'Store Name': 'Streetsboro (OH)',              'DM': 'Alex'},
    {'Store #': '112-0020', 'Store Name': 'Strongsville (OH)',             'DM': 'Alex'},
    {'Store #': '112-0022', 'Store Name': 'Avon (OH)',                     'DM': 'Alex'},
    {'Store #': '112-0023', 'Store Name': 'Lorain (OH)',                   'DM': 'Alex'},
    {'Store #': '112-0015', 'Store Name': 'Centerville (OH)',              'DM': 'Cody'},
    {'Store #': '112-0024', 'Store Name': 'Beavercreek (OH)',              'DM': 'Cody'},
    {'Store #': '112-0028', 'Store Name': 'Englewood - S Main (OH)',       'DM': 'Cody'},
    {'Store #': '112-0029', 'Store Name': 'Grove City (OH)',               'DM': 'Cody'},
    {'Store #': '112-0030', 'Store Name': 'Hilliard (OH)',                 'DM': 'Cody'},
    {'Store #': '112-0002', 'Store Name': 'Muskogee (OK)',                 'DM': 'Dottie'},
    {'Store #': '112-0031', 'Store Name': 'McAlester (OK)',                'DM': 'Dottie'},
    {'Store #': '112-0035', 'Store Name': 'Tulsa - 51st (OK)',             'DM': 'Dottie'},
    {'Store #': '112-0038', 'Store Name': 'Ada (OK)',                      'DM': 'Dottie'},
    {'Store #': '112-0012', 'Store Name': 'Amelia (OH)',                   'DM': 'Renee'},
    {'Store #': '112-0017', 'Store Name': 'West Chester (OH)',             'DM': 'Cody'},
    {'Store #': '112-0018', 'Store Name': 'Harrison (OH)',                 'DM': 'Renee'},
    {'Store #': '112-0026', 'Store Name': 'Fairfield (OH)',                'DM': 'Alex'},
    {'Store #': '112-0027', 'Store Name': 'Florence (KY)',                 'DM': 'Alex'},
    {'Store #': '112-0001', 'Store Name': 'Stillwater (OK)',               'DM': 'Manny'},
    {'Store #': '112-0003', 'Store Name': 'Lawton (OK)',                   'DM': 'Manny'},
    {'Store #': '112-0004', 'Store Name': 'Ardmore (OK)',                  'DM': 'Manny'},
    {'Store #': '112-0005', 'Store Name': 'Durant (OK)',                   'DM': 'Manny'},
    {'Store #': '112-0008', 'Store Name': 'Lufkin (TX)',                   'DM': 'Miriam'},
    {'Store #': '112-0009', 'Store Name': 'Nacogdoches (TX)',              'DM': 'Miriam'},
    {'Store #': '112-0011', 'Store Name': 'Princeton (TX)',                'DM': 'Miriam'},
    {'Store #': '112-0032', 'Store Name': 'Plano (TX)',                    'DM': 'Miriam'},
    {'Store #': '112-0034', 'Store Name': 'Marshall (TX)',                 'DM': 'Miriam'},
    {'Store #': '112-0037', 'Store Name': 'McKinney - Hardin Blvd (TX)',   'DM': 'Miriam'},
    {'Store #': '112-0013', 'Store Name': 'Milford (OH)',                  'DM': 'Renee'},
    {'Store #': '112-0014', 'Store Name': 'Cincinnati - Winton Rd (OH)',   'DM': 'Renee'},
    {'Store #': '112-0016', 'Store Name': 'Cincinnati - Red Bank Rd (OH)', 'DM': 'Renee'},
    {'Store #': '112-0021', 'Store Name': 'Loveland (OH)',                 'DM': 'Renee'},
    {'Store #': '112-0006', 'Store Name': 'Wichita Falls (TX)',            'DM': 'Steve'},
    {'Store #': '112-0007', 'Store Name': 'Gainesville (TX)',              'DM': 'Steve'},
    {'Store #': '112-0010', 'Store Name': 'Little Elm (TX)',               'DM': 'Steve'},
    {'Store #': '112-0033', 'Store Name': 'Brownwood (TX)',                'DM': 'Steve'},
    {'Store #': '112-0036', 'Store Name': 'Denton (TX)',                   'DM': 'Steve'},
])

# ── Revenue Band Hourly Goals (update when goals change) ─────────────────────
rev_bands = {
    '<25k':         422,
    '25k-30k':      420,
    '30k-35k':      504,
    '35k-40k':      588,
    '40k-45k':      605,
    '45k-50k':      849,
    '50k+':         756,
    'NRO Seasoned': 895,
    'NRO':          979,
}

# ── Store Revenue Band Assignments (update weekly as needed) ─────────────────
store_rev_bands = {
    '112-0001': '35k-40k', '112-0002': '35k-40k', '112-0003': '<25k',
    '112-0004': '<25k',    '112-0005': '<25k',    '112-0006': '30k-35k',
    '112-0007': '<25k',    '112-0008': '30k-35k', '112-0009': '25k-30k',
    '112-0010': '<25k',    '112-0011': '<25k',    '112-0012': '<25k',
    '112-0013': '<25k',    '112-0014': '<25k',    '112-0015': '45k-50k',
    '112-0016': '<25k',    '112-0017': '30k-35k', '112-0018': '30k-35k',
    '112-0019': '<25k',    '112-0020': '<25k',    '112-0021': '30k-35k',
    '112-0022': '<25k',    '112-0023': '<25k',    '112-0024': '25k-30k',
    '112-0026': '<25k',    '112-0027': '<25k',    '112-0028': '30k-35k',
    '112-0029': '<25k',    '112-0030': '<25k',    '112-0031': '<25k',
    '112-0032': '<25k',    '112-0033': '<25k',    '112-0034': '<25k',
    '112-0035': 'NRO Seasoned', '112-0036': '35k-40k', '112-0037': '35k-40k',
    '112-0038': 'NRO',
}

sales_raw = pd.read_excel(SALES_FILE)

PAYROLL_TAX_RATE  = 0.0765
WORKERS_COMP_RATE = 0.0100
LOAD_FACTOR       = 1 + PAYROLL_TAX_RATE + WORKERS_COMP_RATE  # 1.0865

# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Aggregate payroll (exclude 112-9001)
# ══════════════════════════════════════════════════════════════════════════════
pf = raw_payroll[~raw_payroll['Batch Description'].str.contains('112-9001', na=False)].copy()
pf['store_num']   = pf['Batch Description'].str.extract(r'(112-\d{4})')
pf['total_hours'] = pf['Reg Hours'] + pf['O/T Hours']
pf['gross_wages'] = pf['Temp Rate'] * pf['total_hours']
pagg = pf.groupby('store_num').agg(
    reg_hours=('Reg Hours','sum'),
    ot_hours=('O/T Hours','sum'),
    gross_wages=('gross_wages','sum')
).reset_index()
pagg['actual_labor_hours'] = (pagg['reg_hours'] + pagg['ot_hours']).round(2)
pagg['gross_wages']        = pagg['gross_wages'].round(2)
pagg['loaded_payroll']     = (pagg['gross_wages'] * LOAD_FACTOR).round(2)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — Load & normalize sales
# ══════════════════════════════════════════════════════════════════════════════
def normalize(name):
    return str(name).lower().replace(',','').replace('(','').replace(')','').replace('-',' ').strip()

name_overrides = {'denton   teasley tx': 'denton tx'}

sales = sales_raw[['Restaurant','Last Week Net Sales']].copy()
sales = sales[~sales['Restaurant'].str.contains('Net Sale Averages|Stillwater- OSU', na=False)]
sales['norm'] = sales['Restaurant'].apply(normalize).replace(name_overrides)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 — Merge everything
# ══════════════════════════════════════════════════════════════════════════════
dm['norm'] = dm['Store Name'].apply(normalize)
df = dm.merge(sales[['norm','Last Week Net Sales']], on='norm', how='left')
df = df.merge(pagg[['store_num','actual_labor_hours','loaded_payroll']], left_on='Store #', right_on='store_num', how='left')
df['Rev Band']    = df['Store #'].map(store_rev_bands)
df['Hourly Goal'] = df['Rev Band'].map(rev_bands)
df['Variance']    = (df['actual_labor_hours'] - df['Hourly Goal']).round(2)
df = df.sort_values(['DM','Store #']).reset_index(drop=True)

# ══════════════════════════════════════════════════════════════════════════════
# STYLES
# ══════════════════════════════════════════════════════════════════════════════
DARK_NAVY = '1F3864'
MID_BLUE  = '2E5FA3'
OVER_RED  = 'FFDADA'
UNDER_GRN = 'D6F0DA'
NEUTRAL   = 'F2F2F2'
SUBTOTAL  = 'E2E8F0'
WHITE     = 'FFFFFF'

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def subtotal_border():
    thick = Side(style='medium', color='2E5FA3')
    thin  = Side(style='thin',   color='CCCCCC')
    return Border(left=thin, right=thin, top=thick, bottom=thin)

center = Alignment(horizontal='center', vertical='center')
left   = Alignment(horizontal='left',   vertical='center', indent=1)

FMT_CURRENCY = '$#,##0.00'
FMT_HOURS    = '#,##0.00'
FMT_VARIANCE = '+#,##0.00;[Red]-#,##0.00;"-"'

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — AvS Summary
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = 'AvS Summary'
ws.sheet_view.showGridLines = False

headers    = ['Store #','Store Name','DM','Rev Band','Hourly Goal',
              'Actual Weekly Sales','Actual Labor Hours','Variance (Hrs)',
              'Est. Payroll Expense','Est. Labor %']
col_widths = [12, 34, 10, 12, 13, 22, 20, 18, 22, 14]

ws.row_dimensions[1].height = 36
ws.merge_cells('A1:J1')
c = ws['A1']
c.value     = f'AvS Weekly Labor Report - {REPORT_DATES}'
c.font      = Font(name='Arial', size=16, bold=True, color=WHITE)
c.fill      = fill(DARK_NAVY)
c.alignment = Alignment(horizontal='center', vertical='center')

ws.row_dimensions[2].height = 24
for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
    c = ws.cell(row=2, column=ci, value=hdr)
    c.font      = Font(name='Arial', size=11, bold=True, color=WHITE)
    c.fill      = fill(MID_BLUE)
    c.alignment = center
    c.border    = thin_border()
    ws.column_dimensions[get_column_letter(ci)].width = w

row_num        = 3
grand_sales    = 0.0
grand_hours    = 0.0
grand_variance = 0.0
grand_goal     = 0.0
grand_payroll  = 0.0

for dm_name, group in df.groupby('DM', sort=True):
    dm_sales         = 0.0
    dm_hours         = 0.0
    dm_variance      = 0.0
    dm_goal          = 0.0
    dm_payroll       = 0.0
    dm_labor_pct_num = 0.0
    dm_labor_pct_den = 0.0

    for _, row in group.iterrows():
        variance  = float(row['Variance'])           if pd.notna(row['Variance'])           else 0.0
        hours     = float(row['actual_labor_hours'])  if pd.notna(row['actual_labor_hours'])  else 0.0
        sales_v   = float(row['Last Week Net Sales']) if pd.notna(row['Last Week Net Sales']) else 0.0
        goal_v    = float(row['Hourly Goal'])          if pd.notna(row['Hourly Goal'])          else 0.0
        payroll_v = float(row['loaded_payroll'])       if pd.notna(row['loaded_payroll'])       else 0.0
        labor_pct = payroll_v / sales_v if sales_v else 0.0

        dm_sales         += sales_v
        dm_hours         += hours
        dm_variance      += variance
        dm_goal          += goal_v
        dm_payroll       += payroll_v
        dm_labor_pct_num += payroll_v
        dm_labor_pct_den += sales_v

        row_fill = fill(OVER_RED) if variance > 0 else fill(UNDER_GRN) if variance < 0 else fill(NEUTRAL)
        ws.row_dimensions[row_num].height = 18

        vals   = [row['Store #'], row['Store Name'], row['DM'], row['Rev Band'],
                  goal_v, sales_v, hours, variance, payroll_v, labor_pct]
        fmts   = [None, None, None, None, FMT_HOURS, FMT_CURRENCY, FMT_HOURS, FMT_VARIANCE, FMT_CURRENCY, '0.0%']
        aligns = [center, left, center, center, center, center, center, center, center, center]

        for ci, (val, fmt, aln) in enumerate(zip(vals, fmts, aligns), 1):
            c = ws.cell(row=row_num, column=ci, value=val)
            c.font      = Font(name='Arial', size=10)
            c.fill      = row_fill
            c.alignment = aln
            c.border    = thin_border()
            if fmt:
                c.number_format = fmt
        row_num += 1

    dm_labor_pct = dm_labor_pct_num / dm_labor_pct_den if dm_labor_pct_den else 0.0
    ws.row_dimensions[row_num].height = 20
    subtotal_vals   = ['', f'{dm_name} — Subtotal', '', '', dm_goal, dm_sales, dm_hours, dm_variance, dm_payroll, dm_labor_pct]
    subtotal_fmts   = [None, None, None, None, FMT_HOURS, FMT_CURRENCY, FMT_HOURS, FMT_VARIANCE, FMT_CURRENCY, '0.0%']
    subtotal_aligns = [center, left, center, center, center, center, center, center, center, center]

    for ci, (val, fmt, aln) in enumerate(zip(subtotal_vals, subtotal_fmts, subtotal_aligns), 1):
        c = ws.cell(row=row_num, column=ci, value=val)
        c.font      = Font(name='Arial', size=10, bold=True, color='1F3864')
        c.fill      = fill(SUBTOTAL)
        c.alignment = aln
        c.border    = subtotal_border()
        if fmt:
            c.number_format = fmt
    row_num += 1

    grand_sales    += dm_sales
    grand_hours    += dm_hours
    grand_variance += dm_variance
    grand_goal     += dm_goal
    grand_payroll  += dm_payroll

grand_labor_pct = grand_payroll / grand_sales if grand_sales else 0.0

ws.row_dimensions[row_num].height = 24
gt_vals  = ['', 'GRAND TOTAL', '', '', grand_goal, grand_sales, grand_hours, grand_variance, grand_payroll, grand_labor_pct]
gt_fmts  = [None, None, None, None, FMT_HOURS, FMT_CURRENCY, FMT_HOURS, FMT_VARIANCE, FMT_CURRENCY, '0.0%']
for ci, (val, fmt) in enumerate(zip(gt_vals, gt_fmts), 1):
    c = ws.cell(row=row_num, column=ci, value=val)
    c.font      = Font(name='Arial', size=11, bold=True, color=WHITE)
    c.fill      = fill(DARK_NAVY)
    c.alignment = center if ci != 2 else left
    c.border    = thin_border()
    if fmt:
        c.number_format = fmt

row_num += 2
ws.merge_cells(f'A{row_num}:C{row_num}')
ws[f'A{row_num}'].value = 'Legend'
ws[f'A{row_num}'].font  = Font(name='Arial', size=10, bold=True)
row_num += 1
for hex_c, label in [(OVER_RED,'Over Goal'),(UNDER_GRN,'Under Goal'),(NEUTRAL,'At Goal')]:
    ws.cell(row=row_num, column=1).fill   = fill(hex_c)
    ws.cell(row=row_num, column=1).border = thin_border()
    c = ws.cell(row=row_num, column=2, value=label)
    c.font      = Font(name='Arial', size=10)
    c.alignment = left
    row_num += 1

row_num += 1
ws.merge_cells(f'A{row_num}:J{row_num}')
note = ws[f'A{row_num}']
note.value     = f'* Est. Payroll Expense = Gross Wages (Temp Rate × Hours) × {LOAD_FACTOR:.4f} load factor (incl. {PAYROLL_TAX_RATE*100:.2f}% payroll tax + {WORKERS_COMP_RATE*100:.1f}% workers comp)'
note.font      = Font(name='Arial', size=9, italic=True, color='666666')
note.alignment = left

ws.freeze_panes = 'A3'

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — Store Rankings
# ══════════════════════════════════════════════════════════════════════════════
wr = wb.create_sheet('Store Rankings')
wr.sheet_view.showGridLines = False

rank_headers    = ['Rank','Store #','Store Name','DM','Hourly Goal','Actual Labor Hours','Variance (Hrs)','Est. Labor %']
rank_col_widths = [8, 12, 34, 12, 13, 20, 18, 14]

wr.row_dimensions[1].height = 36
wr.merge_cells('A1:H1')
c = wr['A1']
c.value     = f'AvS Weekly Labor Report - Store Rankings - {REPORT_DATES}'
c.font      = Font(name='Arial', size=16, bold=True, color=WHITE)
c.fill      = fill(DARK_NAVY)
c.alignment = Alignment(horizontal='center', vertical='center')

wr.row_dimensions[2].height = 24
for ci, (hdr, w) in enumerate(zip(rank_headers, rank_col_widths), 1):
    c = wr.cell(row=2, column=ci, value=hdr)
    c.font      = Font(name='Arial', size=11, bold=True, color=WHITE)
    c.fill      = fill(MID_BLUE)
    c.alignment = center
    c.border    = thin_border()
    wr.column_dimensions[get_column_letter(ci)].width = w

ranked = df.copy()
ranked['labor_pct'] = ranked['loaded_payroll'] / ranked['Last Week Net Sales']
ranked = ranked.sort_values('Variance', ascending=True).reset_index(drop=True)

for rank_idx, row in ranked.iterrows():
    rn       = rank_idx + 3
    variance = float(row['Variance'])          if pd.notna(row['Variance'])          else 0.0
    hours    = float(row['actual_labor_hours']) if pd.notna(row['actual_labor_hours']) else 0.0
    goal_v   = float(row['Hourly Goal'])        if pd.notna(row['Hourly Goal'])        else 0.0
    lp       = float(row['labor_pct'])          if pd.notna(row['labor_pct'])          else 0.0

    row_fill = fill(OVER_RED) if variance > 0 else fill(UNDER_GRN) if variance < 0 else fill(NEUTRAL)
    wr.row_dimensions[rn].height = 18

    vals  = [rank_idx+1, row['Store #'], row['Store Name'], row['DM'], goal_v, hours, variance, lp]
    fmts  = [None, None, None, None, FMT_HOURS, FMT_HOURS, FMT_VARIANCE, '0.0%']
    aligns= [center, center, left, center, center, center, center, center]

    for ci, (val, fmt, aln) in enumerate(zip(vals, fmts, aligns), 1):
        c = wr.cell(row=rn, column=ci, value=val)
        c.font      = Font(name='Arial', size=10)
        c.fill      = row_fill
        c.alignment = aln
        c.border    = thin_border()
        if fmt:
            c.number_format = fmt

leg_row = len(ranked) + 5
wr.merge_cells(f'A{leg_row}:C{leg_row}')
wr[f'A{leg_row}'].value = 'Legend'
wr[f'A{leg_row}'].font  = Font(name='Arial', size=10, bold=True)
leg_row += 1
for hex_c, label in [(UNDER_GRN,'Under Goal — Best'),(OVER_RED,'Over Goal — Worst'),(NEUTRAL,'At Goal')]:
    wr.cell(row=leg_row, column=1).fill   = fill(hex_c)
    wr.cell(row=leg_row, column=1).border = thin_border()
    c = wr.cell(row=leg_row, column=2, value=label)
    c.font      = Font(name='Arial', size=10)
    c.alignment = left
    leg_row += 1

wr.freeze_panes = 'A3'

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — DM Rankings
# ══════════════════════════════════════════════════════════════════════════════
dm_summary = []
for dm_name, group in df.groupby('DM', sort=True):
    dm_sales_t   = group['Last Week Net Sales'].fillna(0).sum()
    dm_hours_t   = group['actual_labor_hours'].fillna(0).sum()
    dm_goal_t    = group['Hourly Goal'].fillna(0).sum()
    dm_payroll_t = group['loaded_payroll'].fillna(0).sum()
    dm_var_t     = round(dm_hours_t - dm_goal_t, 2)
    dm_lp        = dm_payroll_t / dm_sales_t if dm_sales_t else 0.0
    dm_summary.append({'DM': dm_name, 'Hourly Goal': round(dm_goal_t,2),
        'Actual Labor Hours': round(dm_hours_t,2), 'Variance': dm_var_t,
        'Est. Payroll Expense': round(dm_payroll_t,2), 'Est. Labor %': dm_lp})

dm_ranked = sorted(dm_summary, key=lambda x: x['Variance'])

wdm = wb.create_sheet('DM Rankings')
wdm.sheet_view.showGridLines = False

dm_rank_headers    = ['Rank','DM','Hourly Goal','Actual Labor Hours','Variance (Hrs)','Est. Payroll Expense','Est. Labor %']
dm_rank_col_widths = [8, 14, 13, 20, 18, 22, 14]

wdm.row_dimensions[1].height = 36
wdm.merge_cells('A1:G1')
c = wdm['A1']
c.value     = f'AvS Weekly Labor Report - DM Rankings - {REPORT_DATES}'
c.font      = Font(name='Arial', size=16, bold=True, color=WHITE)
c.fill      = fill(DARK_NAVY)
c.alignment = Alignment(horizontal='center', vertical='center')

wdm.row_dimensions[2].height = 24
for ci, (hdr, w) in enumerate(zip(dm_rank_headers, dm_rank_col_widths), 1):
    c = wdm.cell(row=2, column=ci, value=hdr)
    c.font      = Font(name='Arial', size=11, bold=True, color=WHITE)
    c.fill      = fill(MID_BLUE)
    c.alignment = center
    c.border    = thin_border()
    wdm.column_dimensions[get_column_letter(ci)].width = w

for rank_idx, dm_row in enumerate(dm_ranked):
    rn       = rank_idx + 3
    variance = dm_row['Variance']
    row_fill = fill(OVER_RED) if variance > 0 else fill(UNDER_GRN) if variance < 0 else fill(NEUTRAL)
    wdm.row_dimensions[rn].height = 20
    vals  = [rank_idx+1, dm_row['DM'], dm_row['Hourly Goal'], dm_row['Actual Labor Hours'],
             variance, dm_row['Est. Payroll Expense'], dm_row['Est. Labor %']]
    fmts  = [None, None, FMT_HOURS, FMT_HOURS, FMT_VARIANCE, FMT_CURRENCY, '0.0%']
    aligns= [center, center, center, center, center, center, center]
    for ci, (val, fmt, aln) in enumerate(zip(vals, fmts, aligns), 1):
        c = wdm.cell(row=rn, column=ci, value=val)
        c.font      = Font(name='Arial', size=10)
        c.fill      = row_fill
        c.alignment = aln
        c.border    = thin_border()
        if fmt:
            c.number_format = fmt

leg_row = len(dm_ranked) + 5
wdm.merge_cells(f'A{leg_row}:C{leg_row}')
wdm[f'A{leg_row}'].value = 'Legend'
wdm[f'A{leg_row}'].font  = Font(name='Arial', size=10, bold=True)
leg_row += 1
for hex_c, label in [(UNDER_GRN,'Under Goal — Best'),(OVER_RED,'Over Goal — Worst'),(NEUTRAL,'At Goal')]:
    wdm.cell(row=leg_row, column=1).fill   = fill(hex_c)
    wdm.cell(row=leg_row, column=1).border = thin_border()
    c = wdm.cell(row=leg_row, column=2, value=label)
    c.font      = Font(name='Arial', size=10)
    c.alignment = left
    leg_row += 1

wdm.freeze_panes = 'A3'

wb.save(OUTPUT_FILE)
print(f'Done — Report saved to {OUTPUT_FILE}')
print(f'Grand Payroll Expense: ${grand_payroll:,.2f}')
