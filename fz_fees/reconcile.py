"""
Franchise Fee Reconciliation Script
-------------------------------------
Usage:
    python reconcile.py --fz fz_fee_schedule.xlsx --bank bank_data.xlsx

    Dates are pulled automatically from the files where possible.
    If a date cannot be detected, the script will prompt you to enter it.

Weekly workflow:
    The FZ fee schedule and bank data always cover DIFFERENT weeks:
      - FZ fee schedule  = fees owed for the PREVIOUS week
      - Bank data        = payments received in the CURRENT week (for the prior week's fees)

    Example:
      FZ schedule week-end 03/11/2026  →  fees owed for that week
      Bank data   dated    03/20/2026  →  payments pulled from bank that day

    Both dates are stamped in the output filename and Summary sheet so every
    file is unambiguous about which period it covers.

Inputs:
    locations.csv           — Master file with all 38 store IDs and names.
                              Store number is the primary key.

    fz_fee_schedule.xlsx    — Provided weekly by the franchisor. One row per store.
                              Week-end date is pulled automatically from this file.
                              Columns used:
                                Store #, Store Name, Fiscal Year, Week, Week End,
                                Days of Sales, Reported Net Sales,
                                Royalty Fee %, Marketing Fee %,
                                Royalty Fee $, Marketing Fee $, Franchise Fee

    bank_data.xlsx          — Provided weekly. Records what was actually paid.
                              *** BANK DATA INTAKE IS SCAFFOLDED — UPDATE ONCE
                              FILE STRUCTURE IS CONFIRMED ***

    --bank-date MMDDYYYY    — Optional. Date to label the bank data (e.g. 03202026).
                              Defaults to today's date if not provided.

Configuration:
    HELPDESK_FEE_WEEKLY     — Fixed help desk fee per store per week ($25.00).
                              Update here if the rate ever changes.

    PAYMENT_TOLERANCE       — Dollar amount within which a payment is considered
                              matching (default $0.01 — penny rounding tolerance).

    RATE_TOLERANCE_PP       — Percentage point deviation that triggers a rate flag
                              (default 0.10pp — ten basis points).

Output:
    reconciliation_FZ-week-end-MMDDYYYY_bank-MMDDYYYY.xlsx
        Sheet 1 — Full Results   : all stores, all fees, all flags
        Sheet 2 — Anomalies Only : stores with at least one flag
        Sheet 3 — Summary        : totals, flag counts, both dates clearly labeled
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration — edit these values as rates change, not the logic below
# ---------------------------------------------------------------------------

LOCATIONS_FILE      = "locations.csv"
HELPDESK_FEE_WEEKLY = 25.00          # Fixed $ per store per week
PAYMENT_TOLERANCE   = 0.01           # Dollar tolerance for payment match
RATE_TOLERANCE_PP   = 0.10           # Percentage point tolerance for rate check

# Columns to read from the FZ fee schedule (matches actual file headers)
FZ_COLUMNS = {
    "store_id"        : "Store #",
    "store_name"      : "Store Name",
    "fiscal_year"     : "Fiscal Year",
    "week_num"        : "Week",
    "week_end"        : "Week End",
    "days_of_sales"   : "Days of Sales",
    "net_sales"       : "Reported Net Sales",
    "royalty_pct"     : "Royalty Fee %",
    "marketing_pct"   : "Marketing Fee %",
    "royalty_fee"     : "Royalty Fee $",
    "marketing_fee"   : "Marketing Fee $",
    "franchise_fee"   : "Franchise Fee",
}

# ---------------------------------------------------------------------------
# Load locations master
# ---------------------------------------------------------------------------

def load_locations(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, sep="|", dtype=str)
    required = {"location_id", "store_name"}
    missing = required - set(df.columns)
    if missing:
        sys.exit(f"ERROR: locations.csv is missing columns: {missing}")
    df["location_id"] = df["location_id"].str.strip().str.upper()
    return df[["location_id", "store_name"]]


# ---------------------------------------------------------------------------
# Load FZ fee schedule
# ---------------------------------------------------------------------------

def load_fz_schedule(path: str) -> tuple:
    """
    Returns (dataframe, week_label).
    week_label format: '2026-W11_week-end-03182026'
    Skips the Total summary row automatically.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("ERROR: FZ fee schedule appears to be empty.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    # Map internal keys to column indices
    col_idx = {}
    for key, col_name in FZ_COLUMNS.items():
        try:
            col_idx[key] = headers.index(col_name)
        except ValueError:
            sys.exit(
                f"ERROR: FZ fee schedule is missing expected column '{col_name}'.\n"
                f"       Columns found: {headers}"
            )

    records = []
    week_end_date = None

    for row in rows[1:]:
        store_id = row[col_idx["store_id"]]
        if store_id is None or str(store_id).strip().lower() == "total":
            continue

        # Pull week-end date from first valid data row
        if week_end_date is None:
            raw_date = row[col_idx["week_end"]]
            if isinstance(raw_date, datetime):
                week_end_date = raw_date
            elif isinstance(raw_date, str):
                try:
                    week_end_date = datetime.strptime(raw_date.strip(), "%Y-%m-%d")
                except ValueError:
                    pass

        records.append({
            "store_id"          : str(store_id).strip().upper(),
            "store_name_fz"     : row[col_idx["store_name"]],
            "fiscal_year"       : row[col_idx["fiscal_year"]],
            "week_num"          : row[col_idx["week_num"]],
            "week_end"          : row[col_idx["week_end"]],
            "days_of_sales"     : row[col_idx["days_of_sales"]],
            "net_sales"         : row[col_idx["net_sales"]],
            "royalty_pct"       : row[col_idx["royalty_pct"]],
            "marketing_pct"     : row[col_idx["marketing_pct"]],
            "royalty_fee_fz"    : row[col_idx["royalty_fee"]],
            "marketing_fee_fz"  : row[col_idx["marketing_fee"]],
            "franchise_fee_fz"  : row[col_idx["franchise_fee"]],
        })

    if not records:
        sys.exit("ERROR: No store data rows found in FZ fee schedule.")

    df = pd.DataFrame(records)

    # Add fixed help desk fee column
    df["helpdesk_fee_fz"] = HELPDESK_FEE_WEEKLY

    if week_end_date is None:
        print("  WARNING: Could not detect week-end date from FZ file.")
        print()
        while True:
            raw = input("  Enter FZ fee schedule week-end date (MMDDYYYY, e.g. 03112026): ").strip()
            try:
                week_end_date = datetime.strptime(raw, "%m%d%Y")
                break
            except ValueError:
                print("  Invalid format — please use MMDDYYYY (e.g. 03112026)")

    print(f"  Stores loaded        : {len(df)}")
    print(f"  Fiscal year          : {int(df['fiscal_year'].iloc[0])}")
    print(f"  Week                 : {int(df['week_num'].iloc[0])}")
    print(f"  Week end             : {week_end_date.strftime('%m/%d/%Y')}")
    print(f"  Help desk fee added  : ${HELPDESK_FEE_WEEKLY:.2f} per store "
          f"(total: ${HELPDESK_FEE_WEEKLY * len(df):,.2f})")

    return df, week_end_date


# ---------------------------------------------------------------------------
# Auto-detect bank date from file
# ---------------------------------------------------------------------------

def detect_bank_date(path: str) -> str | None:
    """
    Attempts to extract the transaction date from the bank file.
    Looks at the date column (index 1) of the first data row.
    Returns a formatted MM/DD/YYYY string, or None if detection fails.
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                raw = row[1]  # Date is in column index 1
                if isinstance(raw, datetime):
                    return raw.strftime("%m/%d/%Y")
                if isinstance(raw, str):
                    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
                        try:
                            return datetime.strptime(raw.strip(), fmt).strftime("%m/%d/%Y")
                        except ValueError:
                            continue
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Load bank data
# *** UPDATE THIS FUNCTION once bank file structure is confirmed ***
# ---------------------------------------------------------------------------

def load_bank_data(path: str) -> pd.DataFrame:
    """
    Reads actual payment data from the bank export.

    Bank file format: no column headers. Raw transaction rows where:
      - Column 7 (index): description string containing fee type and store number
                          e.g. "Freddy's LLC     Roy Wk 10 Stillwater (OK) 112-0001 ACH Debit"
      - Column 8 (index): dollar amount of the transaction

    Fee type is identified by keywords in the description:
      "Roy"    -> royalty fee
      "Mkt"    -> marketing fee
      "Help D" -> help desk fee

    Store number (e.g. 112-0038) is extracted from the description via regex.

    Each store's transactions are pivoted into a single row with columns:
      store_id, royalty_paid, marketing_paid, helpdesk_paid, franchise_paid
      (franchise_paid = royalty_paid + marketing_paid, to match FZ "Franchise Fee" total)

    Flags printed at load time:
      - Stores with duplicate transactions for the same fee type
      - Stores missing one or more fee types
    """
    import re

    DESCRIPTION_COL = 7   # 0-indexed column containing the transaction description
    AMOUNT_COL      = 8   # 0-indexed column containing the dollar amount

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        sys.exit(f"ERROR: Could not read bank data file: {e}")

    STORE_ID_PATTERN = re.compile(r"(1\d{2}-\d{4})")

    records = []
    skipped = 0
    for i, row in enumerate(raw_rows, 1):
        desc   = row[DESCRIPTION_COL] if len(row) > DESCRIPTION_COL else None
        amount = row[AMOUNT_COL]      if len(row) > AMOUNT_COL      else None

        if not desc or amount is None:
            skipped += 1
            continue

        desc_str = str(desc).strip()

        # Extract store ID
        match = STORE_ID_PATTERN.search(desc_str)
        if not match:
            print(f"  WARNING: Could not extract store ID from row {i}: {desc_str!r}")
            skipped += 1
            continue

        store_id = match.group(1).upper()

        # Identify fee type
        desc_upper = desc_str.upper()
        if "HELP D" in desc_upper:
            fee_type = "helpdesk"
        elif "ROY" in desc_upper:
            fee_type = "royalty"
        elif "MKT" in desc_upper:
            fee_type = "marketing"
        else:
            print(f"  WARNING: Unknown fee type in row {i}: {desc_str!r}")
            skipped += 1
            continue

        try:
            amount_val = float(amount)
        except (ValueError, TypeError):
            print(f"  WARNING: Non-numeric amount in row {i}: {amount!r}")
            skipped += 1
            continue

        records.append({
            "store_id" : store_id,
            "fee_type" : fee_type,
            "amount"   : amount_val,
        })

    if not records:
        sys.exit("ERROR: No valid transactions found in bank data file.")

    df_raw = pd.DataFrame(records)

    # Check for duplicate fee entries per store (e.g. two royalty rows for same store)
    dupes = df_raw.groupby(["store_id", "fee_type"]).size().reset_index(name="count")
    dupes = dupes[dupes["count"] > 1]
    if not dupes.empty:
        for _, d in dupes.iterrows():
            print(f"  WARNING: Duplicate {d['fee_type']} transaction for {d['store_id']} "
                  f"({int(d['count'])} rows) — amounts will be summed.")

    # Pivot: one row per store, one column per fee type
    df_pivot = df_raw.groupby(["store_id", "fee_type"])["amount"].sum().unstack(fill_value=None)
    df_pivot = df_pivot.reset_index()

    # Ensure all three fee columns exist even if no transactions of that type were found
    for col in ["royalty", "marketing", "helpdesk"]:
        if col not in df_pivot.columns:
            df_pivot[col] = None
            print(f"  WARNING: No {col} transactions found in bank file.")

    df_pivot = df_pivot.rename(columns={
        "royalty"   : "royalty_paid",
        "marketing" : "marketing_paid",
        "helpdesk"  : "helpdesk_paid",
    })

    # Franchise paid = royalty + marketing (matches FZ "Franchise Fee" combined total)
    df_pivot["franchise_paid"] = (
        pd.to_numeric(df_pivot["royalty_paid"],   errors="coerce").fillna(0) +
        pd.to_numeric(df_pivot["marketing_paid"], errors="coerce").fillna(0)
    )
    # Set franchise_paid to None if both components are missing
    both_missing = df_pivot["royalty_paid"].isna() & df_pivot["marketing_paid"].isna()
    df_pivot.loc[both_missing, "franchise_paid"] = None

    print(f"  Bank transactions    : {len(records)} rows parsed ({skipped} skipped)")
    print(f"  Unique stores found  : {df_pivot['store_id'].nunique()}")

    # Flag stores missing any fee type
    for fee, col in [("royalty", "royalty_paid"), ("marketing", "marketing_paid"),
                     ("help desk", "helpdesk_paid")]:
        missing_stores = df_pivot[df_pivot[col].isna()]["store_id"].tolist()
        if missing_stores:
            print(f"  WARNING: No {fee} payment found for: {missing_stores}")

    return df_pivot[["store_id", "royalty_paid", "marketing_paid",
                     "franchise_paid", "helpdesk_paid"]]


# ---------------------------------------------------------------------------
# Reconciliation logic
# ---------------------------------------------------------------------------

def reconcile(locations: pd.DataFrame, fz: pd.DataFrame,
              bank: pd.DataFrame) -> pd.DataFrame:

    # Merge FZ schedule with master locations
    df = fz.merge(
        locations,
        left_on="store_id",
        right_on="location_id",
        how="outer",
        indicator=True
    )

    fz_not_in_master = df[df["_merge"] == "left_only"]["store_id"].tolist()
    master_not_in_fz = df[df["_merge"] == "right_only"]["location_id"].tolist()
    if fz_not_in_master:
        print(f"  WARNING: In FZ file but not in master: {fz_not_in_master}")
    if master_not_in_fz:
        print(f"  WARNING: In master but missing from FZ: {master_not_in_fz}")

    df = df[df["_merge"] != "right_only"].copy()
    df.drop(columns=["_merge", "location_id"], inplace=True)

    # Merge bank data
    df = df.merge(bank, on="store_id", how="left")

    # Store 112-9001 historically does not receive a help desk ACH debit.
    # Rather than hardcoding $0, we check what the bank actually paid:
    #   - If bank paid $0 or nothing → set owed to $0 (no flag raised)
    #   - If bank paid any amount   → keep owed at $25 so a match is expected
    mask_9001 = df["store_id"] == "112-9001"
    hd_paid_9001 = pd.to_numeric(
        df.loc[mask_9001, "helpdesk_paid"], errors="coerce"
    ).fillna(0).values
    if mask_9001.any() and (hd_paid_9001 == 0).all():
        df.loc[mask_9001, "helpdesk_fee_fz"] = 0.00

    rows = []
    for _, r in df.iterrows():
        flags = []

        def check_payment(label, paid_col, owed_val):
            """Returns variance float or None; appends to flags if mismatch or missing."""
            paid = r.get(paid_col)
            if pd.isna(paid) and pd.notna(owed_val) and float(owed_val) > 0:
                # Fee was owed but no payment found in bank data at all
                flags.append(
                    f"{label}: NO PAYMENT found in bank data — owed ${float(owed_val):,.2f}"
                )
                return None
            if pd.notna(paid) and pd.notna(owed_val):
                variance = round(float(paid) - float(owed_val), 2)
                if abs(variance) > PAYMENT_TOLERANCE:
                    flags.append(
                        f"{label}: paid ${float(paid):,.2f} vs owed "
                        f"${float(owed_val):,.2f} (variance ${variance:+,.2f})"
                    )
                return variance
            return None

        royalty_var   = check_payment("Royalty",   "royalty_paid",   r.get("royalty_fee_fz"))
        marketing_var = check_payment("Marketing", "marketing_paid", r.get("marketing_fee_fz"))
        franchise_var = check_payment("Franchise", "franchise_paid", r.get("franchise_fee_fz"))
        helpdesk_var  = check_payment("Help desk", "helpdesk_paid",  r.get("helpdesk_fee_fz", HELPDESK_FEE_WEEKLY))

        rows.append({
            # Identity
            "Store #"              : r["store_id"],
            "Store Name"           : r.get("store_name", r.get("store_name_fz", "")),
            "Fiscal Year"          : r.get("fiscal_year"),
            "Week"                 : r.get("week_num"),
            "Week End"             : r.get("week_end"),
            "Days of Sales"        : r.get("days_of_sales"),
            # Sales
            "Reported Net Sales"   : r.get("net_sales"),
            # Royalty
            "Royalty Fee %"        : r.get("royalty_pct"),
            "Royalty Fee Owed"     : r.get("royalty_fee_fz"),
            "Royalty Fee Paid"     : r.get("royalty_paid"),
            "Royalty Variance"     : royalty_var,
            # Marketing
            "Marketing Fee %"      : r.get("marketing_pct"),
            "Marketing Fee Owed"   : r.get("marketing_fee_fz"),
            "Marketing Fee Paid"   : r.get("marketing_paid"),
            "Marketing Variance"   : marketing_var,
            # Franchise total
            "Franchise Fee Owed"   : r.get("franchise_fee_fz"),
            "Franchise Fee Paid"   : r.get("franchise_paid"),
            "Franchise Variance"   : franchise_var,
            # Help desk
            "Help Desk Fee Owed"   : r.get("helpdesk_fee_fz", HELPDESK_FEE_WEEKLY),
            "Help Desk Fee Paid"   : r.get("helpdesk_paid"),
            "Help Desk Variance"   : helpdesk_var,
            # Flags
            "Flag Count"           : len(flags),
            "Flag Details"         : " | ".join(flags) if flags else "",
        })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Invoice generation — three separate AP invoice CSVs for accounting import
# ---------------------------------------------------------------------------

def generate_invoices(results: pd.DataFrame, week_end_dt, fiscal_yr, week_num):
    """
    Generates three AP invoice CSV files using the RECONCILED results dataframe,
    so invoice amounts always reflect the same values used in the reconciliation
    report (including the dynamic 112-9001 help desk adjustment).

      1. Royalty   Week XX YYYY  — GL 7140, per-store royalty fees owed
      2. Marketing Week XX YYYY  — GL 7145, per-store marketing fees owed
      3. Help Desk Week XX YYYY  — GL 7150, per-store help desk fees owed
         ($25 per store; $0 for 112-9001 if bank paid $0 that week)

    Format mirrors the accounting system import template exactly:
      Type | Location | Vendor | Number | Date | Gl Date | Amount |
      Payment Terms | Due Date | Comment | Detail Account |
      Detail Amount | Detail Location | Detail Comment

    Detail Amount is formatted as $#,##0.00 (e.g. $1,323.76).
    Due Date = invoice date + 9 days.
    Amount = sum of all detail amounts for that invoice.
    """

    invoice_date = week_end_dt.strftime("%-m/%-d/%Y")
    due_date_dt  = week_end_dt + pd.Timedelta(days=9)
    due_date     = due_date_dt.strftime("%-m/%-d/%Y")
    yr           = int(fiscal_yr)
    wk           = int(week_num)

    # Map to reconciled results columns (these reflect dynamic adjustments)
    # Single combined invoice — all three fee types in one file
    inv_number = f"Royalty Week {wk} {yr}"

    FEE_SECTIONS = [
        ("Royalty Fee Owed",   7140),
        ("Marketing Fee Owed", 7145),
        ("Help Desk Fee Owed", 7150),
    ]

    COLS = [
        "Type", "Location", "Vendor", "Number", "Date", "Gl Date",
        "Amount", "Payment Terms", "Due Date", "Comment",
        "Detail Account", "Detail Amount", "Detail Location", "Detail Comment",
    ]

    # Calculate grand total across all three fee types
    grand_total = 0.0
    for fee_col, _ in FEE_SECTIONS:
        grand_total += results[fee_col].fillna(0).sum()
    grand_total = round(grand_total, 2)

    rows = []
    for fee_col, gl_acct in FEE_SECTIONS:
        for _, r in results.iterrows():
            amt = float(r.get(fee_col, 0) or 0)
            rows.append({
                "Type"           : "AP Invoice",
                "Location"       : 301,
                "Vendor"         : "Freddy's",
                "Number"         : inv_number,
                "Date"           : invoice_date,
                "Gl Date"        : invoice_date,
                "Amount"         : grand_total,
                "Payment Terms"  : None,
                "Due Date"       : due_date,
                "Comment"        : inv_number,
                "Detail Account" : gl_acct,
                "Detail Amount"  : f"${amt:,.2f}",
                "Detail Location": r["Store #"],
                "Detail Comment" : inv_number,
            })

    df_inv = pd.DataFrame(rows, columns=COLS)
    filename = f"invoice_Week{wk:02d}_{yr}.csv"
    df_inv.to_csv(filename, index=False)

    return [(inv_number, filename, grand_total, len(rows))]


# ---------------------------------------------------------------------------
# Excel report — professional three-tab output
# ---------------------------------------------------------------------------

# ── Palette ────────────────────────────────────────────────────────────────
HDR_DARK   = "1F3864"   # Navy  — title / report-info rows
HDR_MED    = "2E75B6"   # Blue  — section / column headers
HDR_LIGHT  = "BDD7EE"   # Light blue — sub-headers
ROW_ALT    = "F5F8FC"   # Very light blue — alternating data rows
TOTAL_FILL = "D9E1F2"   # Soft indigo — totals row
FLAG_RED   = "FFC7CE"   # Red   — variance mismatch / flag
FLAG_YLW   = "FFEB9C"   # Amber — missing payment
FLAG_GRN   = "C6EFCE"   # Green — zero variance (clean)
WHITE      = "FFFFFF"

# ── Style helpers ──────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def _border(style="thin"):
    s = Side(style=style, color="B8CCE4")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _money():
    return '#,##0.00'

def _pct():
    return '0.000%'

def _int_fmt():
    return '#,##0'

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def _autofit(ws, min_w=8, max_w=40):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value is not None:
                best = max(best, min(len(str(cell.value)) + 3, max_w))
        ws.column_dimensions[col[0].column_letter].width = best

def _style_header_row(ws, row_num, fill_hex, font_color="FFFFFF",
                      bold=True, size=11, h_align="center"):
    for cell in ws[row_num]:
        if cell.value is not None or cell.column <= ws.max_column:
            cell.fill      = _fill(fill_hex)
            cell.font      = _font(bold=bold, color=font_color, size=size)
            cell.alignment = _align(h=h_align, v="center")
            cell.border    = _border()

def _write_title_block(ws, lines, start_row=1):
    """Write a 2-row title block merged across all data columns."""
    for i, (text, fill_hex, font_color, bold, size) in enumerate(lines):
        r = start_row + i
        ws.cell(row=r, column=1, value=text)
        for col in range(1, ws.max_column + 2):
            c = ws.cell(row=r, column=col)
            c.fill      = _fill(fill_hex)
            c.font      = _font(bold=bold, color=font_color, size=size)
            c.alignment = _align(h="left", v="center")
        ws.row_dimensions[r].height = 20 if size >= 13 else 16

def _totals_row(ws, data_start_row, data_end_row, money_cols, label_col=1):
    """Append a totals row with SUM formulas for money columns."""
    tot_row = data_end_row + 1
    ws.cell(row=tot_row, column=label_col, value="TOTAL")
    for c_idx in money_cols:
        col_letter = get_column_letter(c_idx)
        ws.cell(row=tot_row, column=c_idx,
                value=f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})")
        ws.cell(row=tot_row, column=c_idx).number_format = _money()
    for cell in ws[tot_row]:
        cell.fill      = _fill(TOTAL_FILL)
        cell.font      = _font(bold=True, size=11)
        cell.border    = _border()
        cell.alignment = _align(h="right" if cell.column != label_col else "left", v="center")
    ws.row_dimensions[tot_row].height = 18
    return tot_row


# ---------------------------------------------------------------------------
# Tab 1 — FZ Reported
# ---------------------------------------------------------------------------

def build_fz_tab(ws, df, fz_label, fiscal_yr, week_num):
    from openpyxl.utils import get_column_letter

    COLS = [
        ("Store #",           10),
        ("Store Name",        28),
        ("Days of Sales",     12),
        ("Net Sales",         14),
        ("Royalty %",         11),
        ("Royalty $",         13),
        ("Marketing %",       12),
        ("Marketing $",       13),
        ("Franchise Fee",     14),
        ("Help Desk Fee",     13),
    ]
    n_cols = len(COLS)

    # Title block (rows 1-3)
    ws.append(["FZ Fee Schedule — Reported Fees"])
    ws.append([f"Fiscal Year {int(fiscal_yr)}  |  Week {int(week_num)}  |  Week End: {fz_label}"])
    ws.append([""])  # spacer

    _write_title_block(ws, [
        ("FZ Fee Schedule — Reported Fees", HDR_DARK, "FFFFFF", True, 14),
        (f"Fiscal Year {int(fiscal_yr)}  |  Week {int(week_num)}  |  Week End: {fz_label}",
         HDR_MED, "FFFFFF", False, 11),
        ("", HDR_LIGHT, "FFFFFF", False, 10),
    ], start_row=1)

    # Column headers (row 4)
    hdr_row = 4
    for c_idx, (label, _) in enumerate(COLS, 1):
        cell = ws.cell(row=hdr_row, column=c_idx, value=label)
        cell.fill      = _fill(HDR_MED)
        cell.font      = _font(bold=True, color="FFFFFF", size=11)
        cell.alignment = _align(h="center", v="center")
        cell.border    = _border()
    ws.row_dimensions[hdr_row].height = 20

    # Data rows
    data_start = 5
    money_cols = []
    for r_idx, (_, row) in enumerate(df.iterrows(), data_start):
        fill_hex = ROW_ALT if (r_idx - data_start) % 2 == 1 else WHITE
        vals = [
            row.get("store_id", ""),
            row.get("store_name", row.get("store_name_fz", "")),
            row.get("days_of_sales"),
            row.get("net_sales"),
            row.get("royalty_pct"),
            row.get("royalty_fee_fz"),
            row.get("marketing_pct"),
            row.get("marketing_fee_fz"),
            row.get("franchise_fee_fz"),
            row.get("Help Desk Fee Owed", HELPDESK_FEE_WEEKLY),
        ]
        fmts = [None, None, _int_fmt(), _money(), _pct(), _money(),
                _pct(), _money(), _money(), _money()]
        money_cols = [4, 6, 8, 9, 10]

        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = _fill(fill_hex)
            cell.font      = _font(size=11)
            cell.border    = _border()
            if fmt:
                cell.number_format = fmt
                cell.alignment = _align(h="right", v="center")
            else:
                cell.alignment = _align(h="left" if c_idx <= 2 else "center", v="center")
        ws.row_dimensions[r_idx].height = 16

    data_end = data_start + len(df) - 1

    # Totals row
    _totals_row(ws, data_start, data_end, money_cols)

    # Column widths
    for c_idx, (_, width) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    ws.freeze_panes = f"A{data_start}"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6


# ---------------------------------------------------------------------------
# Tab 2 — Bank Reported
# ---------------------------------------------------------------------------

def build_bank_tab(ws, df, bank_date):
    from openpyxl.utils import get_column_letter

    COLS = [
        ("Store #",           10),
        ("Store Name",        28),
        ("Royalty Paid",      13),
        ("Marketing Paid",    14),
        ("Franchise Paid",    14),
        ("Help Desk Paid",    13),
        ("Total Paid",        13),
    ]
    money_cols = [3, 4, 5, 6, 7]

    # Title block
    _write_title_block(ws, [
        ("Bank Data — Payments Received", HDR_DARK, "FFFFFF", True, 14),
        (f"Bank Date: {bank_date}", HDR_MED, "FFFFFF", False, 11),
        ("", HDR_LIGHT, "FFFFFF", False, 10),
    ], start_row=1)

    # Column headers
    hdr_row = 4
    for c_idx, (label, _) in enumerate(COLS, 1):
        cell = ws.cell(row=hdr_row, column=c_idx, value=label)
        cell.fill      = _fill(HDR_MED)
        cell.font      = _font(bold=True, color="FFFFFF", size=11)
        cell.alignment = _align(h="center", v="center")
        cell.border    = _border()
    ws.row_dimensions[hdr_row].height = 20

    # Data rows
    data_start = 5
    for r_idx, (_, row) in enumerate(df.iterrows(), data_start):
        fill_hex   = ROW_ALT if (r_idx - data_start) % 2 == 1 else WHITE
        roy_paid   = row.get("Royalty Fee Paid")
        mkt_paid   = row.get("Marketing Fee Paid")
        hd_paid    = row.get("Help Desk Fee Paid")
        fran_paid  = row.get("Franchise Fee Paid")

        # Total paid = royalty + marketing + helpdesk
        nums = [v for v in [roy_paid, mkt_paid, hd_paid] if pd.notna(v)]
        total_paid = sum(nums) if nums else None

        vals = [
            row.get("Store #", ""),
            row.get("Store Name", ""),
            roy_paid,
            mkt_paid,
            fran_paid,
            hd_paid,
            total_paid,
        ]
        fmts = [None, None, _money(), _money(), _money(), _money(), _money()]

        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = _fill(fill_hex)
            cell.font      = _font(size=11)
            cell.border    = _border()
            if fmt:
                cell.number_format = fmt
                cell.alignment = _align(h="right", v="center")
            else:
                cell.alignment = _align(h="left" if c_idx <= 2 else "center", v="center")
            # Amber if no payment data
            if fmt and val is None:
                cell.fill = _fill(FLAG_YLW)
        ws.row_dimensions[r_idx].height = 16

    data_end = data_start + len(df) - 1
    _totals_row(ws, data_start, data_end, money_cols)

    for c_idx, (_, width) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    ws.freeze_panes = f"A{data_start}"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6


# ---------------------------------------------------------------------------
# Tab 3 — Summary
# ---------------------------------------------------------------------------

def build_summary_detailed_tab(ws, df, fz_label, bank_date, fiscal_yr, week_num):
    from openpyxl.utils import get_column_letter

    # Group headers + sub-columns
    # Layout:
    # A           B            C          D       E        F        G       H        I        J        K        L         M         N         O       P         Q         R        S         T        U
    # Store #  Store Name  Net Sales | Roy%  Roy Owed  Roy Paid  Roy Var | Mkt%  Mkt Owed  Mkt Paid  Mkt Var | Fran Owed  Fran Paid  Fran Var | HD Owed  HD Paid  HD Var | Total Owed  Total Paid  Total Var | Flags

    SECTION_GROUPS = [
        ("",              ["Store #", "Store Name", "Net Sales"]),
        ("ROYALTY",       ["Rate %", "Owed", "Paid", "Variance", "% of Sales"]),
        ("MARKETING",     ["Rate %", "Owed", "Paid", "Variance", "% of Sales"]),
        ("FRANCHISE FEE", ["Owed", "Paid", "Variance"]),
        ("HELP DESK",     ["Owed", "Paid", "Variance"]),
        ("TOTALS",        ["Total Owed", "Total Paid", "Total Variance"]),
        ("",              ["Flags"]),
    ]

    # Build flat column list and section spans
    flat_cols  = []
    spans      = []  # (label, start_col, end_col)
    c = 1
    for section_label, sub_cols in SECTION_GROUPS:
        start = c
        for sc in sub_cols:
            flat_cols.append((section_label, sc))
            c += 1
        spans.append((section_label, start, c - 1))

    # Title rows 1-3
    _write_title_block(ws, [
        ("Weekly Fee Reconciliation — Summary (Detailed)", HDR_DARK, "FFFFFF", True, 14),
        (f"FZ Week End: {fz_label}  |  Bank Date: {bank_date}  |  "
         f"Fiscal Year {int(fiscal_yr)}  Week {int(week_num)}",
         HDR_MED, "FFFFFF", False, 11),
        ("", HDR_LIGHT, "FFFFFF", False, 10),
    ], start_row=1)

    # Section group headers row 4
    SECTION_COLORS = {
        "ROYALTY":       "1F497D",
        "MARKETING":     "375623",
        "FRANCHISE FEE": "7030A0",
        "HELP DESK":     "833C00",
        "TOTALS":        "1F3864",
        "":              HDR_MED,
    }
    for label, start_col, end_col in spans:
        if not label:
            continue
        color = SECTION_COLORS.get(label, HDR_MED)
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=4, column=col)
            cell.value     = label if col == start_col else ""
            cell.fill      = _fill(color)
            cell.font      = _font(bold=True, color="FFFFFF", size=11)
            cell.alignment = _align(h="center", v="center")
            cell.border    = _border()
        ws.row_dimensions[4].height = 18

    # Sub-column headers row 5
    for c_idx, (_, sub) in enumerate(flat_cols, 1):
        cell = ws.cell(row=5, column=c_idx, value=sub)
        cell.fill      = _fill(HDR_LIGHT)
        cell.font      = _font(bold=True, color="1F3864", size=10)
        cell.alignment = _align(h="center", v="center", wrap=True)
        cell.border    = _border()
    ws.row_dimensions[5].height = 30

    # Build column index lookup
    def col_of(sub_name, section=None):
        for i, (sec, sub) in enumerate(flat_cols):
            if sub == sub_name and (section is None or sec == section):
                return i + 1
        return None

    # Data rows start at 6
    data_start = 6
    money_cols_summary = []
    pct_cols_summary   = []

    for r_idx, (_, row) in enumerate(df.iterrows(), data_start):
        fill_hex   = ROW_ALT if (r_idx - data_start) % 2 == 1 else WHITE
        net_sales  = row.get("Reported Net Sales")
        roy_owed   = row.get("Royalty Fee Owed")
        roy_paid   = row.get("Royalty Fee Paid")
        roy_var    = row.get("Royalty Variance")
        mkt_owed   = row.get("Marketing Fee Owed")
        mkt_paid   = row.get("Marketing Fee Paid")
        mkt_var    = row.get("Marketing Variance")
        fran_owed  = row.get("Franchise Fee Owed")
        fran_paid  = row.get("Franchise Fee Paid")
        fran_var   = row.get("Franchise Variance")
        hd_owed    = row.get("Help Desk Fee Owed")
        hd_paid    = row.get("Help Desk Fee Paid")
        hd_var     = row.get("Help Desk Variance")
        flag_count = row.get("Flag Count", 0)
        flag_text  = row.get("Flag Details", "")

        # % of sales
        roy_pct_actual = (roy_paid / net_sales) if (roy_paid and net_sales) else None
        mkt_pct_actual = (mkt_paid / net_sales) if (mkt_paid and net_sales) else None

        nums_owed = [v for v in [roy_owed, mkt_owed, hd_owed] if pd.notna(v)]
        nums_paid = [v for v in [roy_paid, mkt_paid, hd_paid] if pd.notna(v)]
        total_owed = sum(nums_owed) if nums_owed else None
        total_paid = sum(nums_paid) if nums_paid else None
        total_var  = round(total_paid - total_owed, 2) if (total_paid is not None and total_owed is not None) else None

        data = {
            ("", "Store #")              : row.get("Store #", ""),
            ("", "Store Name")           : row.get("Store Name", ""),
            ("", "Net Sales")            : net_sales,
            ("ROYALTY", "Rate %")        : row.get("Royalty Fee %"),
            ("ROYALTY", "Owed")          : roy_owed,
            ("ROYALTY", "Paid")          : roy_paid,
            ("ROYALTY", "Variance")      : roy_var,
            ("ROYALTY", "% of Sales")    : roy_pct_actual,
            ("MARKETING", "Rate %")      : row.get("Marketing Fee %"),
            ("MARKETING", "Owed")        : mkt_owed,
            ("MARKETING", "Paid")        : mkt_paid,
            ("MARKETING", "Variance")    : mkt_var,
            ("MARKETING", "% of Sales")  : mkt_pct_actual,
            ("FRANCHISE FEE", "Owed")    : fran_owed,
            ("FRANCHISE FEE", "Paid")    : fran_paid,
            ("FRANCHISE FEE", "Variance"): fran_var,
            ("HELP DESK", "Owed")        : hd_owed,
            ("HELP DESK", "Paid")        : hd_paid,
            ("HELP DESK", "Variance")    : hd_var,
            ("TOTALS", "Total Owed")     : total_owed,
            ("TOTALS", "Total Paid")     : total_paid,
            ("TOTALS", "Total Variance") : total_var,
            ("", "Flags")                : flag_text,
        }

        for c_idx, (sec, sub) in enumerate(flat_cols, 1):
            val  = data.get((sec, sub))
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill   = _fill(fill_hex)
            cell.font   = _font(size=10)
            cell.border = _border()

            is_money = sub in ("Owed", "Paid", "Total Owed", "Total Paid", "Net Sales")
            is_var   = "Variance" in sub
            is_pct   = sub in ("Rate %", "% of Sales")
            is_flag  = sub == "Flags"

            if is_money:
                cell.number_format = _money()
                cell.alignment = _align(h="right", v="center")
                if c_idx not in money_cols_summary:
                    money_cols_summary.append(c_idx)
            elif is_var:
                cell.number_format = _money()
                cell.alignment = _align(h="right", v="center")
                # Color-code variance
                if val is None and pd.notna(data.get((sec, "Owed"))):
                    cell.fill = _fill(FLAG_YLW)   # missing payment
                elif val is not None and abs(val) > PAYMENT_TOLERANCE:
                    cell.fill = _fill(FLAG_RED)    # mismatch
                elif val is not None:
                    cell.fill = _fill(FLAG_GRN)    # clean
            elif is_pct:
                cell.number_format = _pct()
                cell.alignment = _align(h="center", v="center")
                if c_idx not in pct_cols_summary:
                    pct_cols_summary.append(c_idx)
            elif is_flag:
                cell.alignment = _align(h="left", v="center", wrap=True)
                if flag_count and int(flag_count) > 0:
                    cell.fill = _fill(FLAG_RED)
                    cell.font = _font(size=10, color="9C0006")
            else:
                cell.alignment = _align(
                    h="left" if sub in ("Store #", "Store Name") else "center",
                    v="center"
                )

        ws.row_dimensions[r_idx].height = 16

    data_end = data_start + len(df) - 1

    # Totals row
    tot_row = data_end + 1
    ws.cell(row=tot_row, column=1, value="TOTAL")
    for c_idx in money_cols_summary:
        col_letter = get_column_letter(c_idx)
        cell = ws.cell(row=tot_row, column=c_idx,
                       value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})")
        cell.number_format = _money()
    for cell in ws[tot_row]:
        cell.fill      = _fill(TOTAL_FILL)
        cell.font      = _font(bold=True, size=11)
        cell.border    = _border()
        cell.alignment = _align(h="right" if cell.column > 2 else "left", v="center")
    ws.row_dimensions[tot_row].height = 20

    # Column widths
    COL_WIDTHS = {
        "Store #": 11, "Store Name": 28, "Net Sales": 13,
        "Rate %": 9, "Owed": 12, "Paid": 12, "Variance": 12, "% of Sales": 11,
        "Total Owed": 13, "Total Paid": 13, "Total Variance": 14,
        "Flags": 50,
    }
    for c_idx, (_, sub) in enumerate(flat_cols, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = COL_WIDTHS.get(sub, 12)

    ws.freeze_panes = f"C{data_start}"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 30


# ---------------------------------------------------------------------------
# Tab — Summary (high-level totals)
# ---------------------------------------------------------------------------

def build_summary_tab(ws, df, fz_label, bank_date, fiscal_yr, week_num):
    """
    High-level one-page summary showing totals for each fee type:
    FZ Charged | Bank Paid | Variance — for Royalty, Marketing, Franchise, Help Desk, and Grand Total.
    Also shows store count, net sales, and flag summary.
    """
    from openpyxl.utils import get_column_letter

    fz_label_str  = fz_label
    flagged_count = int((df["Flag Count"] > 0).sum())
    store_count   = len(df)
    clean_count   = store_count - flagged_count

    def v(col):
        return df[col].sum() if df[col].notna().any() else 0.0

    roy_owed  = v("Royalty Fee Owed")
    roy_paid  = v("Royalty Fee Paid")
    mkt_owed  = v("Marketing Fee Owed")
    mkt_paid  = v("Marketing Fee Paid")
    fran_owed = v("Franchise Fee Owed")
    fran_paid = v("Franchise Fee Paid")
    hd_owed   = v("Help Desk Fee Owed")
    hd_paid   = v("Help Desk Fee Paid")
    net_sales = v("Reported Net Sales")

    grand_owed = roy_owed + mkt_owed + hd_owed
    grand_paid = roy_paid + mkt_paid + hd_paid

    has_bank   = df["Royalty Fee Paid"].notna().any()

    # ── Title block ──────────────────────────────────────────────────────────
    _write_title_block(ws, [
        ("Weekly Fee Reconciliation — Summary", HDR_DARK, "FFFFFF", True, 14),
        (f"FZ Week End: {fz_label_str}  |  Bank Date: {bank_date}  |  "
         f"Fiscal Year {int(fiscal_yr)}  Week {int(week_num)}",
         HDR_MED, "FFFFFF", False, 11),
        ("", HDR_LIGHT, "FFFFFF", False, 10),
    ], start_row=1)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    # ── KPI block (rows 4-8): store counts + net sales ──────────────────────
    kpi_items = [
        ("Stores Processed",      store_count,  None),
        ("Stores Clean",          clean_count,  None),
        ("Stores Flagged",        flagged_count, None),
        ("Total Reported Net Sales", net_sales, _money()),
    ]
    for i, (label, val, fmt) in enumerate(kpi_items):
        r = 4 + i
        lbl_cell = ws.cell(row=r, column=1, value=label)
        lbl_cell.font      = _font(bold=True, size=11, color="1F3864")
        lbl_cell.fill      = _fill(ROW_ALT if i % 2 else WHITE)
        lbl_cell.alignment = _align(h="left", v="center")
        lbl_cell.border    = _border()

        val_cell = ws.cell(row=r, column=2, value=val)
        val_cell.font      = _font(bold=True, size=11)
        val_cell.fill      = _fill(ROW_ALT if i % 2 else WHITE)
        val_cell.alignment = _align(h="right", v="center")
        val_cell.border    = _border()
        if fmt:
            val_cell.number_format = fmt
        ws.row_dimensions[r].height = 18

    # ── Spacer row ───────────────────────────────────────────────────────────
    ws.row_dimensions[8].height = 10

    # ── Fee table ────────────────────────────────────────────────────────────
    # Section header row 9
    HDR_ROW = 9
    section_cols = [
        (1, "Fee Type",       16),
        (2, "FZ Charged",     16),
        (3, "Bank Paid",      16),
        (4, "Variance",       16),
        (5, "% of Net Sales", 16),
    ]
    SECTION_FILL = "1F3864"
    for col, label, width in section_cols:
        cell = ws.cell(row=HDR_ROW, column=col, value=label)
        cell.fill      = _fill(SECTION_FILL)
        cell.font      = _font(bold=True, color="FFFFFF", size=11)
        cell.alignment = _align(h="center" if col > 1 else "left", v="center")
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[HDR_ROW].height = 22

    # Fee rows
    SECTION_COLORS_FEE = {
        "Royalty Fee":    ("1F497D", "DEEAF1"),
        "Marketing Fee":  ("375623", "EAF1DD"),
        "Franchise Fee":  ("7030A0", "EAD1F5"),
        "Help Desk Fee":  ("833C00", "FCE4D6"),
        "GRAND TOTAL":    (HDR_DARK, TOTAL_FILL),
    }

    fee_rows = [
        ("Royalty Fee",   roy_owed,  roy_paid,  net_sales),
        ("Marketing Fee", mkt_owed,  mkt_paid,  net_sales),
        ("Franchise Fee", fran_owed, fran_paid, net_sales),
        ("Help Desk Fee", hd_owed,   hd_paid,   None),
        ("GRAND TOTAL",   grand_owed, grand_paid, net_sales),
    ]

    DATA_START = HDR_ROW + 1
    for i, (label, owed, paid, sales_base) in enumerate(fee_rows):
        r = DATA_START + i
        variance  = round(paid - owed, 2) if has_bank else None
        pct_sales = (paid / sales_base)   if (has_bank and sales_base and label not in ("Help Desk Fee", "GRAND TOTAL")) else None

        dark_hex, light_hex = SECTION_COLORS_FEE[label]
        is_total = label == "GRAND TOTAL"

        row_fill = HDR_DARK if is_total else light_hex
        lbl_font_color = "FFFFFF" if is_total else dark_hex

        # Label cell
        lbl = ws.cell(row=r, column=1, value=label)
        lbl.fill      = _fill(row_fill)
        lbl.font      = _font(bold=is_total, color=lbl_font_color, size=11)
        lbl.alignment = _align(h="left", v="center")
        lbl.border    = _border()

        # FZ Charged
        c_owed = ws.cell(row=r, column=2, value=owed)
        c_owed.number_format = _money()
        c_owed.fill      = _fill(row_fill)
        c_owed.font      = _font(bold=is_total, color="FFFFFF" if is_total else "000000", size=11)
        c_owed.alignment = _align(h="right", v="center")
        c_owed.border    = _border()

        # Bank Paid
        c_paid = ws.cell(row=r, column=3, value=paid if has_bank else None)
        c_paid.number_format = _money()
        c_paid.fill      = _fill(row_fill)
        c_paid.font      = _font(bold=is_total, color="FFFFFF" if is_total else "000000", size=11)
        c_paid.alignment = _align(h="right", v="center")
        c_paid.border    = _border()
        if not has_bank:
            c_paid.value = "awaiting bank data"
            c_paid.number_format = "General"

        # Variance
        c_var = ws.cell(row=r, column=4, value=variance)
        c_var.number_format = _money()
        c_var.alignment = _align(h="right", v="center")
        c_var.border    = _border()
        if is_total:
            c_var.fill = _fill(HDR_DARK)
            c_var.font = _font(bold=True, color="FFFFFF", size=11)
        elif variance is None:
            c_var.fill = _fill(light_hex)
            c_var.font = _font(size=11)
        elif abs(variance) <= PAYMENT_TOLERANCE:
            c_var.fill = _fill(FLAG_GRN)
            c_var.font = _font(size=11, color="375623")
        else:
            c_var.fill = _fill(FLAG_RED)
            c_var.font = _font(bold=True, size=11, color="9C0006")

        # % of Net Sales
        c_pct = ws.cell(row=r, column=5, value=pct_sales)
        c_pct.number_format = _pct()
        c_pct.alignment = _align(h="center", v="center")
        c_pct.border    = _border()
        if is_total or label == "Help Desk Fee":
            c_pct.value = "—"
            c_pct.number_format = "General"
            c_pct.fill = _fill(HDR_DARK if is_total else light_hex)
            c_pct.font = _font(bold=is_total, color="FFFFFF" if is_total else dark_hex, size=11)
        else:
            c_pct.fill = _fill(light_hex)
            c_pct.font = _font(size=11)

        ws.row_dimensions[r].height = 20

    # ── Flag detail block ─────────────────────────────────────────────────────
    flagged_df = df[df["Flag Count"] > 0]
    if not flagged_df.empty:
        flag_section_row = DATA_START + len(fee_rows) + 2
        ws.row_dimensions[flag_section_row - 1].height = 10

        hdr = ws.cell(row=flag_section_row, column=1, value="Flagged Stores")
        hdr.fill      = _fill("9C0006")
        hdr.font      = _font(bold=True, color="FFFFFF", size=11)
        hdr.alignment = _align(h="left", v="center")
        hdr.border    = _border()
        for col in range(2, 6):
            c = ws.cell(row=flag_section_row, column=col)
            c.fill   = _fill("9C0006")
            c.border = _border()
        ws.row_dimensions[flag_section_row].height = 18

        for j, (_, frow) in enumerate(flagged_df.iterrows()):
            r = flag_section_row + 1 + j
            fill_hex = ROW_ALT if j % 2 else WHITE

            store_cell = ws.cell(row=r, column=1,
                                  value=f"{frow['Store #']}  {frow['Store Name']}")
            store_cell.font      = _font(bold=True, size=10)
            store_cell.fill      = _fill(fill_hex)
            store_cell.alignment = _align(h="left", v="center")
            store_cell.border    = _border()

            detail_cell = ws.cell(row=r, column=2, value=frow["Flag Details"])
            detail_cell.font      = _font(size=10, color="9C0006")
            detail_cell.fill      = _fill(FLAG_RED)
            detail_cell.alignment = _align(h="left", v="center", wrap=True)
            detail_cell.border    = _border()
            # Merge detail across columns 2-5
            ws.merge_cells(start_row=r, start_column=2,
                           end_row=r,   end_column=5)
            ws.row_dimensions[r].height = 30

    ws.freeze_panes = "A4"



# ---------------------------------------------------------------------------
# Write report — orchestrates all three tabs
# ---------------------------------------------------------------------------

def write_report(df: pd.DataFrame, fz_week_end_dt: object, bank_date_str: str):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    fz_str        = fz_week_end_dt.strftime("%m%d%Y") if fz_week_end_dt else "unknown"
    bank_file_str = bank_date_str.replace("/", "")
    filename      = f"reconciliation_FZ-week-end-{fz_str}_bank-{bank_file_str}.xlsx"
    fz_label      = fz_week_end_dt.strftime("%m/%d/%Y") if fz_week_end_dt else "unknown"

    fiscal_yr = df["Fiscal Year"].iloc[0] if "Fiscal Year" in df.columns else ""
    week_num  = df["Week"].iloc[0]        if "Week"        in df.columns else ""

    flagged = df[df["Flag Count"] > 0]

    wb = Workbook()

    # Tab 1 — Summary (high-level)
    ws_top = wb.active
    ws_top.title = "Summary"

    # Tab 2 — Summary Detailed
    ws_sum = wb.create_sheet("Summary_Detailed")

    # Tab 3 — FZ Reported
    ws_fz = wb.create_sheet("FZ Reported")

    # Reconstruct FZ-only dataframe from reconciled results
    fz_df = df[[
        "Store #", "Store Name", "Fiscal Year", "Week", "Week End",
        "Days of Sales", "Reported Net Sales",
        "Royalty Fee %", "Royalty Fee Owed",
        "Marketing Fee %", "Marketing Fee Owed",
        "Franchise Fee Owed", "Help Desk Fee Owed",
    ]].copy()
    fz_df = fz_df.rename(columns={
        "Royalty Fee Owed"  : "royalty_fee_fz",
        "Marketing Fee Owed": "marketing_fee_fz",
        "Franchise Fee Owed": "franchise_fee_fz",
        "Royalty Fee %"     : "royalty_pct",
        "Marketing Fee %"   : "marketing_pct",
        "Days of Sales"     : "days_of_sales",
        "Reported Net Sales": "net_sales",
        "Store #"           : "store_id",
        "Store Name"        : "store_name",
    })
    build_fz_tab(ws_fz, fz_df, fz_label, fiscal_yr, week_num)

    # Tab 4 — Bank Reported
    ws_bank = wb.create_sheet("Bank Reported")
    build_bank_tab(ws_bank, df, bank_date_str)

    build_summary_tab(ws_top, df, fz_label, bank_date_str, fiscal_yr, week_num)
    build_summary_detailed_tab(ws_sum, df, fz_label, bank_date_str, fiscal_yr, week_num)

    wb.save(filename)

    print(f"\n  Output file          : {filename}")
    print(f"  FZ week end          : {fz_label}")
    print(f"  Bank data date       : {bank_date_str}")
    print(f"  Total stores         : {len(df)}")
    print(f"  Flagged stores       : {len(flagged)}")
    if not flagged.empty:
        print(f"\n  Flagged stores:")
        for _, row in flagged.iterrows():
            print(f"    [{row['Store #']}] {row['Store Name']}: {row['Flag Details']}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Franchise Fee Reconciliation")
    parser.add_argument("--fz",        required=True,  help="Path to FZ fee schedule Excel file")
    parser.add_argument("--bank",      required=False, help="Path to bank data Excel file (optional)")
    parser.add_argument("--bank-date", required=False, dest="bank_date",
                        help="Date label for bank data in MMDDYYYY format (auto-prompted if not provided and cannot be detected)")
    args = parser.parse_args()

    for path, label in [(LOCATIONS_FILE, "locations master"),
                        (args.fz, "FZ fee schedule")]:
        if not Path(path).exists():
            sys.exit(f"ERROR: {label} not found: {path}")
    if args.bank and not Path(args.bank).exists():
        sys.exit(f"ERROR: Bank data file not found: {args.bank}")

    print(f"\n{'='*55}")
    print(f"  Franchise Fee Reconciliation")
    print(f"{'='*55}")

    print(f"\n[1/4] Loading locations master...")
    locations = load_locations(LOCATIONS_FILE)
    print(f"  Stores in master     : {len(locations)}")

    print(f"\n[2/4] Loading FZ fee schedule...")
    fz, fz_week_end_dt = load_fz_schedule(args.fz)

    # Determine bank date label — try CLI arg, then auto-detect from file, then prompt
    bank_date_str = None

    if args.bank_date:
        try:
            bank_date_dt  = datetime.strptime(args.bank_date, "%m%d%Y")
            bank_date_str = bank_date_dt.strftime("%m/%d/%Y")
            print(f"  Bank date (provided) : {bank_date_str}")
        except ValueError:
            sys.exit("ERROR: --bank-date must be in MMDDYYYY format, e.g. 03202026")

    elif args.bank:
        # Try to auto-detect date from the bank file
        bank_date_str = detect_bank_date(args.bank)
        if bank_date_str:
            print(f"  Bank date (detected) : {bank_date_str}")

    if not bank_date_str:
        # Prompt the user
        print()
        while True:
            raw = input("  Enter bank data date (MMDDYYYY, e.g. 03202026): ").strip()
            try:
                bank_date_dt  = datetime.strptime(raw, "%m%d%Y")
                bank_date_str = bank_date_dt.strftime("%m/%d/%Y")
                break
            except ValueError:
                print("  Invalid format — please use MMDDYYYY (e.g. 03202026)")

    if args.bank:
        print(f"\n[3/4] Loading bank data...")
        print(f"  Bank data date       : {bank_date_str}")
        bank = load_bank_data(args.bank)
    else:
        print(f"\n[3/4] Bank data not provided — payment columns will show as blank.")
        print(f"      Re-run with --bank <file> to complete the reconciliation.")
        bank = pd.DataFrame(columns=[
            "store_id", "royalty_paid", "marketing_paid",
            "franchise_paid", "helpdesk_paid"
        ])

    print(f"\n[4/4] Running reconciliation...")
    results = reconcile(locations, fz, bank)
    write_report(results, fz_week_end_dt, bank_date_str)

    print(f"\n[5/5] Generating accounting invoices...")
    fiscal_yr = results["Fiscal Year"].iloc[0]
    week_num  = results["Week"].iloc[0]
    invoices  = generate_invoices(results, fz_week_end_dt, fiscal_yr, week_num)
    for inv_number, filename, total, n_rows in invoices:
        print(f"  {filename}  ({n_rows} detail rows, grand total ${total:,.2f})")
    print(f"\n{'='*55}\n")


if __name__ == "__main__":
    main()
