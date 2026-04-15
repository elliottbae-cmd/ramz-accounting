"""
Ram-Z Accounting Toolbox — Streamlit App
-----------------------------------------
Run with:
    cd C:\\Users\\BretElliott\\ramz-accounting
    streamlit run fz_fees/app.py
"""

import os
import sys
import json
from pathlib import Path
from datetime import datetime, date, timedelta
from io import BytesIO, StringIO

import pandas as pd
import streamlit as st

# ---------------------------------------------------------------------------
# Import paths
# ---------------------------------------------------------------------------
_FZ_DIR = Path(__file__).parent
_LABOR_DIR = _FZ_DIR.parent / "labor"
sys.path.insert(0, str(_FZ_DIR))
sys.path.insert(0, str(_LABOR_DIR))

from reconcile import (
    load_locations, load_fz_schedule, detect_bank_date,
    load_bank_data, reconcile, generate_invoices, write_report,
)
from avs_engine import (
    generate_weekly_report, generate_midweek_report, generate_archive_excel,
)
from weekly_lock import (
    get_week_start, get_next_week_start, format_week_label,
    ensure_current_week_locked, override_locked_value,
)
from supabase_db import (
    load_stores, save_store, delete_store,
    load_reference_data, load_band_goals, load_dm_list,
    save_reference_data_row, save_reference_data_bulk, delete_reference_data, set_store_active,
    save_band_goals, add_dm, remove_dm as db_remove_dm,
    load_all_locks, delete_week_lock, log_change,
    load_locked_config, lock_exists, create_lock, get_locked_weeks,
    load_change_log, is_admin, load_admin_users, add_admin, remove_admin,
    save_weekly_actuals, load_weekly_actuals, delete_weekly_actuals,
    draft_exists, load_draft_config, save_draft_bands, lock_drafts,
    get_week_status,
    load_all_submissions, approve_submission, reject_submission,
    load_email_log, load_app_settings, save_app_setting,
)


# ---------------------------------------------------------------------------
# Revenue band options (for dropdowns)
# ---------------------------------------------------------------------------
BAND_OPTIONS = [
    "<25k", "25k-30k", "30k-35k", "35k-40k", "40k-45k",
    "45k-50k", "50k+", "NRO Seasoned", "NRO",
]

# Hours variance threshold for red/green coloring in performance tables
VARIANCE_GOAL_THRESHOLD = 30

# Days of week used in Email Settings dropdowns
DAYS_OF_WEEK = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

# Revenue band dollar ranges (min, max) — None means no numeric range (NRO stores)
BAND_RANGES = {
    "<25k":         (0,      25_000),
    "25k-30k":      (25_000, 30_000),
    "30k-35k":      (30_000, 35_000),
    "35k-40k":      (35_000, 40_000),
    "40k-45k":      (40_000, 45_000),
    "45k-50k":      (45_000, 50_000),
    "50k+":         (50_000, float("inf")),
    "NRO Seasoned": None,
    "NRO":          None,
}


def _band_classify(band, actual):
    """Return (result, variance) for a band + actual sales pair.
    variance is positive (Over) or negative (Under) or 0 (On Target).
    """
    ranges = BAND_RANGES.get(band)
    if ranges is None or actual is None:
        return "N/A", None
    band_min, band_max = ranges
    if actual < band_min:
        return "Under", actual - band_min          # negative
    elif band_max != float("inf") and actual > band_max:
        return "Over", actual - band_max           # positive
    else:
        return "On Target", 0.0


def _fmt_variance(result, variance):
    """Format variance as '+$2,400 (Over)' / '-$1,800 (Under)' / 'On Target'."""
    if result == "N/A":
        return "N/A"
    if result == "On Target":
        return "On Target"
    if variance is None:
        return result
    sign = "+" if variance >= 0 else ""
    return f"{sign}${variance:,.0f} ({result})"

# ---------------------------------------------------------------------------
# Cached data loaders (avoid re-reading on every Streamlit rerun)
# ---------------------------------------------------------------------------
@st.cache_data(ttl=60)
def _cached_reference_data():
    return load_reference_data()

@st.cache_data(ttl=60)
def _cached_band_goals():
    return load_band_goals()

@st.cache_data(ttl=60)
def _cached_dm_list():
    return load_dm_list()

@st.cache_data(ttl=60)
def _cached_all_locks():
    return load_all_locks()

@st.cache_data(ttl=60)
def _cached_weekly_actuals():
    return load_weekly_actuals()

def _paginate_table(table_name, fields="*", order_col="week_start"):
    """Helper to paginate a Supabase table and return all rows."""
    from supabase_db import get_supabase
    sb = get_supabase()
    all_rows = []
    page_size = 1000
    offset = 0
    while True:
        resp = sb.table(table_name).select(fields).order(
            order_col, desc=True
        ).range(offset, offset + page_size - 1).execute()
        if not resp.data:
            break
        all_rows.extend(resp.data)
        if len(resp.data) < page_size:
            break
        offset += page_size
    return all_rows

@st.cache_data(ttl=300)
def _cached_sos_data():
    """Load all SoS weekly data (uploads with rank). Paginates to get all rows."""
    try:
        rows = _paginate_table("store_sos_weekly")
        return pd.DataFrame(rows) if rows else pd.DataFrame()
    except Exception:
        return pd.DataFrame()

@st.cache_data(ttl=300)
def _cached_votg_data():
    """Load all VOTG weekly data (uploads with rank). Paginates to get all rows."""
    try:
        rows = _paginate_table("store_votg_weekly")
        return pd.DataFrame(rows) if rows else pd.DataFrame()
    except Exception:
        return pd.DataFrame()

@st.cache_data(ttl=1800)
def _cached_tattle_reviews_light():
    """Load Tattle reviews WITHOUT snapshots (lighter query for Tattle Insights page).
    TTL=30min — reviews don't change frequently."""
    from supabase_db import get_supabase
    sb = get_supabase()
    try:
        all_rows = []
        page_size = 1000
        offset = 0
        fields = (
            "id,location_id,location_label,score,cer,"
            "experienced_time,completed_time,day_part_label,channel_label,"
            "comment"
        )
        while True:
            resp = sb.table("tattle_reviews").select(fields).order(
                "experienced_time", desc=True
            ).range(offset, offset + page_size - 1).execute()
            if not resp.data:
                break
            all_rows.extend(resp.data)
            if len(resp.data) < page_size:
                break
            offset += page_size
        return pd.DataFrame(all_rows) if all_rows else pd.DataFrame()
    except Exception:
        return pd.DataFrame()

@st.cache_data(ttl=1800)
def _cached_tattle_with_snapshots():
    """Load Tattle reviews WITH snapshots (heavier query, only for category analysis).
    TTL=30min."""
    from supabase_db import get_supabase
    sb = get_supabase()
    try:
        all_rows = []
        page_size = 1000
        offset = 0
        fields = "id,location_id,score,experienced_time,snapshots"
        while True:
            resp = sb.table("tattle_reviews").select(fields).order(
                "experienced_time", desc=True
            ).not_.is_("snapshots", "null").range(
                offset, offset + page_size - 1
            ).execute()
            if not resp.data:
                break
            all_rows.extend(resp.data)
            if len(resp.data) < page_size:
                break
            offset += page_size
        return pd.DataFrame(all_rows) if all_rows else pd.DataFrame()
    except Exception:
        return pd.DataFrame()

@st.cache_data(ttl=1800)
def _cached_tattle_scored():
    """Load only sentiment-scored reviews (for Sentiment Dashboard).
    TTL=30min."""
    from supabase_db import get_supabase
    sb = get_supabase()
    try:
        all_rows = []
        page_size = 1000
        offset = 0
        fields = (
            "id,location_id,score,experienced_time,comment,"
            "sentiment_themes,sentiment_summary"
        )
        while True:
            resp = sb.table("tattle_reviews").select(fields).order(
                "experienced_time", desc=True
            ).not_.is_("sentiment_themes", "null").range(
                offset, offset + page_size - 1
            ).execute()
            if not resp.data:
                break
            all_rows.extend(resp.data)
            if len(resp.data) < page_size:
                break
            offset += page_size
        return pd.DataFrame(all_rows) if all_rows else pd.DataFrame()
    except Exception:
        return pd.DataFrame()

@st.cache_data(ttl=300)
def _cached_store_sales():
    """Load store sales data (last 52 weeks) aggregated to weekly totals.
    TTL=5min — heavier table, changes less often."""
    from supabase_db import get_supabase
    sb = get_supabase()
    cutoff = str(date.today() - timedelta(weeks=52))
    try:
        resp = sb.table("store_sales").select(
            "location_id,sale_date,net_sales"
        ).gte("sale_date", cutoff).execute()
        if not resp.data:
            return pd.DataFrame()
        df = pd.DataFrame(resp.data)
        # Aggregate daily rows → Thu-anchored weekly totals
        df["sale_date"] = pd.to_datetime(df["sale_date"])
        df["net_sales"] = pd.to_numeric(df["net_sales"], errors="coerce").fillna(0)
        df["days_since_thu"] = (df["sale_date"].dt.weekday - 3) % 7
        df["week_start"] = (df["sale_date"] - pd.to_timedelta(df["days_since_thu"], unit="D")).dt.date.astype(str)
        weekly = df.groupby(["location_id", "week_start"], as_index=False)["net_sales"].sum()
        return weekly
    except Exception:
        return pd.DataFrame()

# ---------------------------------------------------------------------------
# Store name fuzzy matching (for SoS / VOTG file uploads)
# ---------------------------------------------------------------------------
def _match_store_name(file_name: str, ref_df: pd.DataFrame):
    """Return location_id for a store name from an uploaded file.

    Tries (in order): exact match → normalized match → difflib fuzzy match.
    Returns None if no match found above 0.6 similarity.
    """
    import re, difflib

    file_name = str(file_name).strip()

    exact = ref_df[ref_df["store_name"] == file_name]
    if not exact.empty:
        return exact.iloc[0]["location_id"]

    def _norm(s):
        s = re.sub(r"\([^)]+\)", "", str(s))   # strip (STATE)
        s = re.sub(r"[^a-z0-9\s]", " ", s.lower())
        return " ".join(s.split())

    file_norm = _norm(file_name)
    for _, row in ref_df.iterrows():
        if _norm(row["store_name"]) == file_norm:
            return row["location_id"]

    names = ref_df["store_name"].tolist()
    close = difflib.get_close_matches(file_name, names, n=1, cutoff=0.6)
    if close:
        hit = ref_df[ref_df["store_name"] == close[0]]
        if not hit.empty:
            return hit.iloc[0]["location_id"]

    return None


def _get_upload_status(week_start_str: str):
    """Return weekly_data_status row for a given week, or empty dict."""
    from supabase_db import get_supabase
    sb = get_supabase()
    try:
        resp = sb.table("weekly_data_status").select("*").eq("week_start", week_start_str).execute()
        return resp.data[0] if resp.data else {}
    except Exception:
        return {}


def _mark_upload_status(week_start_str: str, **flags):
    """Upsert weekly_data_status flags (e.g. sos_uploaded=True)."""
    from supabase_db import get_supabase
    sb = get_supabase()
    payload = {"week_start": week_start_str, **flags}
    sb.table("weekly_data_status").upsert(payload, on_conflict="week_start").execute()


# ---------------------------------------------------------------------------
# Shared UI Helpers
# ---------------------------------------------------------------------------
def _period_filter(locked_weeks, key_prefix):
    """Render a period filter (Current Week/All/Month/Quarter/Year) and return filtered week list."""
    col_period, col_period_val = st.columns(2)
    with col_period:
        period_filter = st.selectbox(
            "View by", ["Current Week", "All Weeks", "Month", "Quarter", "Year"],
            key=f"{key_prefix}_period",
        )
    with col_period_val:
        if period_filter == "Current Week":
            # Most recent completed week in the dataset
            last_week = max(locked_weeks) if locked_weeks else None
            filtered_weeks = [last_week] if last_week else []
            if last_week:
                st.caption(f"Week of {last_week}")
        elif period_filter == "Month":
            months_available = sorted(set((w.year, w.month) for w in locked_weeks))
            month_labels = [f"{y}-{m:02d}" for y, m in months_available]
            selected_month = st.selectbox("Select Month", month_labels, key=f"{key_prefix}_month")
            sel_year, sel_month = int(selected_month[:4]), int(selected_month[5:])
            filtered_weeks = [w for w in locked_weeks if w.year == sel_year and w.month == sel_month]
        elif period_filter == "Quarter":
            quarters_available = sorted(set((w.year, (w.month - 1) // 3 + 1) for w in locked_weeks))
            quarter_labels = [f"{y} Q{q}" for y, q in quarters_available]
            selected_quarter = st.selectbox("Select Quarter", quarter_labels, key=f"{key_prefix}_quarter")
            sel_year = int(selected_quarter[:4])
            sel_q = int(selected_quarter[-1])
            q_months = [(sel_q - 1) * 3 + 1, (sel_q - 1) * 3 + 2, (sel_q - 1) * 3 + 3]
            filtered_weeks = [w for w in locked_weeks if w.year == sel_year and w.month in q_months]
        elif period_filter == "Year":
            years_available = sorted(set(w.year for w in locked_weeks))
            selected_year = st.selectbox("Select Year", years_available, key=f"{key_prefix}_year")
            filtered_weeks = [w for w in locked_weeks if w.year == selected_year]
        else:
            filtered_weeks = locked_weeks
    return filtered_weeks


def _rename_week_cols(week_cols):
    """Build a {ISO_date: 'Wk M/D'} rename map for pivot table week columns."""
    col_map = {}
    for col in week_cols:
        try:
            d = date.fromisoformat(col)
            end_d = d + timedelta(days=6)
            col_map[col] = f"Wk {end_d.month}/{end_d.day}"
        except (ValueError, TypeError):
            pass
    return col_map


def _color_variance_cells(row, week_col_map, has_actuals_weeks, skip_cols):
    """Apply red/green background to variance cells based on VARIANCE_GOAL_THRESHOLD.
    Total column is colored based on its own value regardless of actuals weeks."""
    styles = [""] * len(row)
    for i, col in enumerate(row.index):
        if col in skip_cols:
            continue
        if col == "Total":
            v = row[col]
            if v > VARIANCE_GOAL_THRESHOLD:
                styles[i] = "background-color: #FADBD8"
            elif v < -VARIANCE_GOAL_THRESHOLD:
                styles[i] = "background-color: #D5F5E3"
            continue
        orig_week = next((k for k, label in week_col_map.items() if label == col), None)
        if orig_week not in has_actuals_weeks:
            styles[i] = ""
        elif abs(row[col]) > VARIANCE_GOAL_THRESHOLD:
            styles[i] = "background-color: #FADBD8"
        else:
            styles[i] = "background-color: #D5F5E3"
    return styles


# ---------------------------------------------------------------------------
# In-App Report Rendering Helpers
# ---------------------------------------------------------------------------
def _render_weekly_preview(df):
    """Render the AVS Weekly Report as a styled in-app table."""
    st.subheader("AvS Summary")

    display_rows = []
    for dm_name, group in df.groupby("DM", sort=True):
        for _, row in group.iterrows():
            hours = float(row["actual_hours"]) if pd.notna(row["actual_hours"]) else 0.0
            goal_v = float(row["Hourly Goal"]) if pd.notna(row["Hourly Goal"]) else 0.0
            variance = float(row["Variance"]) if pd.notna(row["Variance"]) else 0.0
            sales_v = float(row.get("Last Week Net Sales", 0)) if pd.notna(row.get("Last Week Net Sales")) else 0.0
            payroll_v = float(row.get("loaded_payroll", 0)) if pd.notna(row.get("loaded_payroll")) else 0.0
            labor_pct = payroll_v / sales_v if sales_v else 0.0

            display_rows.append({
                "Store": row["Store Name"],
                "DM": row["DM"],
                "Rev Band": row["Rev Band"],
                "Hourly Goal": round(goal_v),
                "Net Sales": round(sales_v, 2),
                "Actual Hours": round(hours, 2),
                "Variance": round(variance),
                "Est. Payroll": round(payroll_v, 2),
                "Est. Labor %": round(labor_pct * 100, 1),
                "_abs_var": abs(variance),
            })

    preview_df = pd.DataFrame(display_rows)
    preview_df = preview_df.sort_values("_abs_var").reset_index(drop=True)
    preview_df.insert(0, "Rank", range(1, len(preview_df) + 1))
    preview_df = preview_df.drop(columns=["_abs_var"])

    def color_weekly(row):
        v = row["Variance"]
        if abs(v) > VARIANCE_GOAL_THRESHOLD:
            color = "#FADBD8"   # red — over threshold (significant deviation)
        else:
            color = "#D5F5E3"   # green — within threshold (on goal)
        return [f"background-color: {color}"] * len(row)

    styled = preview_df.style.apply(color_weekly, axis=1).format({
        "Net Sales": "${:,.2f}",
        "Actual Hours": "{:,.2f}",
        "Est. Payroll": "${:,.2f}",
        "Est. Labor %": "{:.1f}%",
    })

    st.dataframe(
        styled,
        use_container_width=False,
        hide_index=True,
        height=(len(preview_df) + 1) * 35 + 3,
        column_config={
            "Net Sales":    st.column_config.TextColumn("Net Sales",    width=120),
            "Est. Payroll": st.column_config.TextColumn("Est. Payroll", width=110),
        },
    )

    total_goal = preview_df["Hourly Goal"].sum()
    total_hours = preview_df["Actual Hours"].sum()
    total_variance = preview_df["Variance"].sum()
    total_sales = preview_df["Net Sales"].sum()
    total_payroll = preview_df["Est. Payroll"].sum()
    total_labor = (total_payroll / total_sales * 100) if total_sales else 0

    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("Total Goal", f"{total_goal:,.0f}")
    m2.metric("Actual Hours", f"{total_hours:,.0f}")
    m3.metric("Variance", f"{total_variance:+,.0f}")
    m4.metric("Net Sales", f"${total_sales:,.0f}")
    m5.metric("Est. Payroll", f"${total_payroll:,.0f}")
    m6.metric("Est. Labor %", f"{total_labor:.1f}%")


def _render_midweek_preview(df, through_day, thresholds):
    """Render the Mid-Week Pulse as a styled in-app table."""
    green_min = thresholds["green_min"]
    green_max = thresholds["green_max"]
    red_above = thresholds["red_above"]

    display_rows = []
    for _, row in df.iterrows():
        hours = float(row["actual_hours"]) if pd.notna(row["actual_hours"]) else 0.0
        goal_v = float(row["Hourly Goal"]) if pd.notna(row["Hourly Goal"]) else 0.0
        variance = float(row["Variance"]) if pd.notna(row["Variance"]) else 0.0
        pct_used = hours / goal_v if goal_v > 0 else 0.0

        if pct_used > red_above:
            status = "Over Pacing"
        elif green_min <= pct_used <= green_max:
            status = "On Pace"
        elif pct_used < green_min:
            status = "Under Pacing"
        else:
            status = ""

        display_rows.append({
            "Store": row["Store Name"],
            "DM": row["DM"],
            "Hourly Goal": round(goal_v),
            "Actual Hours": round(hours, 2),
            "Variance": round(variance),
            "% Used": round(pct_used * 100, 1),
            "Status": status,
            "_abs_var": abs(variance),
        })

    preview_df = pd.DataFrame(display_rows)
    preview_df = preview_df.sort_values("_abs_var").reset_index(drop=True)
    preview_df.insert(0, "Rank", range(1, len(preview_df) + 1))
    preview_df = preview_df.drop(columns=["_abs_var"])

    def color_midweek(row):
        pct = row["% Used"] / 100
        if pct > red_above:
            return ["background-color: #ffcccc"] * len(row)
        elif green_min <= pct <= green_max:
            return ["background-color: #ccffcc"] * len(row)
        elif pct < green_min:
            return ["background-color: #ffe0b2"] * len(row)
        return [""] * len(row)

    styled = preview_df.style.apply(color_midweek, axis=1).format({
        "Actual Hours": "{:,.2f}",
        "% Used": "{:.1f}%",
    })

    st.dataframe(styled, use_container_width=False, hide_index=True,
                 height=(len(preview_df) + 1) * 35 + 3)

    total_goal = preview_df["Hourly Goal"].sum()
    total_hours = preview_df["Actual Hours"].sum()
    total_variance = preview_df["Variance"].sum()
    total_pct = (total_hours / total_goal * 100) if total_goal else 0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Goal", f"{total_goal:,.0f}")
    m2.metric("Actual Hours", f"{total_hours:,.0f}")
    m3.metric("Variance", f"{total_variance:+,.0f}")
    m4.metric("% Used", f"{total_pct:.1f}%")


# ---------------------------------------------------------------------------
# Page setup
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Ram-Z Accounting Toolbox", layout="wide")

# ---------------------------------------------------------------------------
# Compact sidebar CSS
# ---------------------------------------------------------------------------
st.markdown("""
<style>
/* --- Ram-Z Brand Colors --- */
:root {
    --ramz-navy: #2B3A4E;
    --ramz-gold: #C49A5C;
    --ramz-gold-light: #F5F0EB;
}

/* Sidebar background */
section[data-testid="stSidebar"] {
    background-color: var(--ramz-navy);
}
section[data-testid="stSidebar"] * {
    color: #FFFFFF !important;
}

/* Tighten sidebar spacing */
section[data-testid="stSidebar"] .block-container { padding-top: 1rem; }
section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] { gap: 0.15rem; }

/* Compact nav buttons — left aligned */
section[data-testid="stSidebar"] button[kind="secondary"] {
    padding: 0.2rem 0.5rem;
    font-size: 0.85rem;
    min-height: 0;
    height: auto;
    line-height: 1.3;
    border: none;
    background: transparent;
    text-align: left !important;
    justify-content: flex-start !important;
}
section[data-testid="stSidebar"] button[kind="secondary"] p,
section[data-testid="stSidebar"] button[kind="secondary"] span,
section[data-testid="stSidebar"] button[kind="secondary"] div {
    text-align: left !important;
    width: 100%;
}
section[data-testid="stSidebar"] button[kind="secondary"]:hover {
    background: rgba(196, 154, 92, 0.25);
}

/* Compact expanders */
section[data-testid="stSidebar"] details {
    margin-top: 0.25rem;
    margin-bottom: 0.25rem;
    background: transparent !important;
    border: none !important;
}
section[data-testid="stSidebar"] details summary {
    padding: 0.3rem 0;
    font-size: 0.85rem;
    background: transparent !important;
}
section[data-testid="stSidebar"] details[open] {
    background: transparent !important;
}
section[data-testid="stSidebar"] details[open] [data-testid="stVerticalBlock"] {
    gap: 0.1rem;
    padding-left: 0.5rem;
}
/* Remove white background from expander container */
section[data-testid="stSidebar"] [data-testid="stExpander"] {
    background: transparent !important;
    border: none !important;
}
section[data-testid="stSidebar"] [data-testid="stExpander"] summary,
section[data-testid="stSidebar"] [data-testid="stExpander"] > div {
    background: transparent !important;
}

/* Sidebar header — gold accent */
section[data-testid="stSidebar"] h2 {
    font-size: 1rem;
    margin-bottom: 0.75rem;
    padding-bottom: 0.25rem;
    border-bottom: 1px solid rgba(196, 154, 92, 0.5) !important;
    color: var(--ramz-gold) !important;
}

/* Main content — accent colors */
.stButton > button[kind="primary"] {
    background-color: var(--ramz-gold);
    border-color: var(--ramz-gold);
    color: white;
}
.stButton > button[kind="primary"]:hover {
    background-color: #B08A4E;
    border-color: #B08A4E;
}
h1 { color: var(--ramz-navy) !important; }
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# User identity (Streamlit Cloud provides this)
# ---------------------------------------------------------------------------
def get_current_user_email():
    """Get the current user's email from Streamlit Cloud auth."""
    try:
        # Streamlit >= 1.37 uses st.context
        if hasattr(st, "context") and hasattr(st.context, "user"):
            user_info = st.context.user
            if user_info and hasattr(user_info, "email") and user_info.email:
                return user_info.email.strip().lower()
        # Older Streamlit uses st.experimental_user
        if hasattr(st, "experimental_user"):
            user_info = st.experimental_user
            if user_info and hasattr(user_info, "email") and user_info.email:
                return user_info.email.strip().lower()
    except Exception:
        pass
    return ""

current_user = get_current_user_email()
# If no auth is configured, check if running locally (allow admin) vs cloud (deny admin)
if current_user:
    user_is_admin = is_admin(current_user)
else:
    # Allow admin access locally (no auth configured) but deny on Streamlit Cloud
    user_is_admin = not bool(os.environ.get("STREAMLIT_SHARING_MODE", ""))

# ---------------------------------------------------------------------------
# Navigation — use session state to avoid sticky radio buttons
# ---------------------------------------------------------------------------
ALL_PAGES = {
    "accounting": [
        "FZ Fee Reconciliation",
    ],
    "labor": [
        "AVS Weekly Report",
        "AVS Mid-Week Pulse",
        "Prior Week's Reports",
        "AVS Performance - Store Level",
        "AVS Performance - DMs",
        "GM Hot Streak",
    ],
    "guest_experience": [
        "SoS/VOTG Trends",
        "Tattle Insights",
        "Sentiment Dashboard",
    ],
    "settings": [
        "Manage Stores",
        "Store Revenue Bands",
        "DM Assignments",
        "Hourly Goals",
    ],
    "admin": [
        "Rev Band Approvals",
        "Compliance Report",
        "Rev Band Report",
        "Weekly Config",
        "Email Settings",
        "Change Log",
        "Admin Users",
        "Upload SoS",
        "Upload VOTG",
        "Sales Forecasts",
        "Sales Scenario Analysis",
    ],
}

if "active_page" not in st.session_state:
    st.session_state["active_page"] = "FZ Fee Reconciliation"
if "active_section" not in st.session_state:
    st.session_state["active_section"] = "accounting"

# Upload key counters — incrementing these resets file uploaders
for _ctr in ("fz_upload_ctr", "weekly_upload_ctr", "mw_upload_ctr", "sos_upload_ctr", "votg_upload_ctr"):
    if _ctr not in st.session_state:
        st.session_state[_ctr] = 0


def set_page(section, page_name):
    st.session_state["active_page"] = page_name
    st.session_state["active_section"] = section


def render_nav_section(header, section_key, use_expander=False):
    """Render a navigation section in the sidebar."""
    pages = ALL_PAGES[section_key]
    if use_expander:
        is_expanded = st.session_state["active_section"] == section_key
        with st.sidebar.expander(header, expanded=is_expanded):
            for p in pages:
                is_active = st.session_state["active_page"] == p
                label = f"**{p}**" if is_active else p
                if st.button(label, key=f"nav_{section_key}_{p}", use_container_width=True):
                    set_page(section_key, p)
                    st.rerun()
    else:
        st.sidebar.header(header)
        for p in pages:
            is_active = st.session_state["active_page"] == p
            label = f"**{p}**" if is_active else p
            if st.sidebar.button(label, key=f"nav_{section_key}_{p}", use_container_width=True):
                set_page(section_key, p)
                st.rerun()


# --- Sidebar logo ---
_LOGO_PATH = _FZ_DIR / "ramz_logo.png"
if _LOGO_PATH.exists():
    st.sidebar.image(str(_LOGO_PATH), width=220)
    st.sidebar.markdown("---")

render_nav_section("Accounting", "accounting", use_expander=True)
render_nav_section("Labor", "labor", use_expander=True)
render_nav_section("Guest Experience", "guest_experience", use_expander=True)
render_nav_section("Settings", "settings", use_expander=True)
if user_is_admin:
    render_nav_section("Admin", "admin", use_expander=True)

page = st.session_state["active_page"]

# ---------------------------------------------------------------------------
# Week deadline banner helper
# ---------------------------------------------------------------------------
def show_week_deadline_banner():
    """Show a banner about the weekly lock schedule on settings pages."""
    today = date.today()
    current_ws = get_week_start(today)
    next_ws = get_next_week_start(today)
    current_label = format_week_label(current_ws)
    next_label = format_week_label(next_ws)

    if lock_exists(current_ws):
        st.info(
            f"The current week ({current_label}) is **locked**. "
            f"Changes saved here will take effect starting **{next_label}**. "
            f"Review settings by Wednesday to lock them for next week."
        )
    else:
        st.warning(
            f"The current week ({current_label}) is **not yet locked**. "
            f"Settings will be locked when the first AVS report runs this week. "
            f"Review and update settings now before running a report."
        )


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: FZ Fee Reconciliation
# ═══════════════════════════════════════════════════════════════════════════════
if page == "FZ Fee Reconciliation":
    st.title("FZ Fee Reconciliation")

    # --- Input Files (main content area) ---
    st.subheader("Input Files")

    loc_df = load_stores()
    if not loc_df.empty:
        st.success(f"Locations master: {len(loc_df)} stores")
        locations_source = "supabase"
    else:
        st.warning("No stores found — please upload a locations CSV or add stores in Settings.")
        loc_upload = st.file_uploader("Locations Master (.csv)", type=["csv"], key="loc")
        locations_source = loc_upload

    _fz_ctr = st.session_state["fz_upload_ctr"]
    col_fz, col_bank = st.columns(2)
    with col_fz:
        fz_file = st.file_uploader("FZ Fee Schedule (.xlsx)", type=["xlsx"],
                                    key=f"fz_{_fz_ctr}",
                                    help="Weekly fee schedule from the franchisor.")
    with col_bank:
        bank_file = st.file_uploader("Bank Data (.xlsx) — optional", type=["xlsx"],
                                      key=f"bank_{_fz_ctr}",
                                      help="Bank ACH transaction export. Leave blank to run FZ-only.")

    detected_bank_date = None
    if bank_file is not None:
        detected_str = detect_bank_date(bank_file)
        bank_file.seek(0)
        if detected_str:
            try:
                detected_bank_date = datetime.strptime(detected_str, "%m/%d/%Y").date()
            except ValueError:
                pass

    bank_date = st.date_input("Bank Date",
                               value=detected_bank_date or datetime.today().date(),
                               help="Date label for the bank data. Auto-detected when possible.")
    if detected_bank_date:
        st.caption(f"Auto-detected from file: {detected_bank_date.strftime('%m/%d/%Y')}")

    st.divider()
    run_btn = st.button("Run Reconciliation", type="primary", use_container_width=True)

    if run_btn:
        if locations_source is None:
            st.error("Please upload a locations.csv file.")
            st.stop()
        if fz_file is None:
            st.error("Please upload the FZ fee schedule.")
            st.stop()

        try:
            with st.status("Running reconciliation...", expanded=True) as status:
                st.write("Loading locations master...")
                if locations_source == "supabase":
                    locations = loc_df[["location_id", "store_name"]].copy()
                    locations["location_id"] = locations["location_id"].str.strip().str.upper()
                else:
                    locations = load_locations(locations_source)

                st.write("Loading FZ fee schedule...")
                fz_df, fz_week_end_dt = load_fz_schedule(fz_file)
                if fz_week_end_dt is None:
                    st.error("Could not detect the week-end date from the FZ file.")
                    st.stop()

                fiscal_yr = fz_df["fiscal_year"].iloc[0]
                week_num = fz_df["week_num"].iloc[0]

                bank_date_str = bank_date.strftime("%m/%d/%Y")
                if bank_file is not None:
                    st.write("Loading bank data...")
                    bank_file.seek(0)
                    bank_df = load_bank_data(bank_file)
                else:
                    st.write("No bank data — payment columns will be blank.")
                    bank_df = pd.DataFrame(columns=[
                        "store_id", "royalty_paid", "marketing_paid",
                        "franchise_paid", "helpdesk_paid",
                    ])

                st.write("Reconciling fees vs payments...")
                results = reconcile(locations, fz_df, bank_df)

                st.write("Generating Excel report...")
                report_buf = BytesIO()
                write_report(results, fz_week_end_dt, bank_date_str, output=report_buf)
                report_buf.seek(0)

                st.write("Generating invoice CSV...")
                invoice_buf = StringIO()
                invoices = generate_invoices(results, fz_week_end_dt, fiscal_yr, week_num, output=invoice_buf)
                invoice_csv = invoice_buf.getvalue()

                status.update(label="Reconciliation complete!", state="complete")

            st.session_state["results"] = results
            st.session_state["report_buf"] = report_buf
            st.session_state["invoice_csv"] = invoice_csv
            st.session_state["fz_week_end_dt"] = fz_week_end_dt
            st.session_state["bank_date_str"] = bank_date_str
            st.session_state["fiscal_yr"] = fiscal_yr
            st.session_state["week_num"] = week_num
            st.session_state["invoices"] = invoices
            # Auto-clear file uploaders
            st.session_state["fz_upload_ctr"] += 1
            st.rerun()

        except ValueError as e:
            st.error(f"Error: {e}")
            st.stop()
        except Exception as e:
            st.error(f"Unexpected error: {e}")
            st.stop()

    if "results" in st.session_state:
        results = st.session_state["results"]
        fz_week_end_dt = st.session_state["fz_week_end_dt"]
        bank_date_str = st.session_state["bank_date_str"]
        week_num = st.session_state["week_num"]

        flagged = results[results["Flag Count"] > 0]
        clean = results[results["Flag Count"] == 0]

        st.subheader(
            f"Week {int(week_num)} — FZ Week End: {fz_week_end_dt.strftime('%m/%d/%Y')}  |  "
            f"Bank Date: {bank_date_str}"
        )
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Stores", len(results))
        col2.metric("Clean", len(clean))
        col3.metric("Flagged", len(flagged),
                     delta=f"-{len(flagged)}" if len(flagged) > 0 else None,
                     delta_color="inverse")
        col4.metric("Total Net Sales", f"${results['Reported Net Sales'].sum():,.2f}")

        st.subheader("Results")
        display_cols = [
            "Store #", "Store Name", "Reported Net Sales",
            "Royalty Fee Owed", "Royalty Fee Paid", "Royalty Variance",
            "Marketing Fee Owed", "Marketing Fee Paid", "Marketing Variance",
            "Franchise Fee Owed", "Franchise Fee Paid", "Franchise Variance",
            "Help Desk Fee Owed", "Help Desk Fee Paid", "Help Desk Variance",
            "Flag Count", "Flag Details",
        ]
        display_df = results[display_cols].copy()

        def highlight_rows(row):
            if row["Flag Count"] > 0:
                return ["background-color: #FFC7CE"] * len(row)
            return ["background-color: #C6EFCE"] * len(row)

        styled = display_df.style.apply(highlight_rows, axis=1)
        styled = styled.format({
            "Reported Net Sales": "${:,.2f}",
            "Royalty Fee Owed": "${:,.2f}", "Royalty Fee Paid": "${:,.2f}", "Royalty Variance": "${:+,.2f}",
            "Marketing Fee Owed": "${:,.2f}", "Marketing Fee Paid": "${:,.2f}", "Marketing Variance": "${:+,.2f}",
            "Franchise Fee Owed": "${:,.2f}", "Franchise Fee Paid": "${:,.2f}", "Franchise Variance": "${:+,.2f}",
            "Help Desk Fee Owed": "${:,.2f}", "Help Desk Fee Paid": "${:,.2f}", "Help Desk Variance": "${:+,.2f}",
        }, na_rep="—")
        st.dataframe(styled, use_container_width=True, height=600)

        if not flagged.empty:
            with st.expander(f"Flagged Stores ({len(flagged)})", expanded=True):
                for _, row in flagged.iterrows():
                    st.markdown(f"**{row['Store #']}** {row['Store Name']}  \n_{row['Flag Details']}_")

        st.subheader("Downloads")
        fz_str = fz_week_end_dt.strftime("%m%d%Y")
        bank_str = bank_date_str.replace("/", "")
        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            st.download_button(label="Download Excel Report",
                               data=st.session_state["report_buf"],
                               file_name=f"reconciliation_FZ-week-end-{fz_str}_bank-{bank_str}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        with dl_col2:
            _inv_fname = st.session_state["invoices"][0][1] if st.session_state.get("invoices") else "invoices.csv"
            st.download_button(label="Download Invoice CSV",
                               data=st.session_state["invoice_csv"],
                               file_name=_inv_fname, mime="text/csv", use_container_width=True)



# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: AVS Weekly Report
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "AVS Weekly Report":
    st.title("AVS Weekly Labor Report")
    st.caption("Full weekly report with AvS Summary, Store Rankings, and DM Rankings.")

    _wk_default_start = get_week_start()
    _wk_default_end = _wk_default_start + timedelta(days=6)
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Report Start Date", value=_wk_default_start, key="weekly_start")
    with col2:
        end_date = st.date_input("Report End Date", value=_wk_default_end, key="weekly_end")

    report_dates = f"{start_date.month}.{start_date.day}.{start_date.strftime('%y')} - {end_date.month}.{end_date.day}.{end_date.strftime('%y')}"

    _wk_ctr = st.session_state["weekly_upload_ctr"]
    st.divider()
    adp_file = st.file_uploader("ADP Payroll CSV", type=["csv"], key=f"weekly_adp_{_wk_ctr}",
                                 help="Upload the ADP payroll export for this period.")
    sales_file = st.file_uploader("End of Week Net Sales (.xlsx)", type=["xlsx"], key=f"weekly_sales_{_wk_ctr}",
                                   help="Upload the weekly net sales file.")

    st.divider()
    if st.button("Generate Report", type="primary", use_container_width=True, key="weekly_run"):
        if adp_file is None:
            st.error("Please upload the ADP Payroll CSV.")
            st.stop()
        if sales_file is None:
            st.error("Please upload the Net Sales file.")
            st.stop()

        try:
            with st.spinner("Generating weekly report..."):
                # Use locked config for the report week
                ref_data = _cached_reference_data()
                band_goals = _cached_band_goals()
                locked = ensure_current_week_locked(ref_data, band_goals, start_date)
                # Build ref_data and band_goals from locked values
                locked_band_goals = dict(zip(locked["revenue_band"], locked["hourly_goal"]))
                buf, report_df = generate_weekly_report(adp_file, sales_file, locked, locked_band_goals, report_dates)
                # Save actual hours to Supabase for performance pages
                week_start = get_week_start(start_date)
                actuals_for_db = report_df.rename(columns={
                    "Store #": "location_id", "Variance": "variance", "Hourly Goal": "hourly_goal",
                }).copy()
                if "Last Week Net Sales" in actuals_for_db.columns:
                    actuals_for_db["net_sales"] = actuals_for_db["Last Week Net Sales"].fillna(0)
                if "loaded_payroll" in actuals_for_db.columns and "Last Week Net Sales" in actuals_for_db.columns:
                    actuals_for_db["labor_pct"] = (
                        actuals_for_db["loaded_payroll"] / actuals_for_db["Last Week Net Sales"]
                    ).fillna(0)
                save_weekly_actuals(week_start, actuals_for_db)
                _mark_upload_status(str(week_start), sales_uploaded=True)
            st.cache_data.clear()
            # Store results in session state so they persist after rerun
            st.session_state["weekly_report_buf"] = buf
            st.session_state["weekly_report_df"] = report_df
            st.session_state["weekly_report_fname"] = f"AVS_Labor_Report_{start_date.strftime('%m%d%Y')}.xlsx"
            st.session_state["weekly_report_week"] = format_week_label(get_week_start(start_date))
            # Auto-clear file uploaders
            st.session_state["weekly_upload_ctr"] += 1
            st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")

    if "weekly_report_buf" in st.session_state:
        st.success("Report generated!")
        st.caption(f"Used locked config for week: {st.session_state['weekly_report_week']}")

        # --- In-App Preview ---
        if "weekly_report_df" in st.session_state:
            _render_weekly_preview(st.session_state["weekly_report_df"])

        st.divider()
        st.download_button("Export to Excel",
                           data=st.session_state["weekly_report_buf"],
                           file_name=st.session_state["weekly_report_fname"],
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)



# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Prior Week's Reports
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Prior Week's Reports":
    st.title("📂 Prior Week's Reports")
    st.caption("View previously generated weekly reports from stored data.")

    actuals = _cached_weekly_actuals()
    if not actuals.empty:
        available_weeks = sorted(actuals["week_start"].unique(), reverse=True)
        week_labels = {}
        for ws_str in available_weeks:
            try:
                d = date.fromisoformat(ws_str)
                end_d = d + timedelta(days=6)
                week_labels[ws_str] = f"Thu {d.month}/{d.day} – Wed {end_d.month}/{end_d.day}"
            except (ValueError, TypeError):
                week_labels[ws_str] = ws_str

        selected_archive_week = st.selectbox(
            "Select a past week", available_weeks,
            format_func=lambda x: week_labels.get(x, x), key="archive_week_select",
        )

        if st.button("Load Report", key="load_archive", type="primary", use_container_width=True):
            all_locks = _cached_all_locks()
            week_locks = all_locks[all_locks["week_start"] == selected_archive_week]
            week_actuals = actuals[actuals["week_start"] == selected_archive_week]

            if week_locks.empty:
                st.warning("No lock data found for this week.")
            elif week_actuals.empty:
                st.warning("No actuals data found for this week.")
            else:
                merged = week_locks.merge(week_actuals, on="location_id", how="left", suffixes=("", "_act"))
                archive_df = pd.DataFrame({
                    "Store #": merged["location_id"],
                    "Store Name": merged["store_name"],
                    "DM": merged["dm"],
                    "Rev Band": merged["revenue_band"],
                    "Hourly Goal": pd.to_numeric(merged["hourly_goal"], errors="coerce").fillna(0),
                    "actual_hours": pd.to_numeric(merged["actual_hours"], errors="coerce").fillna(0),
                    "Variance": pd.to_numeric(merged["variance"], errors="coerce").fillna(0),
                    "Last Week Net Sales": pd.to_numeric(merged["net_sales"], errors="coerce").fillna(0),
                })
                labor_pct = pd.to_numeric(merged.get("labor_pct", 0), errors="coerce").fillna(0)
                archive_df["loaded_payroll"] = labor_pct * archive_df["Last Week Net Sales"]

                # Exclude stores with no activity — matches the weekly report engine filter (AND logic)
                archive_df = archive_df[
                    (archive_df["actual_hours"] > 0) & (archive_df["Last Week Net Sales"] > 0)
                ].reset_index(drop=True)

                st.session_state["archive_df"] = archive_df
                st.session_state["archive_week_label"] = week_labels.get(selected_archive_week, selected_archive_week)

        if "archive_df" in st.session_state:
            st.markdown(f"**{st.session_state['archive_week_label']}**")
            _render_weekly_preview(st.session_state["archive_df"])

            archive_label = st.session_state["archive_week_label"]
            archive_buf, _ = generate_archive_excel(
                st.session_state["archive_df"], archive_label
            )
            st.download_button(
                "📥 Export to Excel",
                data=archive_buf,
                file_name=f"AVS_Archive_{archive_label.replace(' ', '_').replace('–', '-')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    else:
        st.info("No archived reports yet. Generate weekly reports to start building the archive.")


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: AVS Mid-Week Pulse
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "AVS Mid-Week Pulse":
    st.title("AVS Mid-Week Labor Pulse")
    st.caption("Cumulative hours vs. weekly goal with day-specific color thresholds.")

    DAY_OPTIONS = ["Friday", "Saturday", "Sunday", "Monday", "Tuesday", "Wednesday"]

    _mw_default_start = get_week_start()
    _mw_default_end = _mw_default_start + timedelta(days=6)
    col1, col2, col3 = st.columns(3)
    with col1:
        start_date = st.date_input("Report Start Date", value=_mw_default_start, key="mw_start")
    with col2:
        end_date = st.date_input("Report End Date", value=_mw_default_end, key="mw_end")
    with col3:
        through_day = st.selectbox("Data Through (Day)", DAY_OPTIONS, key="mw_day",
                                    help="Select the last day of data in this upload. "
                                         "Color thresholds adjust based on the day.")

    report_dates = f"{start_date.month}.{start_date.day}.{start_date.strftime('%y')} - {end_date.month}.{end_date.day}.{end_date.strftime('%y')}"

    _mw_ctr = st.session_state["mw_upload_ctr"]
    st.divider()
    adp_file = st.file_uploader("ADP Payroll CSV", type=["csv"], key=f"mw_adp_{_mw_ctr}",
                                 help="Upload the ADP payroll export.")

    st.divider()
    if st.button("Generate Report", type="primary", use_container_width=True, key="mw_run"):
        if adp_file is None:
            st.error("Please upload the ADP Payroll CSV.")
            st.stop()

        try:
            with st.spinner("Generating mid-week report..."):
                ref_data = _cached_reference_data()
                band_goals = _cached_band_goals()
                locked = ensure_current_week_locked(ref_data, band_goals, start_date)
                locked_band_goals = dict(zip(locked["revenue_band"], locked["hourly_goal"]))
                buf, mw_df = generate_midweek_report(adp_file, locked, locked_band_goals, report_dates, through_day)
            # Store results in session state so they persist after rerun
            st.session_state["mw_report_buf"] = buf
            st.session_state["mw_report_df"] = mw_df
            st.session_state["mw_report_fname"] = f"AVS_MidWeek_Report_{start_date.strftime('%m%d%Y')}.xlsx"
            st.session_state["mw_report_week"] = format_week_label(get_week_start(start_date))
            st.session_state["mw_report_day"] = through_day
            # Auto-clear file uploaders
            st.session_state["mw_upload_ctr"] += 1
            st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")

    if "mw_report_buf" in st.session_state:
        st.success("Report generated!")
        st.caption(f"Used locked config for week: {st.session_state['mw_report_week']} | Thresholds: {st.session_state['mw_report_day']}")

        # --- In-App Preview ---
        if "mw_report_df" in st.session_state:
            from avs_engine import DAY_THRESHOLDS
            day = st.session_state["mw_report_day"]
            thresholds = DAY_THRESHOLDS.get(day, DAY_THRESHOLDS["Friday"])
            _render_midweek_preview(st.session_state["mw_report_df"], day, thresholds)

        st.divider()
        st.download_button("Export to Excel",
                           data=st.session_state["mw_report_buf"],
                           file_name=st.session_state["mw_report_fname"],
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: AVS Performance - Store Level
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "AVS Performance - Store Level":
    st.title("AVS Performance - Store Level")
    st.caption("Compare store performance (Goal vs Actual hours) across weeks.")
    st.info("📊 **Reading this report:** A **negative** number means the store was **under** their hour goal (favorable). A **positive** number means the store was **over** their hour goal (unfavorable). Stores are ranked by Total Variance — the sum of absolute deviations from goal across all weeks.")

    locked_weeks = get_locked_weeks()
    # Only show completed weeks — exclude any week whose Wednesday end date hasn't passed yet
    _today = date.today()
    locked_weeks = [w for w in locked_weeks if (w + timedelta(days=6)) < _today]
    if not locked_weeks:
        st.info("No completed weekly data available yet. Run AVS reports to build history.")
        st.stop()

    # --- Load all data up front (actuals needed for Most Recent / Prior Week logic) ---
    all_locks = _cached_all_locks()
    actuals = _cached_weekly_actuals()
    actuals_week_strs = set(actuals["week_start"].unique()) if not actuals.empty else set()
    # Weeks that have actual data, sorted ascending
    weeks_with_actuals = sorted([w for w in locked_weeks if str(w) in actuals_week_strs])

    # --- Period filter (custom for Store Level — uses actuals to anchor Most Recent / Prior Week) ---
    col_period, col_period_val = st.columns(2)
    with col_period:
        period_filter = st.selectbox(
            "View by",
            ["Most Recent", "Prior Week", "All Weeks", "Month", "Quarter", "Year"],
            key="perf_period",
        )
    with col_period_val:
        if period_filter == "Most Recent":
            # Last week with actual AVS data submitted
            last_week = weeks_with_actuals[-1] if weeks_with_actuals else (max(locked_weeks) if locked_weeks else None)
            filtered_weeks = [last_week] if last_week else []
            if last_week:
                st.caption(format_week_label(last_week))
        elif period_filter == "Prior Week":
            # Second-to-last week with actual data (or let user pick from all weeks with actuals)
            if len(weeks_with_actuals) >= 2:
                prior_week = weeks_with_actuals[-2]
                filtered_weeks = [prior_week]
                st.caption(format_week_label(prior_week))
            elif len(weeks_with_actuals) == 1:
                filtered_weeks = [weeks_with_actuals[0]]
                st.caption(f"{format_week_label(weeks_with_actuals[0])} (only 1 week available)")
            else:
                filtered_weeks = [max(locked_weeks)] if locked_weeks else []
                st.caption("No submitted data yet — showing most recent locked week")
        elif period_filter == "Month":
            months_available = sorted(set((w.year, w.month) for w in locked_weeks))
            month_labels = [f"{y}-{m:02d}" for y, m in months_available]
            selected_month = st.selectbox("Select Month", month_labels, index=len(month_labels)-1, key="perf_month")
            sel_year, sel_month = int(selected_month[:4]), int(selected_month[5:])
            filtered_weeks = [w for w in locked_weeks if w.year == sel_year and w.month == sel_month]
        elif period_filter == "Quarter":
            quarters_available = sorted(set((w.year, (w.month - 1) // 3 + 1) for w in locked_weeks))
            quarter_labels = [f"{y} Q{q}" for y, q in quarters_available]
            selected_quarter = st.selectbox("Select Quarter", quarter_labels, index=len(quarter_labels)-1, key="perf_quarter")
            sel_year = int(selected_quarter[:4])
            sel_q = int(selected_quarter[-1])
            q_months = [(sel_q - 1) * 3 + 1, (sel_q - 1) * 3 + 2, (sel_q - 1) * 3 + 3]
            filtered_weeks = [w for w in locked_weeks if w.year == sel_year and w.month in q_months]
        elif period_filter == "Year":
            years_available = sorted(set(w.year for w in locked_weeks))
            selected_year = st.selectbox("Select Year", years_available, index=len(years_available)-1, key="perf_year")
            filtered_weeks = [w for w in locked_weeks if w.year == selected_year]
        else:  # All Weeks
            filtered_weeks = locked_weeks

    if not filtered_weeks:
        st.warning("No data for the selected period.")
        st.stop()

    # --- Filter by Store and DM ---
    week_strs = [str(w) for w in filtered_weeks]
    # Only include active stores (matches the weekly report engine)
    _active_store_ids = set(_cached_reference_data()["location_id"].tolist())
    perf_data = all_locks[
        all_locks["week_start"].isin(week_strs) &
        all_locks["location_id"].isin(_active_store_ids)
    ].copy()
    perf_data["hourly_goal"] = pd.to_numeric(perf_data["hourly_goal"], errors="coerce").fillna(0)

    col_dm_filter, col_store_filter = st.columns(2)
    with col_dm_filter:
        dm_list = sorted(perf_data["dm"].dropna().unique().tolist())
        selected_dms = st.multiselect("Filter by DM", dm_list, default=[], key="perf_dm_filter")
    with col_store_filter:
        store_pool = perf_data if not selected_dms else perf_data[perf_data["dm"].isin(selected_dms)]
        store_list = sorted(store_pool["store_name"].dropna().unique().tolist())
        selected_stores = st.multiselect("Filter by Store", store_list, default=[], key="perf_store_filter")

    # Apply filters
    if selected_dms:
        perf_data = perf_data[perf_data["dm"].isin(selected_dms)]
    if selected_stores:
        perf_data = perf_data[perf_data["store_name"].isin(selected_stores)]

    if perf_data.empty:
        st.warning("No data matches the selected filters.")
        st.stop()

    st.divider()
    st.subheader(f"Weeks: {format_week_label(filtered_weeks[0])} → {format_week_label(filtered_weeks[-1])}")
    st.caption(f"{len(filtered_weeks)} week(s)  ·  {perf_data['store_name'].nunique()} store(s)")

    week_strs_set = set(week_strs)

    # Filter actuals to the selected period
    if not actuals.empty:
        actuals_filtered = actuals[actuals["week_start"].isin(week_strs_set)].copy()
    else:
        actuals_filtered = pd.DataFrame(columns=["week_start", "location_id", "actual_hours", "variance"])

    # Build pivot using locked config as the backbone so ALL weeks in the
    # selected period always appear as columns — actuals filled in where available,
    # 0 where no report has been run yet for that week.
    grid = perf_data[["week_start", "location_id", "store_name"]].drop_duplicates()
    if not actuals_filtered.empty:
        grid = grid.merge(
            actuals_filtered[["week_start", "location_id", "variance"]],
            on=["week_start", "location_id"],
            how="left",
        )
        # Drop stores that have NO actual data for any week in the period
        # (they were in the locked config but never had a report run)
        stores_with_any_actuals = set(actuals_filtered["location_id"].unique())
        grid = grid[grid["location_id"].isin(stores_with_any_actuals)]
    else:
        grid["variance"] = 0.0
    grid["variance"] = grid["variance"].fillna(0.0)

    pivot = grid.pivot_table(
        index="store_name",
        columns="week_start",
        values="variance",
        aggfunc="first",
    ).fillna(0).reset_index()

    pivot.columns.name = None
    pivot = pivot.rename(columns={"store_name": "Store"})
    week_cols = [c for c in pivot.columns if c != "Store"]

    # Round to whole numbers
    for wc in week_cols:
        pivot[wc] = pivot[wc].round(0).astype(int)

    # --- Ranking + Total ---
    # Rank by sum of absolute values across all weeks (Total Variance)
    if week_cols:
        pivot["_sort_var"] = pivot[week_cols].abs().sum(axis=1)
    else:
        pivot["_sort_var"] = 0

    pivot = pivot.sort_values("_sort_var").reset_index(drop=True)
    pivot.insert(0, "Rank", range(1, len(pivot) + 1))
    pivot = pivot.drop(columns=["_sort_var"])

    # Rename week columns to end date (Wednesday) labels
    week_col_map = _rename_week_cols(week_cols)
    pivot = pivot.rename(columns=week_col_map)
    renamed_week_cols = [week_col_map.get(c, c) for c in week_cols]

    # Net Total and Total Variance columns at the end — only for multi-week views
    if len(renamed_week_cols) > 1:
        pivot["Net Total"] = pivot[renamed_week_cols].sum(axis=1)
        pivot["Total Variance"] = pivot[renamed_week_cols].abs().sum(axis=1)

    # --- Color coding: variance-based ---
    skip_cols = ("Rank", "Store", "Net Total", "Total Variance")
    has_actuals_weeks = set(actuals_filtered["week_start"].unique()) if not actuals_filtered.empty else set()
    styled = pivot.style.apply(
        _color_variance_cells, week_col_map=week_col_map,
        has_actuals_weeks=has_actuals_weeks, skip_cols=skip_cols, axis=1,
    )

    # Display all rows — no internal scroll, user scrolls the browser
    st.dataframe(
        styled,
        use_container_width=False,
        hide_index=True,
        height=(len(pivot) + 1) * 35 + 3,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: AVS Performance - DMs
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "AVS Performance - DMs":
    st.title("AVS Performance - DMs")
    st.caption("Summarized DM performance across all their stores by week.")
    st.info("📊 **Reading this report:** A **negative** number means the DM's stores were **under** their hour goal (favorable). A **positive** number means stores were **over** their hour goal (unfavorable). DMs are ranked by Total Variance — the sum of absolute deviations from goal across all weeks.")

    locked_weeks = get_locked_weeks()
    # Only show completed weeks — exclude any week whose Wednesday end date hasn't passed yet
    locked_weeks = [w for w in locked_weeks if (w + timedelta(days=6)) < date.today()]
    if not locked_weeks:
        st.info("No completed weekly data available yet. Run AVS reports to build history.")
        st.stop()

    # --- Load all data up front (actuals needed for Most Recent / Prior Week logic) ---
    all_locks = _cached_all_locks()
    actuals = _cached_weekly_actuals()
    actuals_week_strs = set(actuals["week_start"].unique()) if not actuals.empty else set()
    weeks_with_actuals = sorted([w for w in locked_weeks if str(w) in actuals_week_strs])

    # --- Period filter (custom — mirrors Store Level page) ---
    col_period, col_period_val = st.columns(2)
    with col_period:
        period_filter = st.selectbox(
            "View by",
            ["Most Recent", "Prior Week", "All Weeks", "Month", "Quarter", "Year"],
            key="dm_perf_period",
        )
    with col_period_val:
        if period_filter == "Most Recent":
            last_week = weeks_with_actuals[-1] if weeks_with_actuals else (max(locked_weeks) if locked_weeks else None)
            filtered_weeks = [last_week] if last_week else []
            if last_week:
                st.caption(format_week_label(last_week))
        elif period_filter == "Prior Week":
            if len(weeks_with_actuals) >= 2:
                prior_week = weeks_with_actuals[-2]
                filtered_weeks = [prior_week]
                st.caption(format_week_label(prior_week))
            elif len(weeks_with_actuals) == 1:
                filtered_weeks = [weeks_with_actuals[0]]
                st.caption(f"{format_week_label(weeks_with_actuals[0])} (only 1 week available)")
            else:
                filtered_weeks = [max(locked_weeks)] if locked_weeks else []
                st.caption("No submitted data yet — showing most recent locked week")
        elif period_filter == "Month":
            months_available = sorted(set((w.year, w.month) for w in locked_weeks))
            month_labels = [f"{y}-{m:02d}" for y, m in months_available]
            selected_month = st.selectbox("Select Month", month_labels, index=len(month_labels)-1, key="dm_perf_month")
            sel_year, sel_month = int(selected_month[:4]), int(selected_month[5:])
            filtered_weeks = [w for w in locked_weeks if w.year == sel_year and w.month == sel_month]
        elif period_filter == "Quarter":
            quarters_available = sorted(set((w.year, (w.month - 1) // 3 + 1) for w in locked_weeks))
            quarter_labels = [f"{y} Q{q}" for y, q in quarters_available]
            selected_quarter = st.selectbox("Select Quarter", quarter_labels, index=len(quarter_labels)-1, key="dm_perf_quarter")
            sel_year = int(selected_quarter[:4])
            sel_q = int(selected_quarter[-1])
            q_months = [(sel_q - 1) * 3 + 1, (sel_q - 1) * 3 + 2, (sel_q - 1) * 3 + 3]
            filtered_weeks = [w for w in locked_weeks if w.year == sel_year and w.month in q_months]
        elif period_filter == "Year":
            years_available = sorted(set(w.year for w in locked_weeks))
            selected_year = st.selectbox("Select Year", years_available, index=len(years_available)-1, key="dm_perf_year")
            filtered_weeks = [w for w in locked_weeks if w.year == selected_year]
        else:  # All Weeks
            filtered_weeks = locked_weeks

    if not filtered_weeks:
        st.warning("No data for the selected period.")
        st.stop()

    # --- Filter by DM ---
    week_strs = [str(w) for w in filtered_weeks]
    # Only include active stores (mirrors weekly report engine)
    _active_store_ids_dm = set(_cached_reference_data()["location_id"].tolist())
    perf_data = all_locks[
        all_locks["week_start"].isin(week_strs) &
        all_locks["location_id"].isin(_active_store_ids_dm)
    ].copy()
    perf_data["hourly_goal"] = pd.to_numeric(perf_data["hourly_goal"], errors="coerce").fillna(0)

    dm_list = sorted(perf_data["dm"].dropna().unique().tolist())
    selected_dms = st.multiselect("Filter by DM", dm_list, default=[], key="dm_perf_dm_filter")
    if selected_dms:
        perf_data = perf_data[perf_data["dm"].isin(selected_dms)]

    if perf_data.empty:
        st.warning("No data matches the selected filters.")
        st.stop()

    st.divider()
    st.subheader(f"Weeks: {format_week_label(filtered_weeks[0])} → {format_week_label(filtered_weeks[-1])}")
    st.caption(f"{len(filtered_weeks)} week(s)  ·  {perf_data['dm'].nunique()} DM(s)  ·  {perf_data['store_name'].nunique()} store(s)")

    week_strs_set = set(week_strs)

    if not actuals.empty:
        actuals_filtered = actuals[actuals["week_start"].isin(week_strs_set)].copy()
    else:
        actuals_filtered = pd.DataFrame(columns=["week_start", "location_id", "variance"])

    # Build pivot using locked config as backbone so ALL weeks in the period
    # appear as columns — actuals filled in where available, 0 where not yet run.
    store_dm = perf_data[["location_id", "dm", "week_start"]].drop_duplicates()

    if not actuals_filtered.empty:
        dm_grid = store_dm.merge(
            actuals_filtered[["week_start", "location_id", "variance"]],
            on=["week_start", "location_id"],
            how="left",
        )
        # Drop stores that have no actual data for any week in the period
        stores_with_any_actuals_dm = set(actuals_filtered["location_id"].unique())
        dm_grid = dm_grid[dm_grid["location_id"].isin(stores_with_any_actuals_dm)]
    else:
        dm_grid = store_dm.copy()
        dm_grid["variance"] = 0.0
    dm_grid["variance"] = dm_grid["variance"].fillna(0.0)

    dm_summary = dm_grid.groupby(["dm", "week_start"]).agg(
        total_variance=("variance", "sum"),
    ).reset_index()

    pivot = dm_summary.pivot_table(
        index="dm",
        columns="week_start",
        values="total_variance",
        aggfunc="first",
    ).fillna(0).reset_index()

    # Add store count per DM
    dm_store_counts = perf_data.groupby("dm")["store_name"].nunique().reset_index()
    dm_store_counts.columns = ["DM", "Stores"]

    pivot.columns.name = None
    pivot = pivot.rename(columns={"dm": "DM"})
    pivot = pivot.merge(dm_store_counts, on="DM", how="left")

    # Move Stores column to second position
    cols = pivot.columns.tolist()
    cols.remove("Stores")
    cols.insert(1, "Stores")
    pivot = pivot[cols]

    week_cols = [c for c in pivot.columns if c not in ("DM", "Stores")]

    # Round to whole numbers
    for wc in week_cols:
        pivot[wc] = pivot[wc].round(0).astype(int)

    # --- Ranking + Total (rank by sum of absolute values across all weeks) ---
    if week_cols:
        pivot["_sort_var"] = pivot[week_cols].abs().sum(axis=1)
    else:
        pivot["_sort_var"] = 0

    pivot = pivot.sort_values("_sort_var").reset_index(drop=True)
    pivot.insert(0, "Rank", range(1, len(pivot) + 1))
    pivot = pivot.drop(columns=["_sort_var"])

    # Rename week columns
    week_col_map = _rename_week_cols(week_cols)
    pivot = pivot.rename(columns=week_col_map)
    renamed_week_cols = [week_col_map.get(c, c) for c in week_cols]

    # Net Total and Total Variance columns at the end (multi-week only)
    if len(renamed_week_cols) > 1:
        pivot["Net Total"] = pivot[renamed_week_cols].sum(axis=1)
        pivot["Total Variance"] = pivot[renamed_week_cols].abs().sum(axis=1)

    # --- Color coding (Total handled inside _color_variance_cells) ---
    has_actuals_weeks = set(actuals_filtered["week_start"].unique()) if not actuals_filtered.empty else set()
    styled = pivot.style.apply(
        _color_variance_cells, week_col_map=week_col_map,
        has_actuals_weeks=has_actuals_weeks, skip_cols=("Rank", "DM", "Stores", "Net Total", "Total Variance"), axis=1,
    )

    st.dataframe(
        styled,
        use_container_width=False,
        hide_index=True,
        height=(len(pivot) + 1) * 35 + 3,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: GM Hot Streak
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "GM Hot Streak":
    st.title("🔥 GM Hot Streak")
    st.caption("Stores that have hit their goal 2 or more consecutive weeks. A store hits its goal when actual hours are within 30 hours of the target.")

    actuals = _cached_weekly_actuals()
    if actuals.empty:
        st.info("No actuals data yet. Run AVS Weekly Reports to build history.")
        st.stop()

    all_locks = _cached_all_locks()

    # Get store info (name, DM) from locks
    store_info = all_locks[["location_id", "store_name", "dm"]].drop_duplicates()
    store_info = store_info.drop_duplicates(subset="location_id", keep="last")

    # Get all weeks with actuals, sorted chronologically
    sorted_weeks = sorted(actuals["week_start"].unique().tolist())
    stores = actuals["location_id"].unique().tolist()

    # Calculate current streak for each store
    streak_data = []
    for store_id in stores:
        store_actuals = actuals[actuals["location_id"] == store_id].copy()
        streak = 0
        for week_str in reversed(sorted_weeks):
            week_row = store_actuals[store_actuals["week_start"] == week_str]
            if week_row.empty:
                break
            variance = abs(week_row["variance"].values[0])
            if variance <= 30:
                streak += 1
            else:
                break
        if streak >= 2:
            info = store_info[store_info["location_id"] == store_id]
            store_name = info["store_name"].values[0] if len(info) > 0 else store_id
            dm_name = info["dm"].values[0] if len(info) > 0 else ""
            streak_data.append({"Store": store_name, "DM": dm_name, "Streak (Weeks)": streak})

    if not streak_data:
        st.info("No stores are currently on a streak (2+ consecutive weeks hitting their goal).")
    else:
        streak_df = pd.DataFrame(streak_data)
        streak_df = streak_df.sort_values("Streak (Weeks)", ascending=False).reset_index(drop=True)
        streak_df.insert(0, "Rank", range(1, len(streak_df) + 1))

        st.metric("Stores on a Streak", len(streak_df))

        st.dataframe(
            streak_df,
            use_container_width=False,
            hide_index=True,
            height=(len(streak_df) + 1) * 35 + 3,
        )

    st.info(
        "A streak means hitting the goal (within 30 hours) for 2+ consecutive weeks. "
        "Streaks are counted backwards from the most recent week."
    )


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Manage Stores
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Manage Stores":
    st.title("Manage Store Locations")
    st.caption("Add or remove stores from the master locations list.")

    stores_df = load_stores()
    if not stores_df.empty:
        stores_df["location_id"] = stores_df["location_id"].str.strip().str.upper()
        stores_df["store_name"] = stores_df["store_name"].str.strip()

    st.subheader(f"Current Stores ({len(stores_df)})")
    if not stores_df.empty:
        display = stores_df.copy()
        display.columns = ["Store Number", "Store Name"]
        display = display.sort_values("Store Number").reset_index(drop=True)
        display.index = display.index + 1
        st.dataframe(display, use_container_width=True, height=400)
    else:
        st.info("No stores in the master list yet.")

    st.divider()
    st.subheader("Add a Store")
    with st.form("add_store_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            new_id = st.text_input("Store Number", placeholder="e.g. 112-0039")
        with col2:
            new_name = st.text_input("Store Name", placeholder="e.g. Springfield (MO)")
        add_btn = st.form_submit_button("Add Store", type="primary")

    if add_btn:
        new_id_clean = new_id.strip().upper()
        new_name_clean = new_name.strip()
        if not new_id_clean:
            st.error("Please enter a store number.")
        elif not new_name_clean:
            st.error("Please enter a store name.")
        elif new_id_clean in stores_df["location_id"].values:
            st.error(f"Store {new_id_clean} already exists.")
        else:
            save_store(new_id_clean, new_name_clean)

            # Also add to reference_data with defaults
            dm_list = _cached_dm_list()
            save_reference_data_row(
                new_id_clean, new_name_clean,
                dm_list[0] if dm_list else "",
                "<25k",
            )

            st.success(f"Added store {new_id_clean} — {new_name_clean}")
            st.cache_data.clear()
            st.rerun()

    st.divider()
    st.subheader("Remove a Store")
    if not stores_df.empty:
        options = stores_df.sort_values("location_id").apply(
            lambda r: f"{r['location_id']}  —  {r['store_name']}", axis=1
        ).tolist()
        selected = st.selectbox("Select store to remove", options, index=None, placeholder="Choose a store...")
        if selected:
            remove_id = selected.split("  —  ")[0].strip()
            st.warning(f"This will remove **{selected}** from the master list.")
            if st.button("Confirm Remove", type="primary"):
                delete_store(remove_id)
                delete_reference_data(remove_id)

                st.success(f"Removed store {remove_id}")
                st.cache_data.clear()
                st.rerun()
    else:
        st.info("No stores to remove.")


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Store Revenue Bands
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Store Revenue Bands":
    st.title("Store Revenue Bands")
    st.caption("Assign revenue bands per store for current and future weeks. Lock weeks to freeze their config.")

    ref_df = _cached_reference_data()
    if ref_df.empty:
        st.error("No reference data found. Please add stores first.")
        st.stop()

    ref_df = ref_df.sort_values("location_id").reset_index(drop=True)
    band_goals = _cached_band_goals()

    # Build 5-week range: current + 4 future
    current_week = get_week_start()
    weeks = [current_week + timedelta(days=7 * i) for i in range(5)]
    week_labels = [format_week_label(w) for w in weeks]
    week_statuses = [get_week_status(w) for w in weeks]

    # Load existing data for each week (locked, draft, or None)
    week_data = {}
    for w, status in zip(weeks, week_statuses):
        if status == "locked":
            cfg = load_locked_config(w)
            week_data[str(w)] = {r["location_id"]: r["revenue_band"] for _, r in cfg.iterrows()} if cfg is not None else {}
        elif status == "draft":
            cfg = load_draft_config(w)
            week_data[str(w)] = {r["location_id"]: r["revenue_band"] for _, r in cfg.iterrows()} if cfg is not None else {}
        else:
            week_data[str(w)] = {}

    # --- Header row with status badges ---
    st.markdown("""<style>
    .week-locked { color: #fff; background: #2B3A4E; padding: 2px 8px; border-radius: 4px; font-size: 0.8em; }
    .week-draft { color: #2B3A4E; background: #C49A5C; padding: 2px 8px; border-radius: 4px; font-size: 0.8em; }
    .week-open { color: #666; background: #eee; padding: 2px 8px; border-radius: 4px; font-size: 0.8em; }
    </style>""", unsafe_allow_html=True)

    header_cols = st.columns([2, 2.5, 0.5] + [2] * 5)
    with header_cols[0]:
        st.markdown("**Store #**")
    with header_cols[1]:
        st.markdown("**Store Name**")
    with header_cols[2]:
        st.markdown("** **")
    for i, (w, label, status) in enumerate(zip(weeks, week_labels, week_statuses)):
        with header_cols[i + 3]:
            if w == current_week:
                st.markdown(f"**{label}**<br><span class='week-locked'>Current</span>", unsafe_allow_html=True)
            elif status == "locked":
                st.markdown(f"**{label}**<br><span class='week-locked'>Locked</span>", unsafe_allow_html=True)
            elif status == "draft":
                st.markdown(f"**{label}**<br><span class='week-draft'>Draft</span>", unsafe_allow_html=True)
            else:
                st.markdown(f"**{label}**<br><span class='week-open'>Open</span>", unsafe_allow_html=True)
    st.divider()

    # --- Load tooltip data (submissions + performance) ---
    _next_week = weeks[1] if len(weeks) > 1 else None
    _submissions = {}
    _sos_last    = {}
    _votg_last   = {}

    try:
        from supabase_db import get_supabase as _get_sb
        _sb = _get_sb()
        if _next_week:
            _sub_resp = _sb.table("rev_band_submissions").select("*").eq(
                "week_start", str(_next_week)).execute()
            _submissions = {r["location_id"]: r for r in (_sub_resp.data or [])}

        _sos_resp = _sb.table("store_sos_weekly").select(
            "location_id,week_start,good_shift_rank,total_stores,total_time"
        ).gte("week_start", str(current_week - timedelta(weeks=4))).lt(
            "week_start", str(current_week)).order("week_start", desc=True).execute()
        for _r in (_sos_resp.data or []):
            if _r["location_id"] not in _sos_last:
                _sos_last[_r["location_id"]] = _r

        _votg_resp = _sb.table("store_votg_weekly").select(
            "location_id,week_start,votg_rank,total_stores,total_negative_reviews"
        ).gte("week_start", str(current_week - timedelta(weeks=4))).lt(
            "week_start", str(current_week)).order("week_start", desc=True).execute()
        for _r in (_votg_resp.data or []):
            if _r["location_id"] not in _votg_last:
                _votg_last[_r["location_id"]] = _r
    except Exception as e:
        print(f"[WARN] VOTG tooltip load failed: {e}")

    # Load sales data for tooltip — use weekly_actuals for recent weeks (more accurate than daily store_sales)
    _lw_start  = current_week - timedelta(weeks=1)   # last complete week start
    _2w_start  = current_week - timedelta(weeks=2)   # two weeks ago start
    _py_week   = _lw_start - timedelta(weeks=52)     # prior year same week
    _store_sales_map = {}
    try:
        from collections import defaultdict

        # Recent weekly sales from weekly_actuals (last 2 weeks)
        _actuals = _sb.table("weekly_actuals").select("location_id,week_start,net_sales").gte(
            "week_start", str(_2w_start)
        ).lt("week_start", str(current_week)).execute().data or []

        _wk_sales = defaultdict(lambda: defaultdict(float))
        for _r in _actuals:
            _wk = date.fromisoformat(str(_r["week_start"])[:10])
            _wk_sales[_r["location_id"]][_wk] += float(_r.get("net_sales") or 0)

        # PY sales from store_sales (weekly_actuals doesn't go back a year yet)
        _py_data = _sb.table("store_sales").select("location_id,sale_date,net_sales").gte(
            "sale_date", str(_py_week)
        ).lt("sale_date", str(_py_week + timedelta(weeks=1))).execute().data or []

        _py_by_store = defaultdict(float)
        for _r in _py_data:
            _py_by_store[_r["location_id"]] += float(_r.get("net_sales") or 0)

        all_store_ids = set(r["location_id"] for r in _actuals)
        for _sid in all_store_ids:
            _lw_val  = _wk_sales[_sid].get(_lw_start)
            _2w_val  = _wk_sales[_sid].get(_2w_start)
            _vals    = [v for v in [_lw_val, _2w_val] if v and v > 0]
            _avg_val = sum(_vals) / len(_vals) if _vals else None
            _py_val  = _py_by_store.get(_sid)
            _store_sales_map[_sid] = {
                "lw":  _lw_val  if _lw_val  and _lw_val  > 0 else None,
                "2w":  _2w_val  if _2w_val  and _2w_val  > 0 else None,
                "avg": _avg_val,
                "py":  _py_val  if _py_val  and _py_val  > 0 else None,
            }
    except Exception as e:
        print(f"[WARN] Sales tooltip load failed: {e}")

    # --- Grid form ---
    with st.form("revenue_bands_grid"):
        grid_data = {str(w): {} for w in weeks}

        for _, row in ref_df.iterrows():
            store_id = row["location_id"]
            store_name = row["store_name"]
            current_band = row["revenue_band"] if pd.notna(row["revenue_band"]) else "<25k"

            cols = st.columns([2, 2.5, 0.5] + [2] * 5)
            with cols[0]:
                st.text(store_id)
            with cols[1]:
                st.text(store_name)
            with cols[2]:
                with st.popover("ⓘ", use_container_width=True):
                    sub        = _submissions.get(store_id, {})
                    status_val = sub.get("status", "")
                    # GM status
                    if not sub or status_val == "pending_gm":
                        st.markdown("**GM:** ⏳ Not submitted")
                    else:
                        st.markdown(f"**GM:** ✅ `{sub.get('selected_band','—')}`")
                    # DM status
                    override = sub.get("dm_override_band")
                    if override:
                        st.markdown(f"**DM:** 🔄 Override `{override}`")
                    elif status_val in ("pending_admin", "approved"):
                        st.markdown("**DM:** ✅ Approved")
                    elif status_val == "pending_dm":
                        st.markdown("**DM:** ⏳ Pending")
                    else:
                        st.markdown("**DM:** —")
                    st.divider()
                    # Sales — label each row with its date range
                    _s = _store_sales_map.get(store_id, {})
                    _lw_s  = f"${_s['lw']:,.0f}"  if _s.get("lw")  else "N/A"
                    _2w_s  = f"${_s['2w']:,.0f}"  if _s.get("2w")  else "N/A"
                    _avg_s = f"${_s['avg']:,.0f}" if _s.get("avg") else "N/A"
                    _py_s  = f"${_s['py']:,.0f}"  if _s.get("py")  else "N/A"
                    # Build compact date labels  e.g. "3/26–4/1"
                    def _wk_label(start):
                        end = start + timedelta(days=6)
                        return f"{start.month}/{start.day}–{end.month}/{end.day}"
                    _lw_lbl = _wk_label(_lw_start)
                    _2w_lbl = _wk_label(_2w_start)
                    _py_lbl = _wk_label(_py_week)
                    st.markdown(f"**Last Week** ({_lw_lbl}): {_lw_s}")
                    st.markdown(f"**2 Wks Ago** ({_2w_lbl}): {_2w_s}")
                    st.markdown(f"**Avg (2 wks):** {_avg_s}")
                    st.markdown(f"**Prior Year** ({_py_lbl}): {_py_s}")
                    # SoS
                    sos = _sos_last.get(store_id, {})
                    if sos:
                        _tt = str(sos.get("total_time") or "")
                        _sos_min = "N/A"
                        if ":" in _tt:
                            try:
                                _m, _s2 = _tt.split(":")
                                _sos_min = f"{(int(_m)*60+int(_s2))/60:.1f} min"
                            except Exception:
                                pass
                        st.markdown(f"**SoS:** {_sos_min} — #{sos.get('good_shift_rank','?')} of {sos.get('total_stores','?')}")
                    # VOTG
                    votg = _votg_last.get(store_id, {})
                    if votg:
                        st.markdown(f"**VOTG:** {votg.get('total_negative_reviews','?')} neg — #{votg.get('votg_rank','?')} of {votg.get('total_stores','?')}")

            for i, (w, status) in enumerate(zip(weeks, week_statuses)):
                w_str = str(w)
                # Get band for this week: from DB if exists, else from current config
                existing_band = week_data[w_str].get(store_id, current_band)
                band_idx = BAND_OPTIONS.index(existing_band) if existing_band in BAND_OPTIONS else 0

                with cols[i + 3]:
                    if status == "locked" or w == current_week:
                        st.text(existing_band)
                        grid_data[w_str][store_id] = existing_band
                    else:
                        grid_data[w_str][store_id] = st.selectbox(
                            f"Band {store_id} {w_str}",
                            BAND_OPTIONS,
                            index=band_idx,
                            key=f"band_{store_id}_{w_str}",
                            label_visibility="collapsed",
                        )

        st.divider()

        # Save buttons — one per unlocked week
        save_cols = st.columns([2, 2.5, 0.5] + [2] * 5)
        with save_cols[0]:
            st.write("")  # spacer
        with save_cols[1]:
            st.write("")  # spacer
        with save_cols[2]:
            st.write("")  # spacer
        for i, (w, label, status) in enumerate(zip(weeks, week_labels, week_statuses)):
            with save_cols[i + 3]:
                if w == current_week or status == "locked":
                    st.write("")  # locked — no button
                else:
                    st.form_submit_button(f"Save", key=f"save_draft_{w}")

    # Handle per-week saves
    for w, label, status in zip(weeks, week_labels, week_statuses):
        if w == current_week or status == "locked":
            continue
        if st.session_state.get(f"save_draft_{w}"):
            w_str = str(w)
            # Always grab the latest hourly goals from Supabase (bypass cache)
            st.cache_data.clear()
            fresh_goals = load_band_goals()
            save_draft_bands(w, grid_data[w_str], ref_df, fresh_goals)
            st.success(f"Draft saved for {label}!")
            st.rerun()

    # Lock buttons (outside form since forms can only have one submit)
    st.subheader("Lock a Week")
    st.caption("Locking a week freezes its bands. Once locked, changes require admin override.")
    lock_cols = st.columns(5)
    for i, (w, label, status) in enumerate(zip(weeks, week_labels, week_statuses)):
        with lock_cols[i]:
            if status == "locked":
                st.success(f"{label}: Locked")
            elif status == "draft":
                if st.button(f"Lock {label}", key=f"lock_btn_{w}"):
                    lock_drafts(w)
                    log_change(
                        user_email=current_user or "admin",
                        week_start=w,
                        location_id="ALL",
                        field_changed="week_lock",
                        old_value="draft",
                        new_value="locked",
                        action="manual-lock",
                    )
                    st.success(f"Locked {label}!")
                    st.cache_data.clear()
                    st.rerun()
            else:
                st.info(f"{label}: Save drafts first")

    # --- Export to Excel ---
    st.divider()
    st.subheader("Export to Excel")
    st.caption("Export store config (Store #, Store Name, Revenue Band, Hourly Goal) for selected weeks.")

    export_options = []
    for w, label, status in zip(weeks, week_labels, week_statuses):
        if status in ("locked", "draft"):
            export_options.append((w, label, status))

    if not export_options:
        st.info("No locked or drafted weeks available to export.")
    else:
        export_choices = [f"{label} ({status})" for _, label, status in export_options]
        selected_exports = st.multiselect(
            "Select weeks to export",
            export_choices,
            default=[export_choices[0]] if export_choices else [],
            key="rev_band_export_select",
        )

        if st.button("Export to Excel", key="rev_band_export_btn", type="primary"):
            if not selected_exports:
                st.warning("Please select at least one week.")
            else:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                wb = Workbook()
                wb.remove(wb.active)

                NAVY = "2B3A4E"
                GOLD = "C49A5C"
                GOLD_LIGHT = "F5F0EB"
                WHITE = "FFFFFF"
                GRAY_BG = "F7F7F7"

                header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
                header_fill = PatternFill("solid", fgColor=NAVY)
                subheader_font = Font(name="Calibri", bold=True, color=NAVY, size=10)
                data_font = Font(name="Calibri", size=10)
                center = Alignment(horizontal="center", vertical="center")
                left = Alignment(horizontal="left", vertical="center")
                thin_border = Border(
                    left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"),
                )
                stripe_fill = PatternFill("solid", fgColor=GRAY_BG)

                for choice in selected_exports:
                    idx = export_choices.index(choice)
                    w, label, status = export_options[idx]

                    if status == "locked":
                        cfg = load_locked_config(w)
                    elif status == "draft":
                        cfg = load_draft_config(w)
                    else:
                        # Open week — build from current settings
                        cfg = None

                    if cfg is None or (hasattr(cfg, 'empty') and cfg.empty):
                        # Build from current base config
                        ref_data = _cached_reference_data()
                        current_goals = _cached_band_goals()
                        cfg = ref_data.copy()
                        cfg["hourly_goal"] = cfg["revenue_band"].map(current_goals).fillna(0)

                    cfg = cfg.sort_values("location_id").reset_index(drop=True)

                    # Load 2-week average sales for each store
                    _export_sales = {}
                    try:
                        from collections import defaultdict as _dd
                        from supabase_db import get_supabase
                        _exp_sb = get_supabase()
                        _exp_2w_start = w - timedelta(weeks=2)
                        _exp_actuals = _exp_sb.table("weekly_actuals").select(
                            "location_id,week_start,net_sales"
                        ).gte("week_start", str(_exp_2w_start)).lt(
                            "week_start", str(w)
                        ).execute().data or []
                        _exp_wk = _dd(list)
                        for _r in _exp_actuals:
                            _v = float(_r.get("net_sales") or 0)
                            if _v > 0:
                                _exp_wk[_r["location_id"]].append(_v)
                        for _sid, _vals in _exp_wk.items():
                            _export_sales[_sid] = sum(_vals) / len(_vals) if _vals else None
                    except Exception:
                        pass

                    # Load most recent LOCKED week before this one to compare band changes
                    prev_bands = {}
                    search_w = w - timedelta(days=7)
                    for _ in range(12):  # Search up to 12 weeks back
                        if lock_exists(search_w):
                            prev_cfg = load_locked_config(search_w)
                            if prev_cfg is not None and not prev_cfg.empty:
                                prev_bands = dict(zip(prev_cfg["location_id"], prev_cfg["revenue_band"]))
                            break
                        search_w -= timedelta(days=7)

                    # Band ordering for comparison (higher index = higher band)
                    BAND_ORDER = [
                        "<25k", "25k-30k", "30k-35k", "35k-40k",
                        "40k-45k", "45k-50k", "50k+", "NRO", "NRO Seasoned"
                    ]
                    def band_rank(b):
                        try:
                            return BAND_ORDER.index(b)
                        except ValueError:
                            return -1

                    # Sheet name (max 31 chars for Excel)
                    sheet_name = label.replace(" – ", "-").replace("/", "-")[:31]
                    ws = wb.create_sheet(title=sheet_name)

                    GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
                    RED_FILL = PatternFill("solid", fgColor="FFC7CE")

                    # Title row
                    ws.merge_cells("A1:F1")
                    title_cell = ws["A1"]
                    title_cell.value = f"Ram-Z Restaurant Group — Store Config: {label}"
                    title_cell.font = Font(name="Calibri", bold=True, color=NAVY, size=14)
                    title_cell.alignment = left

                    # Status row
                    ws.merge_cells("A2:F2")
                    status_cell = ws["A2"]
                    status_cell.value = f"Status: {status.upper()}"
                    status_cell.font = Font(name="Calibri", bold=True, color=GOLD, size=10)

                    # Legend row
                    ws["A3"].value = "Legend:"
                    ws["A3"].font = Font(name="Calibri", size=9, italic=True)
                    ws["B3"].fill = GREEN_FILL
                    ws["B3"].value = "Band Increased"
                    ws["B3"].font = Font(name="Calibri", size=9)
                    ws["C3"].fill = RED_FILL
                    ws["C3"].value = "Band Decreased"
                    ws["C3"].font = Font(name="Calibri", size=9)

                    # Headers
                    headers = ["Store #", "Store Name", "Revenue Band", "Hourly Goal", "2-Wk Avg Sales", "Change"]
                    for col_idx, h in enumerate(headers, 1):
                        cell = ws.cell(row=5, column=col_idx, value=h)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center
                        cell.border = thin_border

                    # Data rows
                    for row_idx, (_, r) in enumerate(cfg.iterrows(), 6):
                        loc_id = r.get("location_id", "")
                        curr_band = r.get("revenue_band", "")
                        prev_band = prev_bands.get(loc_id)

                        # Determine band change
                        change_label = ""
                        row_fill = None
                        if prev_band and prev_band != curr_band:
                            curr_rank = band_rank(curr_band)
                            prev_rank_val = band_rank(prev_band)
                            if curr_rank > prev_rank_val:
                                change_label = f"↑ {prev_band} → {curr_band}"
                                row_fill = GREEN_FILL
                            elif curr_rank < prev_rank_val:
                                change_label = f"↓ {prev_band} → {curr_band}"
                                row_fill = RED_FILL

                        avg_sales = _export_sales.get(loc_id)
                        values = [
                            loc_id,
                            r.get("store_name", ""),
                            curr_band,
                            float(r.get("hourly_goal", 0)) if pd.notna(r.get("hourly_goal")) else 0,
                            avg_sales if avg_sales else "",
                            change_label,
                        ]
                        for col_idx, val in enumerate(values, 1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=val)
                            cell.font = data_font
                            cell.border = thin_border
                            if col_idx in (1, 3, 4, 5):
                                cell.alignment = center
                            else:
                                cell.alignment = left
                            # Apply band change highlight
                            if row_fill:
                                cell.fill = row_fill
                            elif row_idx % 2 == 1:
                                cell.fill = stripe_fill
                            # Format hourly goal as integer
                            if col_idx == 4:
                                cell.number_format = "#,##0"
                            # Format 2-wk avg sales as currency
                            if col_idx == 5 and isinstance(val, (int, float)):
                                cell.number_format = "$#,##0"

                    # Auto-fit column widths
                    from openpyxl.utils import get_column_letter
                    for col_idx in range(1, 7):
                        max_len = len(headers[col_idx - 1])
                        for row_idx in range(6, 6 + len(cfg)):
                            val = ws.cell(row=row_idx, column=col_idx).value
                            if val:
                                max_len = max(max_len, len(str(val)))
                        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)

                week_count = len(selected_exports)
                filename = f"RamZ_Store_Config_{'_'.join(str(export_options[export_choices.index(c)][0].strftime('%m%d')) for c in selected_exports)}.xlsx"
                st.download_button(
                    label=f"Download ({week_count} week{'s' if week_count > 1 else ''})",
                    data=buf,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="rev_band_download",
                )


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: DM Assignments
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "DM Assignments":
    st.title("DM Assignments")
    st.caption("Manage the list of District Managers and assign stores to DMs.")

    show_week_deadline_banner()

    # --- DM List Management ---
    st.subheader("DM List")
    dm_list = _cached_dm_list()
    st.write("Current DMs: " + ", ".join(dm_list) if dm_list else "No DMs defined.")

    col1, col2 = st.columns(2)
    with col1:
        with st.form("add_dm_form", clear_on_submit=True):
            new_dm = st.text_input("New DM Name", placeholder="e.g. John")
            add_dm_btn = st.form_submit_button("Add DM")
        if add_dm_btn and new_dm.strip():
            dm_name = new_dm.strip()
            if dm_name in dm_list:
                st.error(f"{dm_name} already exists.")
            else:
                add_dm(dm_name)
                st.success(f"Added DM: {dm_name}")
                st.cache_data.clear()
                st.rerun()

    with col2:
        if dm_list:
            with st.form("remove_dm_form"):
                remove_dm = st.selectbox("Remove DM", dm_list, index=None, placeholder="Choose a DM...")
                remove_dm_btn = st.form_submit_button("Remove DM")
            if remove_dm_btn and remove_dm:
                db_remove_dm(remove_dm)
                st.success(f"Removed DM: {remove_dm}")
                st.cache_data.clear()
                st.rerun()

    st.divider()

    # --- Store Active / Inactive Toggle ---
    st.subheader("Active Stores")
    st.caption("Inactive stores are excluded from all AVS reports. Use this to remove stores that are closed, corporate, or not yet part of tracking.")

    # Load ALL stores (including inactive) for admin management
    all_ref_df = load_reference_data(active_only=False).sort_values("location_id").reset_index(drop=True)

    if all_ref_df.empty:
        st.error("No reference data found.")
        st.stop()

    active_changes = {}
    hdr1, hdr2, hdr3, hdr4 = st.columns([2, 4, 2, 2])
    hdr1.markdown("**Store #**")
    hdr2.markdown("**Store Name**")
    hdr3.markdown("**DM**")
    hdr4.markdown("**Active in Reports**")

    for _, row in all_ref_df.iterrows():
        c1, c2, c3, c4 = st.columns([2, 4, 2, 2])
        c1.text(row["location_id"])
        c2.text(row["store_name"])
        c3.text(row.get("dm", "") or "")
        is_active = bool(row.get("active", True))
        toggled = c4.checkbox(
            "Active",
            value=is_active,
            key=f"active_{row['location_id']}",
            label_visibility="collapsed",
        )
        if toggled != is_active:
            active_changes[row["location_id"]] = toggled

    if active_changes:
        if st.button("Save Active Status", type="primary"):
            for store_id, new_active in active_changes.items():
                set_store_active(store_id, new_active)
            st.success("Store active status updated.")
            st.cache_data.clear()
            st.rerun()

    st.divider()

    # --- Store-to-DM Assignment ---
    st.subheader("Store-to-DM Assignments")
    st.caption("Only active stores are shown here.")

    ref_df = _cached_reference_data()
    if ref_df.empty:
        st.error("No reference data found.")
        st.stop()

    ref_df = ref_df.sort_values("location_id").reset_index(drop=True)
    dm_list = _cached_dm_list()

    if not dm_list:
        st.warning("No DMs defined. Add DMs above first.")
        st.stop()

    with st.form("dm_assignments_form"):
        new_dms = {}
        for idx, row in ref_df.iterrows():
            current_dm = row["dm"] if pd.notna(row["dm"]) else ""
            current_idx = dm_list.index(current_dm) if current_dm in dm_list else 0

            col1, col2, col3 = st.columns([2, 3, 3])
            with col1:
                st.text(row["location_id"])
            with col2:
                st.text(row["store_name"])
            with col3:
                new_dms[row["location_id"]] = st.selectbox(
                    f"DM for {row['location_id']}",
                    dm_list,
                    index=current_idx,
                    key=f"dm_{row['location_id']}",
                    label_visibility="collapsed",
                )

        save_btn = st.form_submit_button("Save Changes", type="primary", use_container_width=True)

    if save_btn:
        for store_id, dm in new_dms.items():
            ref_df.loc[ref_df["location_id"] == store_id, "dm"] = dm
        save_reference_data_bulk(ref_df)
        st.success("DM assignments saved! These will apply to the next unlocked week.")
        st.cache_data.clear()
        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Hourly Goals
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Hourly Goals":
    st.title("Hourly Goals by Revenue Band")
    st.caption("Edit the weekly hourly goal for each revenue band. Changes apply to the next unlocked week.")

    show_week_deadline_banner()

    band_goals = _cached_band_goals()

    with st.form("hourly_goals_form"):
        new_goals = {}
        for band in BAND_OPTIONS:
            current_goal = band_goals.get(band, 0)
            col1, col2 = st.columns([2, 2])
            with col1:
                st.markdown(f"**{band}**")
            with col2:
                new_goals[band] = st.number_input(
                    f"Goal for {band}",
                    value=int(current_goal),
                    min_value=0,
                    max_value=9999,
                    step=1,
                    key=f"goal_{band}",
                    label_visibility="collapsed",
                )

        save_btn = st.form_submit_button("Save Changes", type="primary", use_container_width=True)

    if save_btn:
        save_band_goals(new_goals)
        st.success("Hourly goals saved! These will apply to the next unlocked week.")
        st.cache_data.clear()
        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Weekly Config (Admin)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Weekly Config":
    st.title("Weekly Config Override")
    st.caption("View and override locked weekly configurations. Admin only.")

    if not user_is_admin:
        st.error("You do not have admin access.")
        st.stop()

    locked_weeks = get_locked_weeks()
    if not locked_weeks:
        st.info("No weeks have been locked yet. Locks are created when AVS reports are run.")
        st.stop()

    # Week selector
    week_labels = {w: format_week_label(w) for w in locked_weeks}
    selected_week = st.selectbox(
        "Select Week",
        locked_weeks,
        format_func=lambda w: f"{w} — {week_labels[w]}",
        index=len(locked_weeks) - 1,  # default to most recent
    )

    locked = load_locked_config(selected_week)
    if locked is None:
        st.error("Could not load locked config for this week.")
        st.stop()

    week_status = get_week_status(selected_week)
    st.subheader(f"Config: {week_labels[selected_week]}  ({week_status or 'unknown'})")

    # Show source info
    all_locks = _cached_all_locks()
    week_data = all_locks[all_locks["week_start"] == str(selected_week)]
    sources = week_data["source"].unique().tolist() if "source" in week_data.columns else []
    if sources:
        source_str = ", ".join(sources)
        if "auto-carry-forward" in sources:
            st.warning(f"Source: {source_str} — This week's config was auto-carried from the previous week.")
        else:
            st.success(f"Source: {source_str}")

    # Display locked config
    display = locked[["location_id", "store_name", "dm", "revenue_band", "hourly_goal"]].copy()
    display.columns = ["Store #", "Store Name", "DM", "Revenue Band", "Hourly Goal"]
    display = display.sort_values("Store #").reset_index(drop=True)
    display.index = display.index + 1
    st.dataframe(display, use_container_width=True, height=400)

    # Override form
    st.divider()
    st.subheader("Override a Value")
    st.caption("Changes are logged with your email and timestamp.")

    store_options = locked.sort_values("location_id").apply(
        lambda r: f"{r['location_id']}  —  {r['store_name']}", axis=1
    ).tolist()

    with st.form("override_form"):
        override_store = st.selectbox("Store", store_options, index=None, placeholder="Choose a store...")
        override_field = st.selectbox("Field to Override", ["dm", "revenue_band", "hourly_goal"])

        if override_field == "revenue_band":
            override_value = st.selectbox("New Value", BAND_OPTIONS)
        elif override_field == "hourly_goal":
            override_value = st.number_input("New Value", min_value=0, max_value=9999, step=1)
        else:
            dm_list = _cached_dm_list()
            override_value = st.selectbox("New Value", dm_list if dm_list else [""])

        override_btn = st.form_submit_button("Apply Override", type="primary")

    if override_btn:
        if override_store is None:
            st.error("Please select a store.")
        else:
            store_id = override_store.split("  —  ")[0].strip()
            try:
                override_locked_value(
                    selected_week, store_id, override_field,
                    override_value, current_user or "admin"
                )
                st.success(f"Override applied: {store_id} {override_field} = {override_value}")
                st.cache_data.clear()
                st.rerun()
            except Exception as e:
                st.error(f"Error: {e}")

    # Delete week lock
    st.divider()
    st.subheader("Delete Week Lock")
    st.caption("Remove the entire lock for this week. This allows the report to be re-run with fresh data.")
    st.warning(f"This will delete the locked config for **{week_labels[selected_week]}** and all associated data.")
    if st.button("Delete This Week's Lock", type="primary"):
        delete_week_lock(selected_week)
        delete_weekly_actuals(selected_week)
        log_change(
            user_email=current_user or "admin",
            week_start=selected_week,
            location_id="ALL",
            field_changed="week_lock",
            old_value=str(selected_week),
            new_value="deleted",
            action="admin-delete",
        )
        st.success(f"Lock deleted for {week_labels[selected_week]}. You can now re-run the report.")
        st.cache_data.clear()
        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Change Log (Admin)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Change Log":
    st.title("Change Log")
    st.caption("Audit trail of all weekly config changes and overrides. Admin only.")

    if not user_is_admin:
        st.error("You do not have admin access.")
        st.stop()

    log_df = load_change_log()

    if log_df.empty:
        st.info("No changes logged yet.")
        st.stop()

    # Filters
    col1, col2, col3 = st.columns(3)
    with col1:
        action_filter = st.multiselect(
            "Filter by Action",
            log_df["action"].unique().tolist(),
            default=log_df["action"].unique().tolist(),
        )
    with col2:
        if "week_start" in log_df.columns:
            week_filter = st.multiselect(
                "Filter by Week",
                sorted(log_df["week_start"].unique().tolist()),
                default=sorted(log_df["week_start"].unique().tolist()),
            )
        else:
            week_filter = []
    with col3:
        if "user_email" in log_df.columns:
            user_filter = st.multiselect(
                "Filter by User",
                sorted(log_df["user_email"].unique().tolist()),
                default=sorted(log_df["user_email"].unique().tolist()),
            )
        else:
            user_filter = []

    filtered = log_df.copy()
    if action_filter:
        filtered = filtered[filtered["action"].isin(action_filter)]
    if week_filter:
        filtered = filtered[filtered["week_start"].isin(week_filter)]
    if user_filter:
        filtered = filtered[filtered["user_email"].isin(user_filter)]

    filtered = filtered.sort_values("timestamp", ascending=False).reset_index(drop=True)
    filtered.index = filtered.index + 1

    st.subheader(f"Log Entries ({len(filtered)})")
    st.dataframe(filtered, use_container_width=True, height=500)


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Admin Users (Admin)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Admin Users":
    st.title("Admin Users")
    st.caption("Manage who has admin access to weekly config overrides and change logs.")

    if not user_is_admin:
        st.error("You do not have admin access.")
        st.stop()

    admins = load_admin_users()

    st.subheader(f"Current Admins ({len(admins)})")
    for i, email in enumerate(sorted(admins), 1):
        st.text(f"{i}. {email}")

    st.divider()
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Add Admin")
        with st.form("add_admin_form", clear_on_submit=True):
            new_admin_email = st.text_input("Email Address", placeholder="e.g. user@example.com")
            add_admin_btn = st.form_submit_button("Add Admin", type="primary")

        if add_admin_btn:
            email_clean = new_admin_email.strip().lower()
            if not email_clean:
                st.error("Please enter an email address.")
            elif "@" not in email_clean:
                st.error("Please enter a valid email address.")
            elif email_clean in admins:
                st.error(f"{email_clean} is already an admin.")
            else:
                add_admin(email_clean)
                st.success(f"Added admin: {email_clean}")
                st.cache_data.clear()
                st.rerun()

    with col2:
        st.subheader("Remove Admin")
        if len(admins) > 1:
            with st.form("remove_admin_form"):
                remove_email = st.selectbox("Select Admin to Remove", sorted(admins), index=None, placeholder="Choose...")
                remove_admin_btn = st.form_submit_button("Remove Admin")

            if remove_admin_btn and remove_email:
                if remove_email == current_user:
                    st.error("You cannot remove yourself as admin.")
                else:
                    remove_admin(remove_email)
                    st.success(f"Removed admin: {remove_email}")
                    st.cache_data.clear()
                    st.rerun()
        else:
            st.info("Cannot remove the last admin.")

# ==========================================
# Rev Band Approvals (Admin Dashboard)
# ==========================================
elif page == "Rev Band Approvals":
    if not user_is_admin:
        st.error("You do not have permission to access this page.")
        st.stop()

    st.title("📋 Rev Band Approvals")
    st.caption("Review GM revenue band submissions. Approve or reject for each store.")

    # Week selector
    all_subs = load_all_submissions()

    if all_subs.empty:
        st.info("No submissions yet. Submissions will appear here once GMs begin selecting revenue bands.")
    else:
        available_weeks = sorted(all_subs["week_start"].unique(), reverse=True)
        selected_week = st.selectbox("Select Week", available_weeks, index=0,
                                      format_func=lambda w: format_week_label(get_week_start(pd.Timestamp(w).date())))

        week_subs = all_subs[all_subs["week_start"] == selected_week].copy()

        # Summary metrics
        total = len(week_subs)
        pending_gm = len(week_subs[week_subs["status"] == "pending_gm"])
        pending_dm = len(week_subs[week_subs["status"] == "pending_dm"])
        pending_admin = len(week_subs[week_subs["status"] == "pending_admin"])
        approved = len(week_subs[week_subs["status"] == "approved"])
        rejected = len(week_subs[week_subs["status"] == "rejected"])

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Awaiting GM", pending_gm)
        c2.metric("Awaiting DM", pending_dm)
        c3.metric("Awaiting Admin", pending_admin)
        c4.metric("Approved", approved)
        c5.metric("Rejected", rejected)

        st.divider()

        # Filter options
        ref_data = _cached_reference_data()
        dm_list_vals = sorted(ref_data["dm"].dropna().unique().tolist())
        filter_col1, filter_col2 = st.columns(2)
        with filter_col1:
            status_filter = st.selectbox("Filter by Status", ["All", "pending_gm", "pending_dm", "pending_admin", "approved", "rejected"])
        with filter_col2:
            dm_filter = st.selectbox("Filter by DM", ["All"] + dm_list_vals)

        filtered = week_subs.copy()
        if status_filter != "All":
            filtered = filtered[filtered["status"] == status_filter]

        # Join with reference data to get DM
        if "dm" not in filtered.columns and not ref_data.empty:
            filtered = filtered.merge(ref_data[["location_id", "dm"]], on="location_id", how="left")

        if dm_filter != "All":
            filtered = filtered[filtered["dm"] == dm_filter]

        if filtered.empty:
            st.info("No submissions match the selected filters.")
        else:
            # Load band goals for hourly goal display
            band_goals = _cached_band_goals()

            for _, row in filtered.iterrows():
                store_name = ref_data[ref_data["location_id"] == row["location_id"]]["store_name"].values
                store_name = store_name[0] if len(store_name) > 0 else row["location_id"]
                current_band = ref_data[ref_data["location_id"] == row["location_id"]]["revenue_band"].values
                current_band = current_band[0] if len(current_band) > 0 else "N/A"

                status = row["status"]
                selected_band = row.get("selected_band", "")

                # Color code status
                if status == "approved":
                    status_color = "🟢"
                elif status == "rejected":
                    status_color = "🔴"
                elif status == "pending_admin":
                    status_color = "🟡"
                elif status == "pending_dm":
                    status_color = "🔵"
                else:
                    status_color = "⚪"

                with st.container():
                    col1, col2, col3, col4 = st.columns([3, 2, 2, 2])
                    with col1:
                        st.markdown(f"**{store_name}** ({row['location_id']})")
                        dm_name = row.get("dm", "N/A")
                        st.caption(f"DM: {dm_name}")
                    with col2:
                        st.markdown(f"Current: **{current_band}**")
                        if selected_band:
                            hourly_goal = band_goals.get(selected_band, "N/A")
                            st.markdown(f"Selected: **{selected_band}** ({hourly_goal} hrs)")
                    with col3:
                        st.markdown(f"{status_color} **{status.replace('_', ' ').title()}**")
                        if row.get("submitted_at"):
                            st.caption(f"Submitted: {str(row['submitted_at'])[:16]}")
                    with col4:
                        if status == "pending_admin":
                            if st.button("✅ Approve", key=f"approve_{row['id']}"):
                                approve_submission(row["id"], current_user or "admin")
                                st.cache_data.clear()
                                st.rerun()
                            reject_reason = st.text_input("Reason", key=f"reason_{row['id']}", placeholder="Optional")
                            if st.button("❌ Reject", key=f"reject_{row['id']}"):
                                reject_submission(row["id"], current_user or "admin", reject_reason)
                                st.cache_data.clear()
                                st.rerun()
                        elif status == "approved":
                            st.caption(f"By: {row.get('admin_approved_by', 'N/A')}")
                        elif status == "rejected":
                            st.caption(f"Reason: {row.get('rejection_reason', 'N/A')}")

                    # Expandable performance data
                    with st.expander(f"📊 View {store_name} Performance Data"):
                        perf_cols = st.columns(3)

                        # Sales data
                        with perf_cols[0]:
                            st.markdown("**📈 Sales**")
                            try:
                                _ap_sb = get_supabase()
                                sales_resp = _ap_sb.table("store_sales").select(
                                    "sale_date,net_sales"
                                ).eq("location_id", row["location_id"]).order(
                                    "sale_date", desc=True
                                ).limit(28).execute()  # up to 4 weeks of daily rows
                                if sales_resp.data:
                                    # Aggregate by Thu-Wed week
                                    from collections import defaultdict as _dd
                                    _wk_map = _dd(float)
                                    for _sr in sales_resp.data:
                                        _sd = date.fromisoformat(str(_sr["sale_date"])[:10])
                                        _wk = _sd - timedelta(days=(_sd.weekday() - 3) % 7)
                                        _wk_map[_wk] += float(_sr.get("net_sales") or 0)
                                    for _wk in sorted(_wk_map, reverse=True)[:4]:
                                        st.caption(f"Wk {_wk}: ${_wk_map[_wk]:,.0f}")
                                else:
                                    st.caption("No sales data available")
                            except Exception:
                                st.caption("No sales data available")

                        # SoS data
                        with perf_cols[1]:
                            st.markdown("**⏱️ Speed of Service**")
                            try:
                                sos_resp = _ap_sb.table("store_sos_weekly").select(
                                    "week_start,good_shift_rank,total_stores,total_time"
                                ).eq("location_id", row["location_id"]).order(
                                    "week_start", desc=True
                                ).limit(4).execute()
                                if sos_resp.data:
                                    secs_list = []
                                    for s in sos_resp.data:
                                        tt = str(s.get("total_time") or "")
                                        if ":" in tt:
                                            try:
                                                _m, _s = tt.split(":")
                                                secs_list.append(int(_m) * 60 + int(_s))
                                            except (ValueError, IndexError):
                                                pass
                                    if secs_list:
                                        st.caption(f"Avg (last {len(secs_list)} wks): {sum(secs_list)/len(secs_list)/60:.1f} min")
                                    for s in sos_resp.data:
                                        rank = s.get("good_shift_rank", "N/A")
                                        total = s.get("total_stores", "")
                                        rank_str = f" (#{rank} of {total})" if rank != "N/A" and total else ""
                                        tt2 = str(s.get("total_time") or "")
                                        st.caption(f"Wk {s['week_start']}: {tt2}{rank_str}")
                                else:
                                    st.caption("No SoS data available")
                            except Exception:
                                st.caption("No SoS data available")

                        # VOTG data
                        with perf_cols[2]:
                            st.markdown("**⭐ Voice of the Guest**")
                            try:
                                votg_resp = _ap_sb.table("store_votg_weekly").select(
                                    "week_start,votg_rank,total_stores,total_negative_reviews"
                                ).eq("location_id", row["location_id"]).order(
                                    "week_start", desc=True
                                ).limit(4).execute()
                                if votg_resp.data:
                                    neg_vals = [v.get("total_negative_reviews", 0) for v in votg_resp.data if v.get("total_negative_reviews") is not None]
                                    if neg_vals:
                                        st.caption(f"Avg Neg Reviews (last {len(neg_vals)} wks): {sum(neg_vals)/len(neg_vals):.1f}")
                                    for v in votg_resp.data:
                                        rank = v.get("votg_rank", "N/A")
                                        total = v.get("total_stores", "")
                                        rank_str = f" (#{rank} of {total})" if rank != "N/A" and total else ""
                                        neg = v.get("total_negative_reviews", "N/A")
                                        st.caption(f"Wk {v['week_start']}: {neg} neg reviews{rank_str}")
                                else:
                                    st.caption("No VOTG data available")
                            except Exception:
                                st.caption("No VOTG data available")

                    st.divider()

# ==========================================
# Email Settings (Admin)
# ==========================================
elif page == "Email Settings":
    if not user_is_admin:
        st.error("You do not have permission to access this page.")
        st.stop()

    st.title("📧 Email Settings")
    st.caption("Configure the email workflow for GM/DM revenue band selection.")

    settings = load_app_settings()

    with st.form("email_settings_form"):
        col1, col2 = st.columns(2)
        with col1:
            gm_send_day = st.selectbox("GM Email Send Day", DAYS_OF_WEEK,
                index=DAYS_OF_WEEK.index(settings.get("gm_email_send_day", "Monday")))
            gm_deadline = st.selectbox("GM Deadline Day", DAYS_OF_WEEK,
                index=DAYS_OF_WEEK.index(settings.get("gm_deadline_day", "Wednesday")))
        with col2:
            ceo_email = st.text_input("CEO Email (for escalation)", value=settings.get("ceo_email", ""))
            # DM deadline is always GM deadline + 1 day — derived automatically
            gm_deadline_idx = DAYS_OF_WEEK.index(settings.get("gm_deadline_day", "Wednesday"))
            dm_deadline_day = DAYS_OF_WEEK[(gm_deadline_idx + 1) % 7]
            st.caption(f"DM Deadline Day: **{dm_deadline_day}** (always GM Deadline + 1)")

        st.subheader("Reminder Times")
        st.caption("Apply to reminder days (Day 2 and Day 3 of the cadence).")
        r1, r2, r3 = st.columns(3)
        with r1:
            reminder_1 = st.text_input("Reminder 1 (morning)", value=settings.get("reminder_1_time", "08:00"))
        with r2:
            reminder_2 = st.text_input("Reminder 2 (midday)", value=settings.get("reminder_2_time", "12:00"))
        with r3:
            reminder_3 = st.text_input("Reminder 3 (evening)", value=settings.get("reminder_3_time", "17:00"))

        save_btn = st.form_submit_button("Save Settings", type="primary")

    if save_btn:
        gm_deadline_idx  = DAYS_OF_WEEK.index(gm_deadline)
        dm_deadline_day  = DAYS_OF_WEEK[(gm_deadline_idx + 1) % 7]
        save_app_setting("gm_email_send_day",  gm_send_day)
        save_app_setting("gm_deadline_day",    gm_deadline)
        save_app_setting("dm_deadline_day",    dm_deadline_day)  # auto-derived
        save_app_setting("ceo_email",          ceo_email)
        save_app_setting("reminder_1_time",    reminder_1)
        save_app_setting("reminder_2_time",    reminder_2)
        save_app_setting("reminder_3_time",    reminder_3)
        st.success(f"Email settings saved! DM Deadline auto-set to {dm_deadline_day}.")
        st.cache_data.clear()
        st.rerun()

    st.divider()

    # DM Email Management
    st.subheader("DM Emails")
    st.caption("Manage DM email addresses for the approval workflow.")
    from supabase_db import get_supabase as _get_supabase
    sb = _get_supabase()
    dm_resp = sb.table("dm_list").select("*").order("dm_name").execute()
    dm_data = dm_resp.data if dm_resp.data else []

    if dm_data:
        for dm in dm_data:
            c1, c2 = st.columns([1, 2])
            with c1:
                st.markdown(f"**{dm['dm_name']}**")
            with c2:
                new_email = st.text_input(f"Email for {dm['dm_name']}", value=dm.get("email", "") or "",
                                          key=f"dm_email_{dm['dm_name']}")
                if new_email != (dm.get("email", "") or ""):
                    if st.button(f"Save", key=f"save_dm_email_{dm['dm_name']}"):
                        sb.table("dm_list").update({"email": new_email}).eq("dm_name", dm["dm_name"]).execute()
                        st.success(f"Updated email for {dm['dm_name']}")
                        st.cache_data.clear()
                        st.rerun()
    else:
        st.info("No DMs found. Add DMs in the DM Assignments page first.")

# ==========================================
# Compliance Report (Admin)
# ==========================================
elif page == "Compliance Report":
    if not user_is_admin:
        st.error("You do not have permission to access this page.")
        st.stop()

    st.title("📋 Compliance Report")
    st.caption(
        "GM submission timeliness and DM approval compliance. "
        "On Time = GM submitted before Tuesday 8am · DM On Time = approved before Wednesday 8am."
    )

    # --- Load base data ---
    ref_data   = _cached_reference_data()
    all_subs   = load_all_submissions()
    email_log_df = load_email_log()

    if all_subs.empty and email_log_df.empty:
        st.info("No compliance data yet. Data will appear once GMs begin submitting revenue bands.")
        st.stop()

    # --- Build unified week list ---
    weeks_set = set()
    if not all_subs.empty:
        weeks_set.update(all_subs["week_start"].dropna().unique())
    if not email_log_df.empty:
        weeks_set.update(email_log_df["week_start"].dropna().unique())

    all_week_dates = sorted(
        [pd.Timestamp(w).date() for w in weeks_set],
        reverse=False,
    )

    if not all_week_dates:
        st.info("No week data available yet.")
        st.stop()

    # --- Filters ---
    filtered_week_dates = _period_filter(all_week_dates, "compliance")
    filtered_week_strs  = [str(w) for w in filtered_week_dates]

    col_f1, col_f2 = st.columns(2)
    dm_options = sorted(ref_data["dm"].dropna().unique().tolist()) if not ref_data.empty else []
    with col_f1:
        dm_filter = st.selectbox("Filter by DM", ["All"] + dm_options, key="compliance_dm")
    with col_f2:
        status_filter = st.selectbox(
            "Filter by GM Status", ["All", "On Time", "Late", "Never Submitted"],
            key="compliance_status",
        )

    # --- Build compliance rows ---
    stores = (
        ref_data[["location_id", "store_name", "dm"]].drop_duplicates()
        if not ref_data.empty
        else pd.DataFrame(columns=["location_id", "store_name", "dm"])
    )

    compliance_rows = []

    for week_str in filtered_week_strs:
        try:
            week_date = pd.Timestamp(week_str).date()
        except Exception:
            continue

        # Week starts Thursday; +5 days = Tuesday, +6 = Wednesday
        gm_deadline  = datetime(week_date.year, week_date.month, week_date.day) + timedelta(days=5, hours=8)
        dm_deadline  = datetime(week_date.year, week_date.month, week_date.day) + timedelta(days=6, hours=8)

        week_log  = email_log_df[email_log_df["week_start"] == week_str].copy() if not email_log_df.empty else pd.DataFrame()
        week_subs = all_subs[all_subs["week_start"] == week_str].copy() if not all_subs.empty else pd.DataFrame()

        for _, store_row in stores.iterrows():
            loc_id     = store_row["location_id"]
            store_name = store_row["store_name"]
            dm_name    = store_row.get("dm", "")

            if dm_filter != "All" and dm_name != dm_filter:
                continue

            # ── Email log for this store ──────────────────────────────────
            store_log = week_log[week_log["location_id"] == loc_id] if not week_log.empty else pd.DataFrame()
            if not store_log.empty and "email_type" in store_log.columns:
                initial_sent   = (store_log["email_type"] == "initial").any()
                reminder_count = int((store_log["email_type"].isin(["tuesday", "wednesday"])).sum())
            else:
                initial_sent   = False
                reminder_count = 0

            # ── Submission for this store ─────────────────────────────────
            store_sub = week_subs[week_subs["location_id"] == loc_id] if not week_subs.empty else pd.DataFrame()

            if store_sub.empty:
                gm_status      = "Never Submitted"
                gm_on_time     = False
                submitted_str  = ""
                final_status   = "Pending GM"
                approved_str   = ""
                dm_on_time_val = False
            else:
                sub = store_sub.iloc[0]
                final_status = sub.get("status", "")
                if isinstance(final_status, str):
                    final_status = final_status.replace("_", " ").title()

                # Only count as submitted if GM actually selected a band (status != pending_gm)
                raw_status   = sub.get("status", "")
                gm_submitted = raw_status != "pending_gm" and bool(sub.get("selected_band"))
                submitted_at = sub.get("submitted_at") if gm_submitted else None
                if submitted_at:
                    try:
                        sub_dt        = pd.Timestamp(submitted_at).to_pydatetime().replace(tzinfo=None)
                        submitted_str = str(submitted_at)[:16]
                        gm_on_time    = sub_dt < gm_deadline
                        gm_status     = "On Time" if gm_on_time else "Late"
                    except Exception:
                        submitted_str = str(submitted_at)[:16]
                        gm_status     = "Unknown"
                        gm_on_time    = False
                else:
                    submitted_str = ""
                    gm_status     = "Never Submitted"
                    gm_on_time    = False

                approved_at = sub.get("admin_approved_at")
                if approved_at:
                    try:
                        app_dt        = pd.Timestamp(approved_at).to_pydatetime().replace(tzinfo=None)
                        approved_str  = str(approved_at)[:16]
                        dm_on_time_val = app_dt < dm_deadline
                    except Exception:
                        approved_str   = str(approved_at)[:16]
                        dm_on_time_val = False
                else:
                    approved_str   = ""
                    dm_on_time_val = False

            if status_filter != "All" and gm_status != status_filter:
                continue

            compliance_rows.append({
                "Week":               week_str,
                "Store":              store_name,
                "DM":                 dm_name,
                "Initial Email Sent": "✅" if initial_sent else "—",
                "Reminders Sent":     reminder_count,
                "GM Submitted At":    submitted_str,
                "GM Status":          gm_status,
                "Final Status":       final_status,
                "Approved At":        approved_str,
                "DM Approved On Time": "✅" if dm_on_time_val else ("—" if not approved_str else "❌"),
                # Raw booleans for DM rollup
                "_gm_on_time":        gm_on_time,
                "_dm_on_time":        dm_on_time_val,
            })

    if not compliance_rows:
        st.info("No compliance data matches the selected filters.")
        st.stop()

    compliance_df = pd.DataFrame(compliance_rows)

    # --- Summary Metrics ---
    total    = len(compliance_df)
    on_time  = int((compliance_df["GM Status"] == "On Time").sum())
    late     = int((compliance_df["GM Status"] == "Late").sum())
    never    = int((compliance_df["GM Status"] == "Never Submitted").sum())
    avg_rem  = compliance_df["Reminders Sent"].mean()

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Total Stores",     total)
    m2.metric("On Time",          f"{on_time} ({on_time / total * 100:.0f}%)")
    m3.metric("Late",             f"{late} ({late / total * 100:.0f}%)")
    m4.metric("Never Submitted",  f"{never} ({never / total * 100:.0f}%)")
    m5.metric("Avg Reminders",    f"{avg_rem:.1f}")

    st.divider()

    # --- GM Detail Table ---
    st.subheader("GM Submission Detail")
    display_cols = [
        "Week", "Store", "DM", "Initial Email Sent", "Reminders Sent",
        "GM Submitted At", "GM Status", "Final Status",
        "Approved At", "DM Approved On Time",
    ]
    st.dataframe(compliance_df[display_cols], use_container_width=True, hide_index=True)

    st.divider()

    # --- DM Portfolio Rollup ---
    st.subheader("DM Portfolio Summary")
    dm_rollup = (
        compliance_df.groupby("DM")
        .agg(
            Stores          = ("Store",       "count"),
            On_Time         = ("_gm_on_time", "sum"),
            Late            = ("GM Status",   lambda x: (x == "Late").sum()),
            Never_Submitted = ("GM Status",   lambda x: (x == "Never Submitted").sum()),
            Total_Reminders = ("Reminders Sent", "sum"),
            DM_On_Time      = ("_dm_on_time", "sum"),
        )
        .reset_index()
    )
    dm_rollup["On Time %"] = (
        (dm_rollup["On_Time"].astype(float) / dm_rollup["Stores"].astype(float) * 100)
        .round(0).astype(int).astype(str) + "%"
    )
    dm_rollup = dm_rollup.rename(columns={
        "On_Time":         "On Time",
        "Never_Submitted": "Never Submitted",
        "Total_Reminders": "Total Reminders",
        "DM_On_Time":      "DM Approved On Time",
    })
    rollup_cols = ["DM", "Stores", "On Time", "Late", "Never Submitted",
                   "On Time %", "Total Reminders", "DM Approved On Time"]
    st.dataframe(dm_rollup[rollup_cols], use_container_width=True, hide_index=True)

    st.divider()

    # --- Excel Export (only generated on click) ---
    st.subheader("Export")
    st.caption("Click the button below to generate the Excel file — it is not built automatically to keep the page fast.")
    if st.button("📥 Generate Excel Report", type="primary"):
        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            # Sheet 1: GM Detail
            compliance_df[display_cols].to_excel(writer, sheet_name="GM Compliance", index=False)
            # Sheet 2: DM Rollup
            dm_rollup[rollup_cols].to_excel(writer, sheet_name="DM Summary", index=False)

            workbook   = writer.book
            hdr_fmt    = workbook.add_format({
                "bold": True, "bg_color": "#2B3A4E",
                "font_color": "#FFFFFF", "border": 1,
            })
            on_time_fmt = workbook.add_format({"bg_color": "#D5F5E3"})
            late_fmt    = workbook.add_format({"bg_color": "#FADBD8"})

            for sheet_name in ["GM Compliance", "DM Summary"]:
                ws = writer.sheets[sheet_name]
                ws.set_row(0, 18, hdr_fmt)
                ws.set_column("A:Z", 18)

        output.seek(0)
        period_label = (
            f"{filtered_week_strs[0]}_to_{filtered_week_strs[-1]}"
            if len(filtered_week_strs) > 1
            else (filtered_week_strs[0] if filtered_week_strs else "all")
        )
        st.download_button(
            label="⬇️ Download Excel",
            data=output,
            file_name=f"compliance_report_{period_label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ==========================================
# Rev Band Report (Admin)
# ==========================================
elif page == "Rev Band Report":
    if not user_is_admin:
        st.error("You do not have permission to access this page.")
        st.stop()

    st.title("🎯 Rev Band Hit/Miss Report")
    st.caption(
        "Compare GM revenue band predictions against actual weekly sales. "
        "Green = Over (outperformed) · Red = Under (missed) · Neutral = On Target."
    )

    # --- Load data ---
    ref_data  = _cached_reference_data()
    all_subs  = load_all_submissions()
    sales_df  = _cached_store_sales()

    if all_subs.empty:
        st.info("No submissions yet. This report will populate once GMs begin selecting revenue bands.")
        st.stop()

    # --- Build week list from submissions ---
    all_week_dates = sorted(
        {pd.Timestamp(w).date() for w in all_subs["week_start"].dropna().unique()},
    )

    # --- Period filter ---
    filtered_week_dates = _period_filter(all_week_dates, "revband")
    filtered_week_strs  = [str(w) for w in filtered_week_dates]

    # --- DM + Result filters ---
    col_f1, col_f2 = st.columns(2)
    dm_options = sorted(ref_data["dm"].dropna().unique().tolist()) if not ref_data.empty else []
    with col_f1:
        dm_filter = st.selectbox("Filter by DM", ["All"] + dm_options, key="revband_dm")
    with col_f2:
        result_filter = st.selectbox(
            "Filter by Result", ["All", "On Target", "Over", "Under", "N/A (NRO)"],
            key="revband_result",
        )

    # --- Join submissions → sales → ref_data ---
    week_subs = all_subs[all_subs["week_start"].isin(filtered_week_strs)].copy()

    # Merge store info
    if not ref_data.empty:
        week_subs = week_subs.merge(
            ref_data[["location_id", "store_name", "dm"]],
            on="location_id", how="left",
        )

    # Merge sales
    if not sales_df.empty:
        week_subs = week_subs.merge(
            sales_df[["location_id", "week_start", "net_sales"]],
            on=["location_id", "week_start"], how="left",
        )
    else:
        week_subs["net_sales"] = None

    # --- Apply band classification ---
    rows = []
    for _, r in week_subs.iterrows():
        band   = r.get("selected_band", "")
        actual = r.get("net_sales")
        dm     = r.get("dm", "")

        if dm_filter != "All" and dm != dm_filter:
            continue

        try:
            actual_num = float(actual) if actual is not None else None
        except (TypeError, ValueError):
            actual_num = None

        result, variance = _band_classify(band, actual_num)
        variance_fmt = _fmt_variance(result, variance)

        if result_filter != "All":
            mapped = "N/A (NRO)" if result == "N/A" else result
            if mapped != result_filter:
                continue

        band_min, band_max = BAND_RANGES.get(band, (None, None)) or (None, None)
        band_range_str = (
            "N/A"
            if band_min is None
            else (
                f"${band_min:,.0f}+"
                if band_max == float("inf")
                else f"${band_min:,.0f} – ${band_max:,.0f}"
            )
        )

        rows.append({
            "Week":          r.get("week_start", ""),
            "Store":         r.get("store_name", r.get("location_id", "")),
            "DM":            dm,
            "Selected Band": band,
            "Band Range":    band_range_str,
            "Actual Sales":  f"${actual_num:,.0f}" if actual_num is not None else "No Data",
            "Result":        result,
            "$ Variance":    variance_fmt,
            "_variance_raw": variance if variance is not None else 0.0,
            "_result_raw":   result,
        })

    if not rows:
        st.info("No data matches the selected filters.")
        st.stop()

    report_df = pd.DataFrame(rows)

    # --- Summary Metrics ---
    has_sales = report_df[report_df["Actual Sales"] != "No Data"]
    total      = len(has_sales)
    on_target  = int((has_sales["Result"] == "On Target").sum())
    over       = int((has_sales["Result"] == "Over").sum())
    under      = int((has_sales["Result"] == "Under").sum())
    no_data    = int((report_df["Actual Sales"] == "No Data").sum())
    avg_var    = has_sales["_variance_raw"].mean() if total else 0

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("With Sales Data",  total)
    m2.metric("On Target",        f"{on_target} ({on_target/total*100:.0f}%)" if total else "—")
    m3.metric("Over",             f"{over} ({over/total*100:.0f}%)"           if total else "—")
    m4.metric("Under",            f"{under} ({under/total*100:.0f}%)"         if total else "—")
    m5.metric("Avg $ Variance",   f"${avg_var:+,.0f}"                         if total else "—")

    if no_data:
        st.caption(f"⚠ {no_data} store/week(s) have submissions but no sales data loaded yet.")

    st.divider()

    # --- Main table with color coding ---
    st.subheader("Store Detail")
    display_cols = ["Week", "Store", "DM", "Selected Band", "Band Range",
                    "Actual Sales", "Result", "$ Variance"]

    def _color_result(row):
        result = row["_result_raw"]
        styles = [""] * len(row)
        for i, col in enumerate(row.index):
            if col in ("Result", "$ Variance"):
                if result == "Over":
                    styles[i] = "background-color: #D5F5E3; color: #1E8449"
                elif result == "Under":
                    styles[i] = "background-color: #FADBD8; color: #922B21"
        return styles

    styled = report_df[display_cols + ["_result_raw"]].style.apply(_color_result, axis=1)
    st.dataframe(
        styled,
        use_container_width=True,
        hide_index=True,
        column_config={"_result_raw": None},  # hide raw column
    )

    st.divider()

    # --- DM Rollup ---
    st.subheader("DM Summary")
    dm_grp = has_sales.groupby("DM").agg(
        Stores      = ("Store",         "count"),
        On_Target   = ("Result",        lambda x: (x == "On Target").sum()),
        Over        = ("Result",        lambda x: (x == "Over").sum()),
        Under       = ("Result",        lambda x: (x == "Under").sum()),
        Avg_Variance= ("_variance_raw", "mean"),
    ).reset_index()
    dm_grp["On Time %"]     = (dm_grp["On_Target"].astype(float) / dm_grp["Stores"].astype(float) * 100).round(0).astype(int).astype(str) + "%"
    dm_grp["Avg $ Variance"]= dm_grp["Avg_Variance"].astype(float).apply(lambda v: f"${v:+,.0f}")
    dm_grp = dm_grp.rename(columns={"On_Target": "On Target"})

    st.dataframe(
        dm_grp[["DM", "Stores", "On Target", "Over", "Under", "On Time %", "Avg $ Variance"]],
        use_container_width=True,
        hide_index=True,
    )

    st.divider()

    # --- Excel Export (generated on click only) ---
    st.subheader("Export")
    st.caption("Click to generate — not built automatically to keep the page fast.")
    if st.button("📥 Generate Excel Report", type="primary", key="revband_export"):
        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            # Sheet 1: Store Detail
            report_df[display_cols].to_excel(writer, sheet_name="Store Detail", index=False)
            # Sheet 2: DM Summary
            dm_grp[["DM", "Stores", "On Target", "Over", "Under", "On Time %", "Avg $ Variance"]].to_excel(
                writer, sheet_name="DM Summary", index=False,
            )

            workbook  = writer.book
            hdr_fmt   = workbook.add_format({"bold": True, "bg_color": "#2B3A4E", "font_color": "#FFFFFF", "border": 1})
            green_fmt = workbook.add_format({"bg_color": "#D5F5E3", "font_color": "#1E8449"})
            red_fmt   = workbook.add_format({"bg_color": "#FADBD8", "font_color": "#922B21"})

            # Style Store Detail sheet
            ws = writer.sheets["Store Detail"]
            ws.set_row(0, 18, hdr_fmt)
            ws.set_column("A:H", 18)

            # Color Result + $ Variance columns per row
            result_col_idx   = display_cols.index("Result")
            variance_col_idx = display_cols.index("$ Variance")
            for row_idx, (_, row_data) in enumerate(report_df[display_cols].iterrows(), start=1):
                result_val = row_data.get("Result", "")
                fmt = green_fmt if result_val == "Over" else (red_fmt if result_val == "Under" else None)
                if fmt:
                    ws.write(row_idx, result_col_idx,   result_val,                  fmt)
                    ws.write(row_idx, variance_col_idx, row_data.get("$ Variance", ""), fmt)

            # Style DM Summary sheet
            ws2 = writer.sheets["DM Summary"]
            ws2.set_row(0, 18, hdr_fmt)
            ws2.set_column("A:G", 18)

        output.seek(0)
        period_label = (
            f"{filtered_week_strs[0]}_to_{filtered_week_strs[-1]}"
            if len(filtered_week_strs) > 1
            else (filtered_week_strs[0] if filtered_week_strs else "all")
        )
        st.download_button(
            label="⬇️ Download Excel",
            data=output,
            file_name=f"revband_report_{period_label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="revband_download",
        )

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Upload SoS  (Admin)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Upload SoS":
    if not user_is_admin:
        st.error("You do not have permission to access this page.")
        st.stop()

    st.title("⏱️ Upload Speed of Service")
    st.caption(
        "Upload the weekly Freddy's Times report. Data is saved to the forecasting database "
        "and the week's SoS status is marked complete."
    )

    from supabase_db import get_supabase as _get_sb

    # --- Week selector ---
    _sos_default_ws = get_week_start()
    sos_week_start = st.date_input(
        "Week Start (Thursday)",
        value=_sos_default_ws,
        key="sos_week_start",
        help="Select the Thursday that begins the week covered by this report.",
    )
    sos_week_str = str(sos_week_start)

    # --- Current upload status for chosen week ---
    _sos_status = _get_upload_status(sos_week_str)
    s1, s2, s3 = st.columns(3)
    s1.metric("Sales", "✅ Uploaded" if _sos_status.get("sales_uploaded") else "⬜ Pending")
    s2.metric("SoS",   "✅ Uploaded" if _sos_status.get("sos_uploaded")   else "⬜ Pending")
    s3.metric("VOTG",  "✅ Uploaded" if _sos_status.get("votg_uploaded")  else "⬜ Pending")

    st.divider()

    _sos_ctr = st.session_state["sos_upload_ctr"]
    sos_file = st.file_uploader(
        "Freddy's Times Report (.xlsx)",
        type=["xlsx"],
        key=f"sos_file_{_sos_ctr}",
        help="The weekly 'Times' spreadsheet exported from the Freddy's ops portal.",
    )

    if st.button("Upload SoS Data", type="primary", use_container_width=True, key="sos_run"):
        if sos_file is None:
            st.error("Please upload the SoS (.xlsx) file first.")
            st.stop()

        try:
            with st.spinner("Parsing SoS file..."):
                raw = pd.read_excel(sos_file, header=None)

                # Row 0 = column headers, Row 1 = "Totals:" — skip both
                data = raw.iloc[2:].reset_index(drop=True)

                ref_data = _cached_reference_data()
                active_ref = ref_data[ref_data["active"] == True][["location_id", "store_name"]].drop_duplicates()

                rows, unmatched = [], []

                for _, r in data.iterrows():
                    store_name = str(r.iloc[0]).strip()
                    if not store_name or store_name.lower() == "nan":
                        continue

                    loc_id = _match_store_name(store_name, active_ref)
                    if not loc_id:
                        unmatched.append(store_name)
                        continue

                    # Parse rank "487 of 532"
                    rank_raw = str(r.iloc[2]) if pd.notna(r.iloc[2]) else ""
                    sos_rank, total_stores = None, None
                    if " of " in rank_raw:
                        parts = rank_raw.split(" of ")
                        try:
                            sos_rank    = int(parts[0].strip())
                            total_stores = int(parts[1].strip())
                        except ValueError:
                            pass

                    good_shift   = float(r.iloc[1]) if pd.notna(r.iloc[1]) else None
                    total_time   = str(r.iloc[3]) if pd.notna(r.iloc[3]) else None
                    red_ticket   = float(r.iloc[4]) if pd.notna(r.iloc[4]) else None
                    shift_streak = int(r.iloc[6])   if pd.notna(r.iloc[6]) else None

                    rows.append({
                        "location_id":    loc_id,
                        "week_start":     sos_week_str,
                        "good_shift":     good_shift,
                        "good_shift_rank": sos_rank,
                        "total_stores":   total_stores,
                        "total_time":     total_time,
                        "red_ticket":     red_ticket,
                        "shift_streak":   shift_streak,
                    })

            if not rows:
                st.error("No store rows could be matched. Check that the file format is correct.")
                st.stop()

            with st.spinner(f"Saving {len(rows)} store records..."):
                sb = _get_sb()
                sb.table("store_sos_weekly").upsert(
                    rows, on_conflict="location_id,week_start"
                ).execute()
                _mark_upload_status(sos_week_str, sos_uploaded=True)

            st.session_state["sos_upload_ctr"] += 1
            st.cache_data.clear()
            st.success(f"✅ SoS data saved for {len(rows)} stores (week of {sos_week_str}).")

            if unmatched:
                st.warning(
                    f"**{len(unmatched)} store(s) could not be matched** — review and re-upload if needed:\n\n"
                    + "\n".join(f"• {n}" for n in unmatched)
                )

            # Show updated status
            _new_status = _get_upload_status(sos_week_str)
            if all([
                _new_status.get("sales_uploaded"),
                _new_status.get("sos_uploaded"),
                _new_status.get("votg_uploaded"),
            ]):
                st.success(
                    "🎉 All three uploads complete for this week! "
                    "The Monday forecast job will run the back-test automatically."
                )
            st.rerun()

        except Exception as _e:
            st.error(f"Upload failed: {_e}")

    # --- Preview last upload for this week ---
    st.divider()
    st.subheader("Last Upload Preview")
    try:
        sb = _get_sb()
        _preview_resp = sb.table("store_sos_weekly").select("*").eq("week_start", sos_week_str).execute()
        if _preview_resp.data:
            _prev_df = pd.DataFrame(_preview_resp.data)[
                ["location_id", "good_shift", "good_shift_rank", "total_stores", "total_time", "red_ticket", "shift_streak"]
            ].rename(columns={
                "location_id":    "Store #",
                "good_shift":     "Good Shift",
                "good_shift_rank": "Rank",
                "total_stores":   "Total Stores",
                "total_time":     "Total Time",
                "red_ticket":     "Red Ticket",
                "shift_streak":   "Shift Streak",
            })
            st.dataframe(_prev_df, use_container_width=True, hide_index=True)
        else:
            st.info("No SoS data uploaded for this week yet.")
    except Exception:
        st.info("No SoS data uploaded for this week yet.")


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Upload VOTG  (Admin)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Upload VOTG":
    if not user_is_admin:
        st.error("You do not have permission to access this page.")
        st.stop()

    st.title("⭐ Upload Voice of the Guest")
    st.caption(
        "Upload the weekly Freddy's VOTG Score Table. Data is saved to the forecasting database "
        "and the week's VOTG status is marked complete."
    )

    from supabase_db import get_supabase as _get_sb

    # --- Week selector ---
    _votg_default_ws = get_week_start()
    votg_week_start = st.date_input(
        "Week Start (Thursday)",
        value=_votg_default_ws,
        key="votg_week_start",
        help="Select the Thursday that begins the week covered by this report.",
    )
    votg_week_str = str(votg_week_start)

    # --- Current upload status ---
    _votg_status = _get_upload_status(votg_week_str)
    v1, v2, v3 = st.columns(3)
    v1.metric("Sales", "✅ Uploaded" if _votg_status.get("sales_uploaded") else "⬜ Pending")
    v2.metric("SoS",   "✅ Uploaded" if _votg_status.get("sos_uploaded")   else "⬜ Pending")
    v3.metric("VOTG",  "✅ Uploaded" if _votg_status.get("votg_uploaded")  else "⬜ Pending")

    st.divider()

    _votg_ctr = st.session_state["votg_upload_ctr"]
    votg_file = st.file_uploader(
        "Freddy's VOTG Score Table (.xlsx)",
        type=["xlsx"],
        key=f"votg_file_{_votg_ctr}",
        help="The weekly 'VOTG Score Table' spreadsheet from the Freddy's ops portal.",
    )

    if st.button("Upload VOTG Data", type="primary", use_container_width=True, key="votg_run"):
        if votg_file is None:
            st.error("Please upload the VOTG (.xlsx) file first.")
            st.stop()

        try:
            with st.spinner("Parsing VOTG file..."):
                raw = pd.read_excel(votg_file, header=None)

                # Row 0 = column headers; data starts at row 1
                data = raw.iloc[1:].reset_index(drop=True)

                ref_data = _cached_reference_data()
                active_ref = ref_data[ref_data["active"] == True][["location_id", "store_name"]].drop_duplicates()

                rows, unmatched = [], []

                for _, r in data.iterrows():
                    store_name = str(r.iloc[0]).strip()
                    if not store_name or store_name.lower() == "nan":
                        continue

                    loc_id = _match_store_name(store_name, active_ref)
                    if not loc_id:
                        unmatched.append(store_name)
                        continue

                    # Parse rank "532 of 532"
                    rank_raw = str(r.iloc[2]) if pd.notna(r.iloc[2]) else ""
                    votg_rank, total_stores = None, None
                    if " of " in rank_raw:
                        parts = rank_raw.split(" of ")
                        try:
                            votg_rank    = int(parts[0].strip())
                            total_stores = int(parts[1].strip())
                        except ValueError:
                            pass

                    total_reviews        = int(r.iloc[1])   if pd.notna(r.iloc[1]) else None
                    total_neg_reviews    = int(r.iloc[3])   if pd.notna(r.iloc[3]) else None
                    guests_per_negative  = float(r.iloc[4]) if pd.notna(r.iloc[4]) else None

                    rows.append({
                        "location_id":         loc_id,
                        "week_start":          votg_week_str,
                        "guests_per_negative": guests_per_negative,
                        "votg_rank":           votg_rank,
                        "total_stores":        total_stores,
                        "total_negative_reviews": total_neg_reviews,
                        "total_reviews":       total_reviews,
                    })

            if not rows:
                st.error("No store rows could be matched. Check that the file format is correct.")
                st.stop()

            with st.spinner(f"Saving {len(rows)} store records..."):
                sb = _get_sb()
                sb.table("store_votg_weekly").upsert(
                    rows, on_conflict="location_id,week_start"
                ).execute()
                _mark_upload_status(votg_week_str, votg_uploaded=True)

            st.session_state["votg_upload_ctr"] += 1
            st.cache_data.clear()
            st.success(f"✅ VOTG data saved for {len(rows)} stores (week of {votg_week_str}).")

            if unmatched:
                st.warning(
                    f"**{len(unmatched)} store(s) could not be matched** — review and re-upload if needed:\n\n"
                    + "\n".join(f"• {n}" for n in unmatched)
                )

            # Show updated status
            _new_status = _get_upload_status(votg_week_str)
            if all([
                _new_status.get("sales_uploaded"),
                _new_status.get("sos_uploaded"),
                _new_status.get("votg_uploaded"),
            ]):
                st.success(
                    "🎉 All three uploads complete for this week! "
                    "The Monday forecast job will run the back-test automatically."
                )
            st.rerun()

        except Exception as _e:
            st.error(f"Upload failed: {_e}")

    # --- Preview last upload for this week ---
    st.divider()
    st.subheader("Last Upload Preview")
    try:
        sb = _get_sb()
        _preview_resp = sb.table("store_votg_weekly").select("*").eq("week_start", votg_week_str).execute()
        if _preview_resp.data:
            _prev_df = pd.DataFrame(_preview_resp.data)[
                ["location_id", "guests_per_negative", "votg_rank", "total_stores",
                 "total_negative_reviews", "total_reviews"]
            ].rename(columns={
                "location_id":            "Store #",
                "guests_per_negative":    "Guests/Neg",
                "votg_rank":              "Rank",
                "total_stores":           "Total Stores",
                "total_negative_reviews": "Neg Reviews",
                "total_reviews":          "Total Reviews",
            })
            st.dataframe(_prev_df, use_container_width=True, hide_index=True)
        else:
            st.info("No VOTG data uploaded for this week yet.")
    except Exception:
        st.info("No VOTG data uploaded for this week yet.")


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Sales Forecasts  (Admin)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Sales Forecasts":
    if not user_is_admin:
        st.error("You do not have permission to access this page.")
        st.stop()

    st.title("📈 Sales Forecasts")
    st.caption("View weekly store forecasts and run manual back-tests for weeks where all three uploads are complete.")

    from supabase_db import get_supabase as _get_sb

    _SF_STATE_MAP = {
        '112-0001': 'OK', '112-0002': 'OK', '112-0003': 'OK', '112-0004': 'OK',
        '112-0005': 'OK', '112-0031': 'OK', '112-0035': 'OK', '112-0038': 'OK',
        '112-0006': 'TX', '112-0007': 'TX', '112-0008': 'TX', '112-0009': 'TX',
        '112-0010': 'TX', '112-0011': 'TX', '112-0032': 'TX', '112-0033': 'TX',
        '112-0034': 'TX', '112-0036': 'TX', '112-0037': 'TX',
        '112-0012': 'OH', '112-0013': 'OH', '112-0014': 'OH', '112-0015': 'OH',
        '112-0016': 'OH', '112-0017': 'OH', '112-0018': 'OH', '112-0019': 'OH',
        '112-0020': 'OH', '112-0021': 'OH', '112-0022': 'OH', '112-0023': 'OH',
        '112-0024': 'OH', '112-0025': 'OH', '112-0026': 'OH', '112-0028': 'OH',
        '112-0029': 'OH', '112-0030': 'OH',
        '112-0027': 'KY',
    }

    _SF_BANDS = [
        ('<25k',    0,     25000),
        ('25k-30k', 25000, 30000),
        ('30k-35k', 30000, 35000),
        ('35k-40k', 35000, 40000),
        ('40k-45k', 40000, 45000),
        ('45k-50k', 45000, 50000),
        ('50k+',    50000, 9_999_999),
    ]

    def _sf_get_band(sales):
        for name, lo, hi in _SF_BANDS:
            if lo <= sales < hi:
                return name
        return '50k+'

    def _run_backtest_for_week(week_monday_str: str):
        """Back-fill actuals onto forecast rows for a given Monday week_start."""
        sb = _get_sb()
        wk_start = pd.Timestamp(week_monday_str)
        wk_end   = wk_start + timedelta(days=6)

        fc_resp = sb.table("sales_forecasts").select("*").eq("week_start", week_monday_str).execute()
        forecast_rows = fc_resp.data or []
        if not forecast_rows:
            return 0, "No forecast rows found for this week."

        # Load gas prices once
        gas_resp = sb.table("gas_price_history").select("state,week_start,price_per_gallon").execute()
        gas_df = pd.DataFrame(gas_resp.data or [])
        if not gas_df.empty:
            gas_df['week_start'] = pd.to_datetime(gas_df['week_start'])
            gas_df['price_per_gallon'] = pd.to_numeric(gas_df['price_per_gallon'], errors='coerce')

        updated = 0
        for fc in forecast_rows:
            loc_id = fc['location_id']

            # Actual sales
            s_resp = sb.table("store_sales").select("net_sales").eq(
                "location_id", loc_id
            ).gte("sale_date", str(wk_start.date())).lte("sale_date", str(wk_end.date())).execute()
            actual_sales = sum(float(r['net_sales']) for r in s_resp.data) if s_resp.data else None

            # Actual weather
            w_resp = sb.table("weather_history").select(
                "temp_high_f,temp_low_f,precipitation_in"
            ).eq("location_id", loc_id).eq("is_forecast", False).gte(
                "date", str(wk_start.date())
            ).lte("date", str(wk_end.date())).execute()
            if w_resp.data:
                wdf = pd.DataFrame(w_resp.data)
                actual_temp_high = float(pd.to_numeric(wdf['temp_high_f'], errors='coerce').mean())
                actual_temp_low  = float(pd.to_numeric(wdf['temp_low_f'],  errors='coerce').mean())
                actual_precip    = float(pd.to_numeric(wdf['precipitation_in'], errors='coerce').sum())
            else:
                actual_temp_high = actual_temp_low = actual_precip = None

            # Actual gas
            actual_gas = None
            if not gas_df.empty:
                state = _SF_STATE_MAP.get(loc_id, 'OH')
                state_gas = gas_df[gas_df['state'] == state].sort_values('week_start')
                past_gas  = state_gas[state_gas['week_start'] <= wk_start]
                if len(past_gas):
                    actual_gas = float(past_gas['price_per_gallon'].iloc[-1])

            # Error metrics
            fc_point = float(fc['forecast_point']) if fc.get('forecast_point') is not None else None
            fc_low   = float(fc['forecast_low'])   if fc.get('forecast_low')   is not None else None
            fc_high  = float(fc['forecast_high'])  if fc.get('forecast_high')  is not None else None
            rec_band = fc.get('recommended_band')

            forecast_error     = round(actual_sales - fc_point, 2) if actual_sales and fc_point else None
            forecast_error_pct = round((forecast_error / actual_sales) * 100, 2) if forecast_error and actual_sales else None
            band_hit           = (_sf_get_band(actual_sales) == rec_band) if actual_sales and rec_band else None
            within_ci          = bool(fc_low <= actual_sales <= fc_high) if all([fc_low, fc_high, actual_sales]) else None

            sb.table("sales_forecasts").update({
                'actual_sales':              actual_sales,
                'actual_temp_high':          actual_temp_high,
                'actual_temp_low':           actual_temp_low,
                'actual_precip':             actual_precip,
                'actual_gas_price':          actual_gas,
                'forecast_error':            forecast_error,
                'forecast_error_pct':        forecast_error_pct,
                'band_hit':                  band_hit,
                'within_confidence_interval': within_ci,
            }).eq("id", fc['id']).execute()
            updated += 1

        return updated, None

    # ── Load available forecast weeks (cached) ───────────────────────────────
    sb = _get_sb()

    @st.cache_data(ttl=300)
    def _get_forecast_weeks():
        _sb = _get_sb()
        resp = _sb.table("sales_forecasts").select("week_start").execute()
        return sorted({r['week_start'] for r in (resp.data or [])}, reverse=True)

    try:
        fc_weeks = _get_forecast_weeks()
    except Exception:
        fc_weeks = []

    if not fc_weeks:
        st.info("No forecast data yet. Forecasts are generated each Monday morning by the automated job.")
        st.stop()

    # ── Week selector ──────────────────────────────────────────────────────────
    selected_fc_week = st.selectbox(
        "Model Output Date (week starting)",
        fc_weeks,
        format_func=lambda w: f"Week of {date.fromisoformat(w).strftime('%-m/%-d/%y')}",
        key="sf_week_select",
    )

    # ── Forecast table for selected week ───────────────────────────────────────
    try:
        _fc_thu  = date.fromisoformat(selected_fc_week)
        _fc_wed  = _fc_thu + timedelta(days=6)
        _fc_label = f"{_fc_thu.strftime('%-m/%-d')} – {_fc_wed.strftime('%-m/%-d/%y')}"
    except Exception:
        _fc_label = selected_fc_week

    st.markdown(f"**Model Output Date:** week of {date.fromisoformat(selected_fc_week).strftime('%-m/%-d/%y') if selected_fc_week else ''}")
    st.subheader(f"Sales Forecast: {_fc_label}")
    try:
        fc_resp = sb.table("sales_forecasts").select(
            "location_id,store_name,recommended_band,forecast_low,forecast_point,"
            "forecast_high,confidence_pct,actual_sales,forecast_error_pct,band_hit"
        ).eq("week_start", selected_fc_week).order("location_id").execute()

        if fc_resp.data:
            fc_df = pd.DataFrame(fc_resp.data).rename(columns={
                "location_id":       "Store #",
                "store_name":        "Store",
                "recommended_band":  "Band",
                "forecast_low":      "Low",
                "forecast_point":    "Point",
                "forecast_high":     "High",
                "confidence_pct":    "Conf %",
                "actual_sales":      "Actual",
                "forecast_error_pct":"Error %",
                "band_hit":          "Band Hit",
            })
            # Compute totals before formatting converts to strings
            _totals = {col: pd.to_numeric(fc_df[col], errors="coerce").sum()
                       for col in ["Low", "Point", "High", "Actual"] if col in fc_df.columns}

            for col in ["Low", "Point", "High", "Actual"]:
                if col in fc_df.columns:
                    fc_df[col] = fc_df[col].apply(
                        lambda x: f"${x:,.0f}" if pd.notna(x) and x != "" else "—"
                    )
            if "Error %" in fc_df.columns:
                fc_df["Error %"] = fc_df["Error %"].apply(
                    lambda x: f"{x:+.1f}%" if pd.notna(x) and x != "" else "—"
                )
            if "Band Hit" in fc_df.columns:
                fc_df["Band Hit"] = fc_df["Band Hit"].apply(
                    lambda x: "✅" if x is True else ("❌" if x is False else "—")
                )

            # Append totals row
            _totals_row = {c: "" for c in fc_df.columns}
            _totals_row["Store #"] = "TOTAL"
            _totals_row["Store"] = ""
            for col, val in _totals.items():
                _totals_row[col] = f"${val:,.0f}" if val > 0 else "—"
            fc_df = pd.concat([fc_df, pd.DataFrame([_totals_row])], ignore_index=True)

            st.dataframe(fc_df, use_container_width=True, hide_index=True)
        else:
            st.info("No forecast rows for this week.")
    except Exception as _e:
        st.error(f"Could not load forecasts: {_e}")

    # ── Manual back-test ───────────────────────────────────────────────────────
    st.divider()
    st.subheader("Manual Back-Test")
    st.caption(
        "Run this after uploading Sales, SoS, and VOTG for a past week to back-fill "
        "actual results onto the forecast rows and calculate accuracy metrics."
    )

    # Both sales_forecasts and weekly_data_status use Thursday week_start — no offset needed.
    try:
        _bt_thursday = selected_fc_week
    except Exception:
        _bt_thursday = None

    _bt_status = _get_upload_status(_bt_thursday) if _bt_thursday else {}
    b1, b2, b3 = st.columns(3)
    b1.metric("Sales", "✅ Uploaded" if _bt_status.get("sales_uploaded") else "⬜ Pending")
    b2.metric("SoS",   "✅ Uploaded" if _bt_status.get("sos_uploaded")   else "⬜ Pending")
    b3.metric("VOTG",  "✅ Uploaded" if _bt_status.get("votg_uploaded")  else "⬜ Pending")

    _all_ready = all([
        _bt_status.get("sales_uploaded"),
        _bt_status.get("sos_uploaded"),
        _bt_status.get("votg_uploaded"),
    ])

    _already_run = bool(
        fc_resp.data and any(r.get("actual_sales") is not None for r in fc_resp.data)
    ) if "fc_resp" in dir() and fc_resp.data else False

    if _already_run:
        st.success("✅ Back-test already run for this week — actuals are filled in above.")
        if st.button("Re-run Back-Test", key="bt_rerun"):
            with st.spinner("Running back-test..."):
                _n, _err = _run_backtest_for_week(selected_fc_week)
            if _err:
                st.error(_err)
            else:
                st.success(f"Updated {_n} store forecast rows.")
                st.cache_data.clear()
                st.rerun()
    elif not _all_ready:
        _missing_labels = [
            label for key, label in [
                ("sales_uploaded", "Sales"), ("sos_uploaded", "SoS"), ("votg_uploaded", "VOTG")
            ] if not _bt_status.get(key)
        ]
        st.warning(f"Cannot run back-test — missing uploads: {', '.join(_missing_labels)}")
    else:
        if st.button("▶ Run Back-Test for this week", type="primary", use_container_width=True, key="bt_run"):
            with st.spinner("Running back-test — this may take a moment..."):
                _n, _err = _run_backtest_for_week(selected_fc_week)
            if _err:
                st.error(_err)
            else:
                st.success(f"✅ Back-test complete — updated {_n} store forecast rows.")
                st.cache_data.clear()
                st.rerun()

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Sales Scenario Analysis  (Admin)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Sales Scenario Analysis":
    if not user_is_admin:
        st.error("You do not have permission to access this page.")
        st.stop()

    import os
    import altair as alt
    from scenario_engine import ScenarioEngine, SCENARIOS, MODEL_PATH

    _sc_title_col, _sc_refresh_col = st.columns([6, 1])
    with _sc_title_col:
        st.title("📊 Sales Scenario Analysis")
        st.caption(
            "6-month forward-looking projection using the live LightGBM model. "
            "Scenarios update automatically after each monthly model retrain."
        )
    with _sc_refresh_col:
        st.write("")
        st.write("")
        if st.button("🔄 Refresh", key="sc_refresh", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

    # ── Load data (cached; invalidates when model file is updated) ────────────
    @st.cache_data(ttl=1800, show_spinner="Running 6-month projections…")
    def _sc_load(_sb, _mtime):
        """
        _sb    : supabase client (underscore → not hashed by Streamlit)
        _mtime : model file modification time — acts as the cache-bust key
        """
        engine         = ScenarioEngine(_sb)
        results_df     = engine.run_all_scenarios()
        meta           = engine.get_model_meta()
        port_assump    = {sc: engine.get_assumption_summary(sc) for sc in SCENARIOS}
        store_assump   = {
            loc_id: {sc: engine.get_assumption_summary(sc, loc_id) for sc in SCENARIOS}
            for loc_id in engine.stores
        }
        store_names    = {
            loc_id: s['store_name'] for loc_id, s in engine.stores.items()
        }
        return results_df, meta, port_assump, store_assump, store_names

    from supabase_db import get_supabase as _sc_get_sb
    _sb_sc       = _sc_get_sb()
    _model_mtime = os.path.getmtime(str(MODEL_PATH))

    try:
        _sc_df, _sc_meta, _port_assump, _store_assump, _store_names = \
            _sc_load(_sb_sc, _model_mtime)
    except Exception as _sc_err:
        st.error(f"Failed to run scenario engine: {_sc_err}")
        st.stop()

    if _sc_df.empty:
        st.warning("No forecast data returned — ensure stores have sales history uploaded.")
        st.stop()

    # ── Model info badge ──────────────────────────────────────────────────────
    _mv   = _sc_meta.get("model_version", "v1")
    _mape = _sc_meta.get("overall_mape", "—")
    _bacc = _sc_meta.get("band_accuracy", "—")
    _tr   = _sc_meta.get("train_rows", 0)
    st.caption(
        f"🤖 Model **{_mv}** &nbsp;|&nbsp; "
        f"MAPE **{_mape}%** &nbsp;|&nbsp; "
        f"Band accuracy **{_bacc}%** &nbsp;|&nbsp; "
        f"Trained on **{_tr:,}** rows &nbsp;|&nbsp; "
        f"Horizon **26 weeks**"
    )
    st.divider()

    # ── Controls ──────────────────────────────────────────────────────────────
    _ctrl1, _ctrl2, _ctrl3 = st.columns([1, 1, 2])

    with _ctrl1:
        _view_mode = st.radio("View", ["Portfolio", "Per Store"], horizontal=True)

    with _ctrl2:
        _show_all = st.checkbox("Overlay all 3 scenarios", value=True)
        if not _show_all:
            _sc_choice = st.selectbox(
                "Scenario",
                ["conservative", "base", "optimistic"],
                index=1,
                format_func=lambda s: s.capitalize(),
            )
        else:
            _sc_choice = None

    with _ctrl3:
        if _view_mode == "Per Store":
            _sorted_stores = sorted(_store_names.items(), key=lambda x: x[1])
            _store_options  = [loc for loc, _ in _sorted_stores]
            _store_labels   = {loc: name for loc, name in _sorted_stores}
            _sel_store = st.selectbox(
                "Store",
                _store_options,
                format_func=lambda x: _store_labels.get(x, x),
            )
        else:
            _sel_store = None

    st.divider()

    # ── Prepare chart data ────────────────────────────────────────────────────
    _plot_df = _sc_df.copy()
    _plot_df['week_start'] = pd.to_datetime(_plot_df['week_start'])

    # Filter to selected scenario(s)
    if not _show_all and _sc_choice:
        _plot_df = _plot_df[_plot_df['scenario'] == _sc_choice]

    # Portfolio: sum forecast_point; use sqrt-of-sum-of-variances for the
    # confidence band so it reflects portfolio diversification rather than
    # naively stacking each store's individual interval (which inflates the
    # band ~5x vs the statistically correct portfolio-level uncertainty).
    if _view_mode == "Portfolio":
        _plot_df['_half_width'] = (_plot_df['forecast_high'] - _plot_df['forecast_low']) / 2
        _port_agg = (
            _plot_df
            .groupby(['scenario', 'week_num', 'week_start'], as_index=False)
            .agg(
                forecast_point=('forecast_point', 'sum'),
                _var_sum=('_half_width', lambda x: (x ** 2).sum()),
            )
        )
        _port_agg['_port_hw']    = _port_agg['_var_sum'].apply(lambda v: v ** 0.5)
        _port_agg['forecast_low']  = (_port_agg['forecast_point'] - _port_agg['_port_hw']).clip(lower=0)
        _port_agg['forecast_high'] = _port_agg['forecast_point'] + _port_agg['_port_hw']
        _chart_df = _port_agg.drop(columns=['_var_sum', '_port_hw'])
        _chart_title = "Portfolio — Weekly Sales Projection"
    else:
        _chart_df   = _plot_df[_plot_df['location_id'] == _sel_store].copy()
        _chart_title = f"{_store_labels.get(_sel_store, _sel_store)} — Weekly Sales Projection"

    # ── Altair chart ──────────────────────────────────────────────────────────
    _sc_colors = {
        'conservative': '#4A90D9',
        'base':         '#27AE60',
        'optimistic':   '#E67E22',
    }
    _color_scale = alt.Scale(
        domain=list(_sc_colors.keys()),
        range=list(_sc_colors.values()),
    )
    _color_enc = alt.Color(
        'scenario:N',
        scale=_color_scale,
        legend=alt.Legend(
            title='Scenario',
            orient='top-right',
            labelExpr="datum.value == 'conservative' ? 'Conservative' : "
                       "datum.value == 'base' ? 'Base' : 'Optimistic'",
        ),
    )

    _band = (
        alt.Chart(_chart_df)
        .mark_area(opacity=0.12)
        .encode(
            x=alt.X('week_start:T', title='Week Starting',
                     axis=alt.Axis(format='%b %d', labelAngle=-45)),
            y=alt.Y('forecast_low:Q',  title='Weekly Sales',
                     axis=alt.Axis(format='$,.0f')),
            y2=alt.Y2('forecast_high:Q'),
            color=_color_enc,
        )
    )

    _line = (
        alt.Chart(_chart_df)
        .mark_line(strokeWidth=2.5)
        .encode(
            x=alt.X('week_start:T'),
            y=alt.Y('forecast_point:Q',
                     axis=alt.Axis(format='$,.0f')),
            color=_color_enc,
            tooltip=[
                alt.Tooltip('week_start:T',     title='Week',     format='%b %d, %Y'),
                alt.Tooltip('scenario:N',        title='Scenario'),
                alt.Tooltip('forecast_point:Q',  title='Projection', format='$,.0f'),
                alt.Tooltip('forecast_low:Q',    title='Low',        format='$,.0f'),
                alt.Tooltip('forecast_high:Q',   title='High',       format='$,.0f'),
            ],
        )
    )

    _chart = (
        alt.layer(_band, _line)
        .properties(
            height=420,
            title=alt.TitleParams(
                _chart_title,
                fontSize=15,
                anchor='start',
            ),
        )
        .resolve_scale(y='shared')
    )

    st.altair_chart(_chart, use_container_width=True)

    st.caption(
        "Shaded bands show the widening confidence interval "
        "(±13.5% at week 1 → ±34% at week 26). "
        "Optimistic band carries an additional 15% buffer."
    )

    # ── Key Assumptions expander ──────────────────────────────────────────────
    with st.expander("📋 Key Assumptions", expanded=False):
        _assump_src = (
            _store_assump.get(_sel_store, {})
            if _view_mode == "Per Store" and _sel_store
            else _port_assump
        )

        _assump_rows = {
            "SOS Today (good-shift %)": lambda a: f"{a['sos_week1']*100:.1f}%",
            "SOS at Week 26":           lambda a: f"{a['sos_week26']*100:.1f}%",
            "VOTG Today (guests/neg)":  lambda a: f"{a['votg_week1']:.0f}",
            "VOTG at Week 26":          lambda a: f"{a['votg_week26']:.0f}",
            "Gas Price Today":          lambda a: f"${a['gas_week1']:.3f}",
            "Gas Price at Week 26":     lambda a: f"${a['gas_week26']:.3f}",
            "Weather Source":           lambda a: a['weather_src'],
            "Confidence Band":          lambda a: f"{a['conf_week1']} → {a['conf_week26']}",
        }

        _assump_display = {"Assumption": list(_assump_rows.keys())}
        for sc in SCENARIOS:
            a = _assump_src.get(sc, {})
            _assump_display[sc.capitalize()] = [
                fn(a) if a else "—" for fn in _assump_rows.values()
            ]

        st.dataframe(
            pd.DataFrame(_assump_display).set_index("Assumption"),
            use_container_width=True,
        )

    # ── Weekly detail table ───────────────────────────────────────────────────
    st.subheader("Weekly Detail")

    _tbl_df = _chart_df.copy()
    _tbl_df['week_start'] = _tbl_df['week_start'].dt.strftime('%Y-%m-%d')
    _tbl_df['forecast_point'] = _tbl_df['forecast_point'].apply(lambda v: f"${v:,.0f}")
    _tbl_df['forecast_low']   = _tbl_df['forecast_low'].apply(  lambda v: f"${v:,.0f}")
    _tbl_df['forecast_high']  = _tbl_df['forecast_high'].apply( lambda v: f"${v:,.0f}")
    _tbl_df['scenario']       = _tbl_df['scenario'].str.capitalize()

    _tbl_df = _tbl_df.rename(columns={
        'scenario':       'Scenario',
        'week_num':       'Week #',
        'week_start':     'Week Starting',
        'forecast_point': 'Projection',
        'forecast_low':   'Low',
        'forecast_high':  'High',
    })

    _display_cols = ['Scenario', 'Week #', 'Week Starting', 'Projection', 'Low', 'High']
    if _view_mode == "Per Store" and 'recommended_band' in _chart_df.columns:
        _tbl_df['Band'] = _chart_df['recommended_band'].values
        _display_cols.append('Band')

    st.dataframe(
        _tbl_df[_display_cols].reset_index(drop=True),
        use_container_width=True,
        hide_index=True,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: SoS/VOTG Trends
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "SoS/VOTG Trends":
    st.title("SoS / VOTG Trends")
    st.caption("Week-over-week and month-over-month rank tracking for Speed of Service and Voice of the Guest.")

    ref_data = _cached_reference_data()
    sos_df = _cached_sos_data()
    votg_df = _cached_votg_data()

    if sos_df.empty and votg_df.empty:
        st.info("No SoS or VOTG data uploaded yet. Upload data via Admin > Upload SoS / Upload VOTG.")
        st.stop()

    # Build store-name lookup from reference data
    store_lookup = dict(zip(ref_data["location_id"], ref_data["store_name"]))
    dm_lookup = dict(zip(ref_data["location_id"], ref_data["dm"]))

    # --- Filters ---
    dm_list = sorted(ref_data["dm"].dropna().unique().tolist())
    col_f1, col_f2 = st.columns(2)
    with col_f1:
        dm_filter = st.multiselect("Filter by DM", dm_list, key="sos_dm_filter")
    with col_f2:
        if dm_filter:
            store_ids = ref_data[ref_data["dm"].isin(dm_filter)]["location_id"].tolist()
        else:
            store_ids = ref_data["location_id"].tolist()
        store_options = [f"{sid} — {store_lookup.get(sid, sid)}" for sid in sorted(store_ids)]
        store_filter = st.multiselect("Filter by Store", store_options, key="sos_store_filter")
        if store_filter:
            selected_ids = [s.split(" — ")[0] for s in store_filter]
        else:
            selected_ids = store_ids

    period = st.radio("Period", ["All Weeks", "Last Month", "Last Quarter", "Last Year"],
                      horizontal=True, key="sos_period")
    cutoff = None
    if period == "Last Month":
        cutoff = str(date.today() - timedelta(days=30))
    elif period == "Last Quarter":
        cutoff = str(date.today() - timedelta(days=90))
    elif period == "Last Year":
        cutoff = str(date.today() - timedelta(days=365))

    # --- SoS Trends ---
    st.subheader("Speed of Service (SoS) — Rank Trends")
    if not sos_df.empty:
        sos = sos_df[sos_df["location_id"].isin(selected_ids)].copy()
        if cutoff:
            sos = sos[sos["week_start"] >= cutoff]
        if sos.empty:
            st.info("No SoS data for the selected filters.")
        else:
            sos["store"] = sos["location_id"].map(store_lookup)
            sos["dm"] = sos["location_id"].map(dm_lookup)
            sos["week_start"] = pd.to_datetime(sos["week_start"])
            sos = sos.sort_values("week_start")

            # Convert numeric columns from strings
            for col in ["good_shift_rank", "good_shift", "total_stores", "red_ticket", "shift_streak"]:
                if col in sos.columns:
                    sos[col] = pd.to_numeric(sos[col], errors="coerce")

            import altair as alt

            # Portfolio average rank by week
            portfolio_avg = sos.groupby("week_start").agg(
                avg_rank=("good_shift_rank", "mean"),
                stores=("location_id", "nunique"),
            ).reset_index()
            portfolio_avg["avg_rank"] = portfolio_avg["avg_rank"].round(1)

            avg_chart = alt.Chart(portfolio_avg).mark_line(
                point=True, strokeWidth=3, color="#2B3A4E"
            ).encode(
                x=alt.X("week_start:T", title="Week", axis=alt.Axis(format="%m/%d")),
                y=alt.Y("avg_rank:Q", title="Avg Rank",
                        scale=alt.Scale(zero=False)),
                tooltip=[alt.Tooltip("week_start:T", title="Week"),
                         alt.Tooltip("avg_rank:Q", title="Avg Rank", format=".1f"),
                         alt.Tooltip("stores:Q", title="Stores")]
            ).properties(height=350).interactive()
            st.altair_chart(avg_chart, use_container_width=True)
            st.caption("Portfolio average SoS rank. Line going up = rank number increasing (getting worse). Line going down = improving.")

            # Identify stores on a negative trend (rank getting worse)
            st.markdown("**⚠️ Stores Trending Negative** (rank getting worse over recent weeks)")

            weeks = sorted(sos["week_start"].unique())
            if len(weeks) >= 2:
                # Compare each store's most recent rank to their earliest rank
                store_trends = []
                for store_name in sos["store"].unique():
                    store_data = sos[sos["store"] == store_name].sort_values("week_start")
                    if len(store_data) >= 2:
                        first_rank = store_data["good_shift_rank"].iloc[0]
                        last_rank = store_data["good_shift_rank"].iloc[-1]
                        dm = store_data["dm"].iloc[0]
                        if pd.notna(first_rank) and pd.notna(last_rank):
                            change = last_rank - first_rank
                            store_trends.append({
                                "Store": store_name,
                                "DM": dm,
                                "First Rank": int(first_rank),
                                "Latest Rank": int(last_rank),
                                "Change": int(change),
                                "Direction": "📉 Declining" if change > 0 else ("📈 Improving" if change < 0 else "➡️ Flat"),
                            })

                if store_trends:
                    trend_df = pd.DataFrame(store_trends).sort_values("Change", ascending=False)
                    declining = trend_df[trend_df["Change"] > 0]
                    if not declining.empty:
                        st.dataframe(declining, use_container_width=True, hide_index=True)
                    else:
                        st.success("No stores are trending negative — all ranks are holding or improving.")

                    with st.expander("View all stores"):
                        st.dataframe(trend_df, use_container_width=True, hide_index=True)
            else:
                st.info("Need at least 2 weeks of data to calculate trends.")
    else:
        st.info("No SoS data uploaded yet.")

    st.divider()

    # --- VOTG Trends ---
    st.subheader("Voice of the Guest (VOTG) — Rank Trends")
    if not votg_df.empty:
        votg = votg_df[votg_df["location_id"].isin(selected_ids)].copy()
        if cutoff:
            votg = votg[votg["week_start"] >= cutoff]
        if votg.empty:
            st.info("No VOTG data for the selected filters.")
        else:
            votg["store"] = votg["location_id"].map(store_lookup)
            votg["dm"] = votg["location_id"].map(dm_lookup)
            votg["week_start"] = pd.to_datetime(votg["week_start"])
            votg = votg.sort_values("week_start")

            # Convert numeric columns from strings
            for col in ["votg_rank", "total_stores", "total_negative_reviews", "total_reviews", "guests_per_negative"]:
                if col in votg.columns:
                    votg[col] = pd.to_numeric(votg[col], errors="coerce")

            import altair as alt

            # Portfolio average rank by week
            votg_portfolio = votg.groupby("week_start").agg(
                avg_rank=("votg_rank", "mean"),
                avg_neg=("total_negative_reviews", "mean"),
                stores=("location_id", "nunique"),
            ).reset_index()
            votg_portfolio["avg_rank"] = votg_portfolio["avg_rank"].round(1)
            votg_portfolio["avg_neg"] = votg_portfolio["avg_neg"].round(1)

            votg_avg_chart = alt.Chart(votg_portfolio).mark_line(
                point=True, strokeWidth=3, color="#2B3A4E"
            ).encode(
                x=alt.X("week_start:T", title="Week", axis=alt.Axis(format="%m/%d")),
                y=alt.Y("avg_rank:Q", title="Avg Rank",
                        scale=alt.Scale(zero=False)),
                tooltip=[alt.Tooltip("week_start:T", title="Week"),
                         alt.Tooltip("avg_rank:Q", title="Avg Rank", format=".1f"),
                         alt.Tooltip("avg_neg:Q", title="Avg Neg Reviews", format=".1f"),
                         alt.Tooltip("stores:Q", title="Stores")]
            ).properties(height=350).interactive()
            st.altair_chart(votg_avg_chart, use_container_width=True)
            st.caption("Portfolio average VOTG rank. Line going up = rank number increasing (getting worse). Line going down = improving.")

            # Identify stores on a negative trend
            st.markdown("**⚠️ Stores Trending Negative** (rank getting worse over recent weeks)")

            votg_weeks = sorted(votg["week_start"].unique())
            if len(votg_weeks) >= 2:
                votg_trends = []
                for store_name in votg["store"].unique():
                    store_data = votg[votg["store"] == store_name].sort_values("week_start")
                    if len(store_data) >= 2:
                        first_rank = store_data["votg_rank"].iloc[0]
                        last_rank = store_data["votg_rank"].iloc[-1]
                        dm = store_data["dm"].iloc[0]
                        if pd.notna(first_rank) and pd.notna(last_rank):
                            change = last_rank - first_rank
                            votg_trends.append({
                                "Store": store_name,
                                "DM": dm,
                                "First Rank": int(first_rank),
                                "Latest Rank": int(last_rank),
                                "Change": int(change),
                                "Direction": "📉 Declining" if change > 0 else ("📈 Improving" if change < 0 else "➡️ Flat"),
                            })

                if votg_trends:
                    votg_trend_df = pd.DataFrame(votg_trends).sort_values("Change", ascending=False)
                    votg_declining = votg_trend_df[votg_trend_df["Change"] > 0]
                    if not votg_declining.empty:
                        st.dataframe(votg_declining, use_container_width=True, hide_index=True)
                    else:
                        st.success("No stores are trending negative — all VOTG ranks are holding or improving.")

                    with st.expander("View all stores"):
                        st.dataframe(votg_trend_df, use_container_width=True, hide_index=True)
            else:
                st.info("Need at least 2 weeks of data to calculate trends.")
    else:
        st.info("No VOTG data uploaded yet.")


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Tattle Insights
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Tattle Insights":
    st.title("Tattle Insights")
    st.caption("Time-of-day patterns, category trends, and day-part analysis from guest feedback.")

    ref_data = _cached_reference_data()

    # Date range selector — default to last 6 months for performance
    date_range = st.radio("Date Range", ["Last 3 Months", "Last 6 Months", "Last Year", "All Time"],
                          index=1, horizontal=True, key="tattle_date_range")
    if date_range == "Last 3 Months":
        date_cutoff = str(date.today() - timedelta(days=90))
    elif date_range == "Last 6 Months":
        date_cutoff = str(date.today() - timedelta(days=180))
    elif date_range == "Last Year":
        date_cutoff = str(date.today() - timedelta(days=365))
    else:
        date_cutoff = None

    # Load reviews with date filter applied at the query level
    @st.cache_data(ttl=1800)
    def _load_tattle_filtered(cutoff):
        from supabase_db import get_supabase
        sb = get_supabase()
        all_rows = []
        page_size = 1000
        offset = 0
        fields = (
            "id,location_id,location_label,score,cer,"
            "experienced_time,completed_time,day_part_label,channel_label,"
            "comment"
        )
        while True:
            q = sb.table("tattle_reviews").select(fields).order("experienced_time", desc=True)
            if cutoff:
                q = q.gte("experienced_time", cutoff)
            q = q.range(offset, offset + page_size - 1)
            resp = q.execute()
            if not resp.data:
                break
            all_rows.extend(resp.data)
            if len(resp.data) < page_size:
                break
            offset += page_size
        return pd.DataFrame(all_rows) if all_rows else pd.DataFrame()

    reviews_df = _load_tattle_filtered(date_cutoff)

    if reviews_df.empty:
        st.info("No Tattle review data available. Run the Tattle ingestion workflow first.")
        st.stop()

    # Build lookups
    store_lookup = dict(zip(ref_data["location_id"], ref_data["store_name"]))
    dm_lookup = dict(zip(ref_data["location_id"], ref_data["dm"]))

    reviews_df["store"] = reviews_df["location_id"].map(store_lookup)
    reviews_df["dm"] = reviews_df["location_id"].map(dm_lookup)

    # Parse timestamps
    reviews_df["experienced_time"] = pd.to_datetime(reviews_df["experienced_time"], errors="coerce")
    reviews_df["hour"] = reviews_df["experienced_time"].dt.hour
    reviews_df["day_of_week"] = reviews_df["experienced_time"].dt.day_name()
    reviews_df["month"] = reviews_df["experienced_time"].dt.to_period("M").astype(str)
    reviews_df["week"] = reviews_df["experienced_time"].dt.isocalendar().week.astype(int)

    # Parse snapshots (loaded separately — only reviews with snapshots)
    def parse_snapshots(snap_str):
        if pd.isna(snap_str) or not snap_str:
            return {}
        try:
            snaps = json.loads(snap_str) if isinstance(snap_str, str) else snap_str
            result = {}
            for s in snaps:
                cat = s.get("category") or s.get("name") or s.get("label") or "Unknown"
                rating = s.get("rating")
                if rating is not None:
                    result[cat] = rating
            return result
        except Exception:
            return {}

    # --- Filters ---
    dm_list = sorted(reviews_df["dm"].dropna().unique().tolist())
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        dm_filter = st.multiselect("Filter by DM", dm_list, key="tattle_dm")
    with col_f2:
        if dm_filter:
            store_opts = sorted(reviews_df[reviews_df["dm"].isin(dm_filter)]["store"].dropna().unique().tolist())
        else:
            store_opts = sorted(reviews_df["store"].dropna().unique().tolist())
        store_filter = st.multiselect("Filter by Store", store_opts, key="tattle_store")
    with col_f3:
        channel_opts = sorted(reviews_df["channel_label"].dropna().unique().tolist())
        channel_filter = st.multiselect("Filter by Channel", channel_opts, key="tattle_channel")

    filtered = reviews_df.copy()
    if dm_filter:
        filtered = filtered[filtered["dm"].isin(dm_filter)]
    if store_filter:
        filtered = filtered[filtered["store"].isin(store_filter)]
    if channel_filter:
        filtered = filtered[filtered["channel_label"].isin(channel_filter)]

    if filtered.empty:
        st.warning("No reviews match the selected filters.")
        st.stop()

    # --- Metrics ---
    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("Total Reviews", f"{len(filtered):,}")
    avg_score = filtered["score"].mean()
    mc2.metric("Avg Score", f"{avg_score:.1f}" if pd.notna(avg_score) else "—")
    with_comments = filtered["comment"].notna().sum()
    mc3.metric("With Comments", f"{with_comments:,}")
    unique_stores = filtered["location_id"].nunique()
    mc4.metric("Stores", unique_stores)

    st.divider()

    # --- Time-of-Day Heatmap ---
    st.subheader("🕐 Time-of-Day Heatmap — When Do Issues Happen?")
    st.caption("Based on the guest's experienced time. Darker = more low-scoring reviews.")

    # Focus on below-average reviews for the heatmap
    low_reviews = filtered[pd.to_numeric(filtered["score"], errors="coerce") < 70].copy()
    if not low_reviews.empty and low_reviews["hour"].notna().any():
        heatmap_data = low_reviews.groupby(
            ["day_of_week", "hour"]
        ).size().reset_index(name="count")

        day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        heatmap_data["day_of_week"] = pd.Categorical(
            heatmap_data["day_of_week"], categories=day_order, ordered=True
        )

        import altair as alt
        # Convert to 12-hour format for display
        heatmap_data["hour_label"] = heatmap_data["hour"].apply(
            lambda h: f"{h % 12 or 12}:00"
        )
        # Sort order based on original hour
        hour_order = [f"{h % 12 or 12}:00" for h in sorted(heatmap_data["hour"].unique())]

        heatmap = alt.Chart(heatmap_data).mark_rect().encode(
            x=alt.X("hour_label:N", title="Hour of Day",
                     sort=hour_order),
            y=alt.Y("day_of_week:N", title="Day",
                     sort=day_order),
            color=alt.Color("count:Q", title="Low-Score Reviews",
                           scale=alt.Scale(scheme="reds")),
            tooltip=["day_of_week", "hour_label", "count"]
        ).properties(height=280).configure_view(strokeWidth=0)
        st.altair_chart(heatmap, use_container_width=True)

        # Top problem time slots
        top_slots = heatmap_data.nlargest(5, "count")
        if not top_slots.empty:
            st.markdown("**Top 5 Problem Time Slots** (most low-score reviews):")
            for _, slot in top_slots.iterrows():
                h = int(slot["hour"])
                ampm = f"{h % 12 or 12}{'am' if h < 12 else 'pm'}"
                st.markdown(f"- **{slot['day_of_week']} at {ampm}** — {int(slot['count'])} low-score reviews")
    else:
        st.info("Not enough low-score reviews to build a heatmap.")

    st.divider()

    # --- Category Trends (Month-over-Month) ---
    st.subheader("📊 Category Trends — Month over Month")
    st.caption("Average rating by category across months. Track which areas are improving or declining.")

    # Load snapshots separately (heavier query, only for category analysis)
    snap_df = _cached_tattle_with_snapshots()
    if not snap_df.empty:
        snap_df["experienced_time"] = pd.to_datetime(snap_df["experienced_time"], errors="coerce")
        snap_df["month"] = snap_df["experienced_time"].dt.to_period("M").astype(str)
        snap_df["store"] = snap_df["location_id"].map(store_lookup)
        snap_df["categories"] = snap_df["snapshots"].apply(parse_snapshots)
        # Apply same filters
        if dm_filter:
            snap_df["dm"] = snap_df["location_id"].map(dm_lookup)
            snap_df = snap_df[snap_df["dm"].isin(dm_filter)]
        if store_filter:
            snap_df = snap_df[snap_df["store"].isin(store_filter)]
        filtered_snap = snap_df
    else:
        filtered_snap = pd.DataFrame()

    # Explode categories into rows
    cat_rows = []
    for _, row in (filtered_snap.iterrows() if not filtered_snap.empty else []):
        cats = row["categories"]
        if cats:
            for cat_name, rating in cats.items():
                cat_rows.append({
                    "month": row["month"],
                    "category": cat_name,
                    "rating": float(rating),
                    "store": row["store"],
                })

    if cat_rows:
        cat_df = pd.DataFrame(cat_rows)
        cat_monthly = cat_df.groupby(["month", "category"])["rating"].mean().round(2).reset_index()

        import altair as alt
        cat_chart = alt.Chart(cat_monthly).mark_line(point=True).encode(
            x=alt.X("month:N", title="Month", sort=None),
            y=alt.Y("rating:Q", title="Avg Rating",
                     scale=alt.Scale(domain=[
                         max(0, cat_monthly["rating"].min() - 0.3),
                         min(5, cat_monthly["rating"].max() + 0.3)
                     ])),
            color=alt.Color("category:N", title="Category"),
            tooltip=["category", "month", "rating"]
        ).properties(height=400).interactive()
        st.altair_chart(cat_chart, use_container_width=True)

        # Category comparison table
        cat_pivot = cat_monthly.pivot_table(
            index="category", columns="month", values="rating"
        ).round(2)
        if not cat_pivot.empty:
            # Add trend arrow
            cols = cat_pivot.columns.tolist()
            if len(cols) >= 2:
                cat_pivot["Trend"] = cat_pivot.apply(
                    lambda r: "📈 Improving" if r[cols[-1]] > r[cols[-2]]
                    else ("📉 Declining" if r[cols[-1]] < r[cols[-2]] else "➡️ Flat"),
                    axis=1
                )
            # Style: widen category column, center headers
            styled_cat = cat_pivot.reset_index()
            styled_cat = styled_cat.rename(columns={"category": "Category"})
            st.dataframe(
                styled_cat,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Category": st.column_config.TextColumn("Category", width="large"),
                },
            )
    else:
        st.info("No category-level data available in the reviews.")

    st.divider()

    # --- Negative Review Count by Category (Month-over-Month) ---
    st.subheader("📉 Negative Reviews by Category — Month over Month")
    st.caption("Count of reviews scoring below 70 per category. Fewer = improving. A downward trend is good.")

    neg_cat_rows = []
    for _, row in (filtered_snap.iterrows() if not filtered_snap.empty else []):
        score_val = pd.to_numeric(row.get("score"), errors="coerce")
        if pd.notna(score_val) and score_val < 70:
            cats = row["categories"]
            if cats:
                for cat_name in cats.keys():
                    neg_cat_rows.append({
                        "month": row["month"],
                        "category": cat_name,
                    })

    if neg_cat_rows:
        neg_cat_df = pd.DataFrame(neg_cat_rows)
        neg_cat_monthly = neg_cat_df.groupby(["month", "category"]).size().reset_index(name="neg_count")

        import altair as alt
        neg_cat_chart = alt.Chart(neg_cat_monthly).mark_line(point=True).encode(
            x=alt.X("month:N", title="Month", sort=None),
            y=alt.Y("neg_count:Q", title="Negative Reviews"),
            color=alt.Color("category:N", title="Category"),
            tooltip=["category", "month", "neg_count"]
        ).properties(height=400).interactive()
        st.altair_chart(neg_cat_chart, use_container_width=True)

        # Table with trend
        neg_cat_pivot = neg_cat_monthly.pivot_table(
            index="category", columns="month", values="neg_count", fill_value=0
        ).astype(int)
        if not neg_cat_pivot.empty:
            cols = neg_cat_pivot.columns.tolist()
            if len(cols) >= 2:
                neg_cat_pivot["Trend"] = neg_cat_pivot.apply(
                    lambda r: "📈 Improving" if r[cols[-1]] < r[cols[-2]]
                    else ("📉 Getting Worse" if r[cols[-1]] > r[cols[-2]] else "➡️ Flat"),
                    axis=1
                )
            styled_neg = neg_cat_pivot.reset_index().rename(columns={"category": "Category"})
            st.dataframe(
                styled_neg,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Category": st.column_config.TextColumn("Category", width="large"),
                },
            )
    else:
        st.info("No negative reviews with category data for the selected filters.")

    st.divider()

    # --- Day Part Analysis ---
    st.subheader("🍽️ Day Part Analysis")
    st.caption("Compare performance across meal periods and service channels.")

    if filtered["day_part_label"].notna().any():
        daypart_stats = filtered.groupby("day_part_label").agg(
            count=("score", "size"),
            avg_score=("score", "mean"),
            low_pct=("score", lambda x: (pd.to_numeric(x, errors="coerce") < 70).mean()),
        ).round(3).reset_index()
        daypart_stats.columns = ["Day Part", "Reviews", "Avg Score", "% Low Score"]
        daypart_stats["% Low Score"] = (daypart_stats["% Low Score"] * 100).round(1)
        daypart_stats = daypart_stats.sort_values("Avg Score")

        import altair as alt

        # Chart 1: Avg Score by Day Part (color by relative performance)
        median_score = daypart_stats["Avg Score"].median()
        daypart_stats["_color"] = daypart_stats["Avg Score"].apply(
            lambda x: "Above Median" if x >= median_score else "Below Median"
        )
        dp_chart = alt.Chart(daypart_stats).mark_bar().encode(
            x=alt.X("Avg Score:Q", scale=alt.Scale(zero=False)),
            y=alt.Y("Day Part:N", sort="-x"),
            color=alt.Color("_color:N", scale=alt.Scale(
                domain=["Above Median", "Below Median"],
                range=["#C6EFCE", "#FFC7CE"]
            ), title=""),
            tooltip=["Day Part", "Reviews", "Avg Score", "% Low Score"]
        ).properties(height=250)
        st.altair_chart(dp_chart, use_container_width=True)

        # Chart 2: % Negative by Day Part (the actionable view)
        st.markdown("**% Negative Reviews by Day Part** — which periods have the most complaints?")
        neg_sorted = daypart_stats.sort_values("% Low Score", ascending=False)
        neg_chart = alt.Chart(neg_sorted).mark_bar().encode(
            x=alt.X("% Low Score:Q", title="% Negative (score < 70)"),
            y=alt.Y("Day Part:N", sort="-x"),
            color=alt.Color("% Low Score:Q", scale=alt.Scale(
                scheme="reds"
            ), legend=None),
            tooltip=["Day Part", "Reviews", "% Low Score", "Avg Score"]
        ).properties(height=250)
        st.altair_chart(neg_chart, use_container_width=True)

        st.dataframe(daypart_stats, use_container_width=True, hide_index=True)
    else:
        st.info("No day-part data available in the reviews.")


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Sentiment Dashboard
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Sentiment Dashboard":
    st.title("Sentiment Dashboard")
    st.caption("AI-powered analysis of guest review themes — powered by Claude sentiment scoring.")

    ref_data = _cached_reference_data()
    scored = _cached_tattle_scored()
    # Get total comment count from light query
    light_df = _cached_tattle_reviews_light()
    total_reviews = len(light_df[light_df["comment"].notna()]) if not light_df.empty else 0

    if scored.empty:
        st.info("No Tattle review data available.")
        st.stop()

    # Build lookups
    store_lookup = dict(zip(ref_data["location_id"], ref_data["store_name"]))
    dm_lookup = dict(zip(ref_data["location_id"], ref_data["dm"]))
    scored["store"] = scored["location_id"].map(store_lookup)
    scored["dm"] = scored["location_id"].map(dm_lookup)
    scored["experienced_time"] = pd.to_datetime(scored["experienced_time"], errors="coerce")
    scored["month"] = scored["experienced_time"].dt.to_period("M").astype(str)

    scored_count = len(scored)

    st.metric("Scored Reviews", f"{scored_count:,} / {total_reviews:,} with comments")
    if scored_count == 0:
        st.warning("No reviews have been sentiment-scored yet. Run the Tattle Sentiment Scoring workflow.")
        st.stop()

    # --- Filters ---
    dm_list = sorted(scored["dm"].dropna().unique().tolist())
    col_f1, col_f2 = st.columns(2)
    with col_f1:
        dm_filter = st.multiselect("Filter by DM", dm_list, key="sent_dm")
    with col_f2:
        if dm_filter:
            store_opts = sorted(scored[scored["dm"].isin(dm_filter)]["store"].dropna().unique().tolist())
        else:
            store_opts = sorted(scored["store"].dropna().unique().tolist())
        store_filter = st.multiselect("Filter by Store", store_opts, key="sent_store")

    if dm_filter:
        scored = scored[scored["dm"].isin(dm_filter)]
    if store_filter:
        scored = scored[scored["store"].isin(store_filter)]

    # Parse sentiment themes
    def parse_themes(themes_val):
        if pd.isna(themes_val) or not themes_val:
            return {}
        try:
            if isinstance(themes_val, str):
                return json.loads(themes_val)
            return themes_val
        except Exception:
            return {}

    scored["parsed_themes"] = scored["sentiment_themes"].apply(parse_themes)

    st.divider()

    # --- Theme Frequency ---
    st.subheader("🔍 Most Common Themes — Negative vs Positive")
    st.caption("Side-by-side comparison of what guests complain about vs. what they praise.")

    # Split scored reviews into negative (score < 70) and positive (score >= 70)
    scored["_score_num"] = pd.to_numeric(scored["score"], errors="coerce")
    neg_scored = scored[scored["_score_num"] < 70]
    pos_scored = scored[scored["_score_num"] >= 70]

    def count_themes(df):
        counts = {}
        for themes in df["parsed_themes"]:
            for theme in themes.get("themes", []):
                counts[theme] = counts.get(theme, 0) + 1
        return counts

    neg_counts = count_themes(neg_scored)
    pos_counts = count_themes(pos_scored)

    import altair as alt
    col_neg, col_pos = st.columns(2)

    with col_neg:
        st.markdown(f"**Negative Reviews** ({len(neg_scored):,} reviews)")
        if neg_counts:
            neg_df = pd.DataFrame([
                {"Theme": k, "Count": v} for k, v in neg_counts.items()
            ]).sort_values("Count", ascending=False)
            neg_chart = alt.Chart(neg_df.head(10)).mark_bar().encode(
                x=alt.X("Count:Q", title="Mentions"),
                y=alt.Y("Theme:N", sort="-x", title=""),
                color=alt.value("#FFC7CE"),
                tooltip=["Theme", "Count"]
            ).properties(height=300)
            st.altair_chart(neg_chart, use_container_width=True)
        else:
            st.info("No negative review themes.")

    with col_pos:
        st.markdown(f"**Positive Reviews** ({len(pos_scored):,} reviews)")
        if pos_counts:
            pos_df = pd.DataFrame([
                {"Theme": k, "Count": v} for k, v in pos_counts.items()
            ]).sort_values("Count", ascending=False)
            pos_chart = alt.Chart(pos_df.head(10)).mark_bar().encode(
                x=alt.X("Count:Q", title="Mentions"),
                y=alt.Y("Theme:N", sort="-x", title=""),
                color=alt.value("#C6EFCE"),
                tooltip=["Theme", "Count"]
            ).properties(height=300)
            st.altair_chart(pos_chart, use_container_width=True)
        else:
            st.info("No positive review themes.")

    st.divider()

    # --- Theme Trends over Time (Negative Only) ---
    st.subheader("📉 Negative Review Theme Trends — Month over Month")
    st.caption("Tracking negative review themes over time. A downward trend = improving.")

    theme_time_rows = []
    for _, row in neg_scored.iterrows():
        themes = row["parsed_themes"]
        for theme in themes.get("themes", []):
            theme_time_rows.append({
                "month": row["month"],
                "theme": theme,
            })

    if theme_time_rows:
        tt_df = pd.DataFrame(theme_time_rows)
        all_themes = tt_df["theme"].value_counts().index.tolist()

        # Theme filter
        theme_filter = st.multiselect(
            "Filter by Theme", all_themes, default=all_themes[:7],
            key="neg_theme_filter",
            help="Select which themes to display. Default shows top 7."
        )

        if not theme_filter:
            theme_filter = all_themes[:7]

        tt_filtered = tt_df[tt_df["theme"].isin(theme_filter)]
        tt_monthly = tt_filtered.groupby(["month", "theme"]).size().reset_index(name="mentions")

        # Calculate overall average across all selected themes per month
        overall_avg = tt_monthly.groupby("month")["mentions"].mean().reset_index()
        overall_avg["theme"] = "── OVERALL AVG ──"

        import altair as alt

        # Individual theme lines (thinner, with points)
        theme_lines = alt.Chart(tt_monthly).mark_line(point=True, strokeWidth=1.5, opacity=0.6).encode(
            x=alt.X("month:N", title="Month", sort=None),
            y=alt.Y("mentions:Q", title="Mentions"),
            color=alt.Color("theme:N", title="Theme"),
            tooltip=["theme", "month", "mentions"]
        )

        # Overall average line (bold, dark)
        avg_line = alt.Chart(overall_avg).mark_line(
            strokeWidth=4, strokeDash=[0], point=True
        ).encode(
            x=alt.X("month:N", sort=None),
            y=alt.Y("mentions:Q"),
            color=alt.value("#2B3A4E"),
            tooltip=[alt.Tooltip("theme:N", title=""), "month", alt.Tooltip("mentions:Q", title="Avg Mentions", format=".1f")]
        )

        combined = (theme_lines + avg_line).properties(height=400).interactive()
        st.altair_chart(combined, use_container_width=True)
        st.caption("**Bold dark line** = overall average across selected themes. Individual themes shown lighter.")
    else:
        st.info("Not enough data to show theme trends.")

    st.divider()

    # --- Worst Stores by Sentiment ---
    st.subheader("⚠️ Stores with Most Negative Themes")

    store_theme_counts = {}
    for _, row in scored.iterrows():
        themes = row["parsed_themes"]
        sid = row["store"]
        if sid and themes.get("themes"):
            store_theme_counts[sid] = store_theme_counts.get(sid, 0) + len(themes["themes"])

    if store_theme_counts:
        worst_df = pd.DataFrame([
            {"Store": k, "Total Issue Mentions": v} for k, v in store_theme_counts.items()
        ]).sort_values("Total Issue Mentions", ascending=False).head(15)

        import altair as alt
        worst_chart = alt.Chart(worst_df).mark_bar().encode(
            x=alt.X("Total Issue Mentions:Q"),
            y=alt.Y("Store:N", sort="-x"),
            color=alt.condition(
                alt.datum["Total Issue Mentions"] > worst_df["Total Issue Mentions"].median(),
                alt.value("#FFC7CE"),
                alt.value("#C6EFCE")
            ),
            tooltip=["Store", "Total Issue Mentions"]
        ).properties(height=400)
        st.altair_chart(worst_chart, use_container_width=True)

    st.divider()

    # --- Overall Score Distribution ---
    st.subheader("📊 Guest Score Distribution")
    st.caption("How guest scores are spread across the 0–100 scale. Green bars are scores 70+, red bars are below 70 (negative).")

    valid_scores = scored[scored["score"].notna()].copy()
    valid_scores["score_num"] = pd.to_numeric(valid_scores["score"], errors="coerce")
    valid_scores = valid_scores[valid_scores["score_num"].notna()]

    if not valid_scores.empty:
        import altair as alt
        # Pre-bin and assign color category so Altair colors correctly
        valid_scores["score_bin"] = (valid_scores["score_num"] // 5 * 5).astype(int)
        valid_scores["rating"] = valid_scores["score_num"].apply(
            lambda x: "Positive (70+)" if x >= 70 else "Negative (<70)"
        )
        binned = valid_scores.groupby(["score_bin", "rating"]).size().reset_index(name="count")

        hist = alt.Chart(binned).mark_bar().encode(
            x=alt.X("score_bin:O", title="Guest Score (0–100)",
                     axis=alt.Axis(labelAngle=0)),
            y=alt.Y("count:Q", title="Reviews"),
            color=alt.Color("rating:N", scale=alt.Scale(
                domain=["Positive (70+)", "Negative (<70)"],
                range=["#C6EFCE", "#FFC7CE"]
            ), title=""),
            tooltip=["score_bin", "count", "rating"]
        ).properties(height=250)
        st.altair_chart(hist, use_container_width=True)

        avg_score = valid_scores["score_num"].mean()
        neg_pct = (valid_scores["score_num"] < 70).mean() * 100
        sc1, sc2, sc3 = st.columns(3)
        sc1.metric("Average Score", f"{avg_score:.2f}")
        sc2.metric("Negative Reviews", f"{neg_pct:.1f}%")
        sc3.metric("Total Scored", f"{len(valid_scores):,}")

        # --- Monthly trend: Avg Score & Negative % over time ---
        st.markdown("**Historical Trend — Average Score & Negative %**")
        st.caption("Track how guest satisfaction is moving month over month. Score going up + negative % going down = improving.")

        valid_scores["month"] = valid_scores["experienced_time"].dt.to_period("M").astype(str)
        monthly_stats = valid_scores.groupby("month").agg(
            avg_score=("score_num", "mean"),
            neg_pct=("score_num", lambda x: (x < 70).mean() * 100),
            reviews=("score_num", "count"),
        ).reset_index()
        monthly_stats["avg_score"] = monthly_stats["avg_score"].round(1)
        monthly_stats["neg_pct"] = monthly_stats["neg_pct"].round(1)

        if len(monthly_stats) >= 2:
            # Dual-axis chart: avg score (line) + negative % (bars)
            score_line = alt.Chart(monthly_stats).mark_line(
                point=True, strokeWidth=3, color="#2B3A4E"
            ).encode(
                x=alt.X("month:N", title="Month", sort=None),
                y=alt.Y("avg_score:Q",
                         title="Avg Score",
                         axis=alt.Axis(titlePadding=15),
                         scale=alt.Scale(domain=[
                             max(0, monthly_stats["avg_score"].min() - 10),
                             100
                         ])),
                tooltip=[alt.Tooltip("month:N", title="Month"),
                         alt.Tooltip("avg_score:Q", title="Avg Score", format=".1f"),
                         alt.Tooltip("reviews:Q", title="Reviews")]
            )

            neg_y_scale = alt.Scale(domain=[0, max(50, monthly_stats["neg_pct"].max() + 5)])

            neg_bars = alt.Chart(monthly_stats).mark_bar(opacity=0.4, color="#FFC7CE").encode(
                x=alt.X("month:N", sort=None),
                y=alt.Y("neg_pct:Q", title="Negative %", scale=neg_y_scale),
                tooltip=[alt.Tooltip("month:N", title="Month"),
                         alt.Tooltip("neg_pct:Q", title="Negative %", format=".1f")]
            )

            # Trend line for negative % (manual linear regression)
            import numpy as np
            _x_nums = np.arange(len(monthly_stats))
            _y_vals = monthly_stats["neg_pct"].values
            if len(_x_nums) >= 2:
                _slope, _intercept = np.polyfit(_x_nums, _y_vals, 1)
                monthly_stats["neg_trend"] = _intercept + _slope * _x_nums
            else:
                monthly_stats["neg_trend"] = monthly_stats["neg_pct"]

            neg_trend = alt.Chart(monthly_stats).mark_line(
                strokeDash=[6, 3], strokeWidth=2, color="#D32F2F"
            ).encode(
                x=alt.X("month:N", sort=None),
                y=alt.Y("neg_trend:Q", scale=neg_y_scale),
            )

            combined = alt.layer(neg_bars, neg_trend, score_line).resolve_scale(
                y="independent"
            ).properties(height=300)
            st.altair_chart(combined, use_container_width=True)
            st.caption("**Red bars** = Negative % (left axis) · **Red dashed line** = Negative % trend · **Dark line** = Average Score (right axis)")
        else:
            st.info("Need at least 2 months of data to show trend.")
